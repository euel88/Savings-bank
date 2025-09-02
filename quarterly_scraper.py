"""
저축은행 중앙회 통일경영공시 데이터 자동 스크래핑 도구 (MD 기능 포함)
목적: 79개 저축은행의 재무정보를 빠르고 효율적으로 스크래핑 (Excel + Markdown)
작성일: 2025-03-31
버전: 2.5 (당기/전년동기 비교 및 MD 통합 지원)
개선사항:
- 당기와 전년동기 데이터 비교 분석 기능 추가 (v2.5)
- 증감률 계산 및 표시 기능 추가 (v2.5)
- 재무데이터 통합 시 엑셀과 MD 파일 동시 생성 (v2.5)
- MD 파일 생성 기능 추가 (v2.4)
- 요약 보고서 MD 형식 지원 (v2.4)
- 통합 재무 보고서 MD 형식 지원 (v2.4)
- 재무데이터 통합 대화상자 버튼 표시 문제 수정 (v2.3)
- "친애" 은행을 "JT친애"로 수정 (v2.3)
- 날짜 추출 로직 개선 - 당기 데이터 우선 추출 (v2.3)
- 파일 경로 오류 수정 및 경로 처리 강화 (v2.2)
- 은행명 정확한 매칭 기능 강화 (v2.1)
- 재무 데이터 통합 보고서 기능 추가 (v2.1)
- 재무 데이터 통합 시 폴더/파일 선택 기능 추가 (v2.1)
- 자기자본/총자산 구분 추출 개선 (v2.1)
"""

import os
import sys
import time
import random
import json
import re
import concurrent.futures
import zipfile
from datetime import datetime
from io import StringIO
import io
from contextlib import contextmanager
from pathlib import Path
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from tqdm import tqdm
import selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
from bs4 import BeautifulSoup
import pandas as pd
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)  # 웹드라이버 관련 경고 무시
warnings.filterwarnings("ignore", category=UserWarning)  # 기타 경고 무시

# 표준 에러 출력을 억제하는 컨텍스트 매니저
@contextmanager
def suppress_stderr():
    """표준 에러 출력을 임시로 억제합니다."""
    original_stderr = sys.stderr
    sys.stderr = io.StringIO()
    try:
        yield
    finally:
        sys.stderr = original_stderr

# 상수 및 기본 설정
class Config:
    """프로그램 설정을 관리하는 클래스"""
    # 기본 설정 값
    VERSION = "2.5"
    BASE_URL = "https://www.fsb.or.kr/busmagequar_0100.act"  # 통일경영공시 URL
    MAX_RETRIES = 2  # 재시도 횟수
    PAGE_LOAD_TIMEOUT = 8  # 페이지 로드 타임아웃
    WAIT_TIMEOUT = 4  # 대기 시간
    MAX_WORKERS = 3  # 병렬 처리 워커 수 (기본값 감소)
    
    # 전체 79개 저축은행 목록 (친애 → JT친애로 수정)
    BANKS = [
        "다올", "대신", "더케이", "민국", "바로", "스카이", "신한", "애큐온", "예가람", "웰컴",
        "유안타", "조은", "키움YES", "푸른", "하나", "DB", "HB", "JT", "JT친애", "KB",  # "친애" → "JT친애"로 변경
        "NH", "OK", "OSB", "SBI", "금화", "남양", "모아", "부림", "삼정", "상상인",
        "세람", "안국", "안양", "영진", "융창", "인성", "인천", "키움", "페퍼", "평택",
        "한국투자", "한화", "고려", "국제", "동원제일", "솔브레인", "에스앤티", "우리", "조흥", "진주",
        "흥국", "BNK", "DH", "IBK", "대백", "대아", "대원", "드림", "라온", "머스트삼일",
        "엠에스", "오성", "유니온", "참", "CK", "대한", "더블", "동양", "삼호",
        "센트럴", "스마트", "스타", "대명", "상상인플러스", "아산", "오투", "우리금융", "청주", "한성"
    ]
    
    # 카테고리 목록
    CATEGORIES = ["영업개황", "재무현황", "손익현황", "기타"]
    
    def __init__(self):
        """Config 초기화 - 경로 처리 개선"""
        self.today = datetime.now().strftime("%Y%m%d")
        
        # 설정 파일 경로 (먼저 설정)
        self.config_dir = os.path.join(os.path.expanduser("~"), ".bank_scraper")
        self.config_file = os.path.join(self.config_dir, "settings.json")
        
        # 기본값 설정
        self.chrome_driver_path = None
        self.auto_zip = True
        
        # 기본 출력 디렉토리 설정
        self.output_dir = self._get_default_output_dir()
        
        # 설정 로드 (있으면 덮어쓰기)
        self.load_settings()
        
        # 경로 정규화 및 검증
        self._validate_and_create_paths()
    
    def _get_default_output_dir(self):
        """기본 출력 디렉토리 경로 생성"""
        # 사용자 홈 디렉토리 사용
        base_dir = os.path.expanduser("~")
        # Downloads 폴더가 있으면 사용, 없으면 홈 디렉토리 사용
        downloads_dir = os.path.join(base_dir, "Downloads")
        if os.path.exists(downloads_dir):
            base_dir = downloads_dir
        
        # 오늘 날짜로 폴더명 생성
        return os.path.join(base_dir, f"저축은행_데이터_{self.today}")
    
    def _validate_and_create_paths(self):
        """모든 경로 검증 및 생성"""
        # 경로 정규화 (슬래시 통일)
        self.output_dir = os.path.normpath(self.output_dir)
        
        # 출력 디렉토리 생성
        try:
            os.makedirs(self.output_dir, exist_ok=True)
            print(f"✅ 출력 디렉토리 준비 완료: {self.output_dir}")
        except Exception as e:
            # 권한 문제 등으로 실패하면 임시 디렉토리 사용
            import tempfile
            self.output_dir = tempfile.mkdtemp(prefix=f"저축은행_데이터_{self.today}_")
            print(f"⚠️ 기본 경로 생성 실패, 임시 경로 사용: {self.output_dir}")
        
        # 파일 경로 설정 (정규화된 경로 사용)
        self.progress_file = os.path.join(self.output_dir, 'bank_scraping_progress.json')
        self.log_file = os.path.join(self.output_dir, f'bank_scraping_log_{self.today}.txt')
    
    def update_output_dir(self, new_dir):
        """출력 디렉토리를 업데이트합니다."""
        # 경로 정규화
        new_dir = os.path.normpath(new_dir)
        
        # 디렉토리 존재 확인 및 생성
        try:
            os.makedirs(new_dir, exist_ok=True)
            self.output_dir = new_dir
            self.progress_file = os.path.join(self.output_dir, 'bank_scraping_progress.json')
            self.log_file = os.path.join(self.output_dir, f'bank_scraping_log_{self.today}.txt')
            self.save_settings()
            print(f"✅ 출력 디렉토리 변경 완료: {self.output_dir}")
        except Exception as e:
            print(f"❌ 출력 디렉토리 변경 실패: {str(e)}")
            raise
    
    def update_chrome_driver_path(self, new_path):
        """ChromeDriver 경로를 업데이트합니다."""
        if new_path:
            new_path = os.path.normpath(new_path)
        self.chrome_driver_path = new_path
        self.save_settings()
    
    def update_auto_zip(self, auto_zip):
        """자동 압축 옵션을 업데이트합니다."""
        self.auto_zip = auto_zip
        self.save_settings()
    
    def load_settings(self):
        """설정 파일에서 설정을 로드합니다."""
        if not os.path.exists(self.config_file):
            return
        
        try:
            with open(self.config_file, 'r', encoding='utf-8') as f:
                settings = json.load(f)
                
                # output_dir 처리 - 날짜 확인
                if 'output_dir' in settings:
                    saved_dir = settings['output_dir']
                    # 저장된 경로에 오늘 날짜가 포함되어 있는지 확인
                    if self.today in os.path.basename(saved_dir):
                        # 오늘 날짜가 포함된 경로면 그대로 사용
                        self.output_dir = os.path.normpath(saved_dir)
                    else:
                        # 날짜가 다르면 새 디렉토리 생성
                        print(f"📅 날짜가 변경되어 새 출력 폴더를 생성합니다.")
                        self.output_dir = self._get_default_output_dir()
                
                # 기타 설정 로드
                if 'chrome_driver_path' in settings:
                    path = settings['chrome_driver_path']
                    if path and os.path.exists(path):
                        self.chrome_driver_path = os.path.normpath(path)
                
                if 'max_workers' in settings:
                    self.MAX_WORKERS = int(settings['max_workers'])
                    
                if 'auto_zip' in settings:
                    self.auto_zip = bool(settings['auto_zip'])
                    
        except Exception as e:
            print(f"⚠️ 설정 파일 로드 중 오류 (기본값 사용): {str(e)}")
    
    def save_settings(self):
        """설정을 파일에 저장합니다."""
        try:
            os.makedirs(self.config_dir, exist_ok=True)
            
            settings = {
                'output_dir': self.output_dir,
                'chrome_driver_path': self.chrome_driver_path,
                'max_workers': self.MAX_WORKERS,
                'auto_zip': self.auto_zip,
                'last_saved': datetime.now().isoformat()
            }
            
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(settings, f, ensure_ascii=False, indent=2)
                
        except Exception as e:
            print(f"⚠️ 설정 저장 실패 (프로그램은 계속 실행됩니다): {str(e)}")


# 로그 관리 클래스
class Logger:
    def __init__(self, config, gui=None):
        """Logger 초기화 - 에러 처리 강화"""
        self.config = config
        self.gui = gui
        self.fallback_log_file = None  # 대체 로그 파일 경로
        
        # 로그 파일 초기화
        self._initialize_log_file()
    
    def _initialize_log_file(self):
        """로그 파일 초기화 - 실패 시 대체 경로 사용"""
        try:
            # 디렉토리 확인 및 생성
            log_dir = os.path.dirname(self.config.log_file)
            if log_dir and not os.path.exists(log_dir):
                os.makedirs(log_dir, exist_ok=True)
            
            # 로그 파일 생성/초기화
            with open(self.config.log_file, 'w', encoding='utf-8') as f:
                f.write(f"=== 저축은행 스크래핑 로그 시작 ===\n")
                f.write(f"시작 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"출력 폴더: {self.config.output_dir}\n")
                f.write(f"=" * 50 + "\n\n")
            
            print(f"✅ 로그 파일 초기화 완료: {self.config.log_file}")
            
        except Exception as e:
            # 기본 경로 실패 시 대체 경로 사용
            print(f"⚠️ 기본 로그 파일 생성 실패: {str(e)}")
            
            # 대체 경로 시도 (현재 작업 디렉토리)
            self.fallback_log_file = os.path.join(
                os.getcwd(), 
                f'bank_scraping_log_{self.config.today}.txt'
            )
            
            try:
                with open(self.fallback_log_file, 'w', encoding='utf-8') as f:
                    f.write(f"=== 저축은행 스크래핑 로그 (대체 경로) ===\n")
                    f.write(f"시작 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write(f"원래 경로: {self.config.log_file}\n")
                    f.write(f"대체 경로: {self.fallback_log_file}\n")
                    f.write(f"=" * 50 + "\n\n")
                
                print(f"✅ 대체 로그 파일 사용: {self.fallback_log_file}")
                
            except Exception as e2:
                # 모든 파일 쓰기 실패 시
                print(f"❌ 로그 파일 생성 완전 실패: {str(e2)}")
                print("⚠️ 로그가 콘솔에만 출력됩니다.")
    
    def log_message(self, message, print_to_console=True, verbose=True):
        """로그 메시지를 파일에 기록하고 필요한 경우 콘솔에 출력합니다."""
        if not verbose:
            return
        
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_entry = f"[{timestamp}] {message}"
        
        # 로그 파일에 기록 시도
        log_files = [self.config.log_file]
        if self.fallback_log_file:
            log_files.append(self.fallback_log_file)
        
        file_written = False
        for log_file in log_files:
            try:
                # 파일이 있는 디렉토리 확인
                log_dir = os.path.dirname(log_file)
                if log_dir and not os.path.exists(log_dir):
                    os.makedirs(log_dir, exist_ok=True)
                
                # 로그 추가
                with open(log_file, 'a', encoding='utf-8') as f:
                    f.write(log_entry + '\n')
                    f.flush()  # 즉시 디스크에 쓰기
                
                file_written = True
                break  # 성공하면 중단
                
            except Exception as e:
                continue  # 다음 파일 시도
        
        # GUI에 출력
        if self.gui:
            try:
                self.gui.update_log(message)
            except:
                pass  # GUI 업데이트 실패는 무시
        
        # 콘솔에 출력
        if print_to_console:
            if file_written:
                print(f"📝 {message}")
            else:
                print(f"⚠️ [로그 파일 쓰기 실패] {message}")
    
    def get_log_location(self):
        """현재 사용 중인 로그 파일 경로 반환"""
        if self.fallback_log_file and os.path.exists(self.fallback_log_file):
            return self.fallback_log_file
        return self.config.log_file


# 웹드라이버 관리 클래스
class DriverManager:
    def __init__(self, config, logger):
        self.config = config
        self.logger = logger
        self.max_drivers = config.MAX_WORKERS
        self.drivers = []
        self.available_drivers = []
        
    def initialize_drivers(self):
        """드라이버 풀 초기화"""
        self.logger.log_message(f"{self.max_drivers}개의 드라이버 초기화 중...")
        
        # 환경 변수 설정 (경고 메시지 억제)
        os.environ['WDM_LOG_LEVEL'] = '0'  # 로깅 레벨 최소화

        # stdout 및 stderr 임시 리다이렉션
        original_stdout = sys.stdout
        original_stderr = sys.stderr
        sys.stdout = io.StringIO()
        sys.stderr = io.StringIO()
        
        try:
            # 첫 번째 드라이버 생성 시도
            try:
                first_driver = self.create_driver()
                self.drivers.append(first_driver)
                self.available_drivers.append(first_driver)
                
                # 나머지 드라이버 생성
                for _ in range(self.max_drivers - 1):
                    driver = self.create_driver()
                    self.drivers.append(driver)
                    self.available_drivers.append(driver)
                    
            except Exception as e:
                self.logger.log_message(f"드라이버 초기화 중 오류 발생: {str(e)}")
                # 이미 생성된 드라이버가 있다면 모두 종료
                self.close_all()
                raise
        finally:
            # 원래 stdout 및 stderr 복원
            sys.stdout = original_stdout
            sys.stderr = original_stderr
    
    def create_driver(self):
        """최적화된 Chrome 웹드라이버를 생성합니다."""
        # stderr 출력 억제 (ChromeDriver 경고 메시지 숨기기)
        with suppress_stderr():
            options = webdriver.ChromeOptions()
            options.add_argument('--headless=new')
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('--disable-gpu')
            options.add_argument('--window-size=1280,800')
            options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36')
            
            # 로그 레벨 설정 (경고 숨기기)
            options.add_argument('--log-level=3')  # 오류만 표시
            options.add_argument('--silent')       # 메시지 억제
            
            # 최적화 옵션
            options.add_argument('--disable-extensions')
            options.add_argument('--disable-browser-side-navigation')
            options.add_argument('--disable-infobars')
            options.add_argument('--disable-notifications')
            options.add_argument('--disable-popup-blocking')
            
            # 이미지 로딩 활성화 (필요시)
            prefs = {
                'profile.default_content_setting_values': {
                    'images': 1,      # 이미지 로딩 활성화 (1=허용)
                    'plugins': 2,     # 플러그인 차단
                    'javascript': 1,  # JavaScript 허용 (필요)
                    'notifications': 2  # 알림 차단
                },
                'disk-cache-size': 4096,
            }
            options.add_experimental_option('prefs', prefs)
            
            # 브라우저 자동 종료 방지를 위한 경고 무시
            options.add_experimental_option("detach", True)
            options.add_experimental_option('excludeSwitches', ['enable-logging'])  # 콘솔 로깅 비활성화
            
            try:
                # 수동으로 지정된 ChromeDriver 경로 사용
                if self.config.chrome_driver_path and os.path.exists(self.config.chrome_driver_path):
                    from selenium.webdriver.chrome.service import Service
                    service = Service(executable_path=self.config.chrome_driver_path)
                    service.log_path = os.devnull  # 로그 비활성화
                    driver = webdriver.Chrome(service=service, options=options)
                    self.logger.log_message(f"지정된 ChromeDriver 사용: {self.config.chrome_driver_path}", verbose=False)
                else:
                    # ChromeDriver 자동 다운로드 오류 방지
                    from selenium.webdriver.chrome.service import Service
                    from webdriver_manager.chrome import ChromeDriverManager
                    
                    # 캐시 모드 설정 (오프라인 사용)
                    os.environ['WDM_LOCAL_CACHE_DIR'] = os.path.join(os.path.expanduser("~"), ".wdm", "drivers")
                    os.environ['WDM_OFFLINE'] = "true"  # 오프라인 모드로 설정
                    os.environ['WDM_LOG_LEVEL'] = '0'   # 로깅 레벨 최소화
                    
                    # 캐시된 드라이버 사용
                    service = Service(ChromeDriverManager().install())
                    
                    # 서비스 로그 수준 설정
                    service.log_path = os.devnull  # 로그 출력을 /dev/null로 리다이렉션
                    
                    driver = webdriver.Chrome(service=service, options=options)
            except Exception as e:
                self.logger.log_message(f"ChromeDriver 자동 설치 실패, 기본 방식으로 시도: {str(e)}", verbose=False)
                # 자동 설치 실패 시 기본 방식으로 시도
                driver = webdriver.Chrome(options=options)
            
            driver.set_page_load_timeout(self.config.PAGE_LOAD_TIMEOUT)
            return driver
    
    def get_driver(self):
        """사용 가능한 드라이버를 가져옵니다."""
        if not self.available_drivers:
            self.logger.log_message("모든 드라이버가 사용 중입니다. 대기 중...", verbose=False)
            time.sleep(1)  # 잠시 대기 후 재시도
            return self.get_driver()
        
        driver = self.available_drivers.pop(0)
        return driver
    
    def return_driver(self, driver):
        """드라이버를 풀에 반환합니다."""
        if driver in self.drivers and driver not in self.available_drivers:
            try:
                # 드라이버 상태 확인
                driver.current_url  # 접근 가능한지 확인
                self.available_drivers.append(driver)
            except:
                # 오류 발생 시 해당 드라이버를 종료하고 새 드라이버 생성
                try:
                    driver.quit()
                except:
                    pass
                
                self.drivers.remove(driver)
                new_driver = self.create_driver()
                self.drivers.append(new_driver)
                self.available_drivers.append(new_driver)
    
    def close_all(self):
        """모든 드라이버를 종료합니다."""
        for driver in self.drivers:
            try:
                driver.quit()
            except:
                pass
        self.drivers = []
        self.available_drivers = []


# 진행 상황 관리 클래스
class ProgressManager:
    def __init__(self, config, logger):
        self.config = config
        self.logger = logger
        self.file_path = config.progress_file
        self.progress = self.load()
    
    def load(self):
        """저장된 진행 상황을 로드합니다."""
        if os.path.exists(self.file_path):
            try:
                with open(self.file_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except json.JSONDecodeError:
                self.logger.log_message(f"진행 파일 손상: {self.file_path}, 새로 생성합니다.")
            except Exception as e:
                self.logger.log_message(f"진행 파일 로드 실패: {str(e)}")
        
        return {
            'completed': [],
            'failed': [],
            'stats': {
                'last_run': None,
                'success_count': 0,
                'failure_count': 0
            }
        }
    
    def is_completed(self, bank_name):
        """특정 은행의 스크래핑이 완료되었는지 확인합니다."""
        return bank_name in self.progress.get('completed', [])
    
    def mark_completed(self, bank_name):
        """은행을 완료 목록에 추가합니다."""
        if bank_name not in self.progress.get('completed', []):
            self.progress.setdefault('completed', []).append(bank_name)
            self.progress['stats']['success_count'] = len(self.progress.get('completed', []))
        
        # 실패 목록에서 제거 (재시도 후 성공한 경우)
        if bank_name in self.progress.get('failed', []):
            self.progress['failed'].remove(bank_name)
        
        self.save()
    
    def mark_failed(self, bank_name):
        """은행을 실패 목록에 추가합니다."""
        if bank_name not in self.progress.get('failed', []) and bank_name not in self.progress.get('completed', []):
            self.progress.setdefault('failed', []).append(bank_name)
            self.progress['stats']['failure_count'] = len(self.progress.get('failed', []))
            self.save()
    
    def save(self):
        """진행 상황을 파일에 저장합니다."""
        try:
            self.progress['stats']['last_run'] = datetime.now().isoformat()
            # 디렉토리 확인
            progress_dir = os.path.dirname(self.file_path)
            if progress_dir and not os.path.exists(progress_dir):
                os.makedirs(progress_dir, exist_ok=True)
            
            with open(self.file_path, 'w', encoding='utf-8') as f:
                json.dump(self.progress, f, ensure_ascii=False, indent=2)
        except Exception as e:
            self.logger.log_message(f"진행 상황 저장 실패: {str(e)}")
    
    def get_pending_banks(self, all_banks=None):
        """처리할 은행 목록을 반환합니다."""
        if all_banks is None:
            all_banks = self.config.BANKS
        completed = set(self.progress.get('completed', []))
        return [bank for bank in all_banks if bank not in completed]
    
    def reset_progress(self):
        """진행 상황을 초기화합니다."""
        self.progress = {
            'completed': [],
            'failed': [],
            'stats': {
                'last_run': None,
                'success_count': 0,
                'failure_count': 0
            }
        }
        self.save()


# 명시적 웨이팅 유틸리티 클래스
class WaitUtils:
    @staticmethod
    def wait_for_element(driver, locator, timeout):
        """요소가 나타날 때까지 명시적으로 대기합니다."""
        try:
            element = WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located(locator)
            )
            return element
        except TimeoutException:
            return None
    
    @staticmethod
    def wait_for_clickable(driver, locator, timeout):
        """요소가 클릭 가능할 때까지 명시적으로 대기합니다."""
        try:
            element = WebDriverWait(driver, timeout).until(
                EC.element_to_be_clickable(locator)
            )
            return element
        except TimeoutException:
            return None
    
    @staticmethod
    def wait_for_page_load(driver, timeout):
        """페이지가 완전히 로드될 때까지 대기합니다."""
        try:
            WebDriverWait(driver, timeout).until(
                lambda d: d.execute_script('return document.readyState') == 'complete'
            )
            return True
        except TimeoutException:
            return False
    
    @staticmethod
    def wait_with_random(min_time=0.3, max_time=0.7):
        """무작위 시간 동안 대기합니다."""
        time.sleep(random.uniform(min_time, max_time))


# 재무 데이터 소스 선택 대화상자 (수정된 버전)
class FinancialDataSourceDialog:
    """재무 데이터 소스 선택 대화상자"""
    
    def __init__(self, parent, config):
        self.parent = parent
        self.config = config
        self.result = None
        
        # 대화상자 창 생성
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("재무 데이터 소스 선택")
        # 창 크기를 늘려서 버튼이 보이도록 수정
        self.dialog.geometry("500x480")  # 높이를 480으로 증가
        self.dialog.resizable(False, False)
        
        # 모달 대화상자로 설정
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # 위젯 생성
        self.create_widgets()
        
        # 창 가운데 정렬
        self.center_window()
        
        # 키보드 단축키 바인딩
        self.dialog.bind('<Return>', lambda e: self.on_ok())
        self.dialog.bind('<Escape>', lambda e: self.on_cancel())
        
        # 창 크기 자동 조정
        self.dialog.update_idletasks()
        self.dialog.minsize(self.dialog.winfo_reqwidth(), self.dialog.winfo_reqheight())
    
    def create_widgets(self):
        """대화상자 위젯 생성"""
        # 메인 프레임
        main_frame = ttk.Frame(self.dialog, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 설명 라벨
        ttk.Label(main_frame, text="재무 데이터를 추출할 소스를 선택하세요:", 
                  font=("", 10, "bold")).pack(pady=(0, 15))
        
        # 라디오 버튼 변수
        self.source_var = tk.StringVar(value="default_folder")
        
        # 옵션 1: 기본 출력 폴더
        option1_frame = ttk.LabelFrame(main_frame, text="옵션 1: 기본 출력 폴더", padding="10")
        option1_frame.pack(fill=tk.X, pady=5)
        
        ttk.Radiobutton(option1_frame, text="현재 설정된 출력 폴더의 모든 엑셀 파일", 
                       variable=self.source_var, value="default_folder").pack(anchor=tk.W)
        ttk.Label(option1_frame, text=f"경로: {self.config.output_dir}", 
                 font=("", 9), foreground="gray").pack(anchor=tk.W, padx=20)
        
        # 옵션 2: 다른 폴더 선택
        option2_frame = ttk.LabelFrame(main_frame, text="옵션 2: 다른 폴더 선택", padding="10")
        option2_frame.pack(fill=tk.X, pady=5)
        
        ttk.Radiobutton(option2_frame, text="다른 폴더에서 엑셀 파일 읽기", 
                       variable=self.source_var, value="custom_folder").pack(anchor=tk.W)
        
        folder_frame = ttk.Frame(option2_frame)
        folder_frame.pack(fill=tk.X, padx=20, pady=5)
        
        self.folder_path_var = tk.StringVar()
        ttk.Entry(folder_frame, textvariable=self.folder_path_var, width=40).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(folder_frame, text="폴더 선택", command=self.browse_folder).pack(side=tk.LEFT)
        
        # 옵션 3: 개별 파일 선택
        option3_frame = ttk.LabelFrame(main_frame, text="옵션 3: 개별 파일 선택", padding="10")
        option3_frame.pack(fill=tk.X, pady=5)
        
        ttk.Radiobutton(option3_frame, text="특정 엑셀 파일들만 선택", 
                       variable=self.source_var, value="selected_files").pack(anchor=tk.W)
        
        files_frame = ttk.Frame(option3_frame)
        files_frame.pack(fill=tk.X, padx=20, pady=5)
        
        self.selected_files = []
        self.files_label = ttk.Label(files_frame, text="선택된 파일: 0개", foreground="gray")
        self.files_label.pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(files_frame, text="파일 선택", command=self.browse_files).pack(side=tk.LEFT)
        
        # 선택된 파일 목록 표시 (스크롤 가능)
        list_frame = ttk.Frame(option3_frame)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=20)
        
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.files_listbox = tk.Listbox(list_frame, height=3, yscrollcommand=scrollbar.set)
        self.files_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.files_listbox.yview)
        
        # 버튼 프레임 - 중앙 정렬로 변경
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, side=tk.BOTTOM, pady=(20, 10))
        
        # 버튼을 중앙에 배치
        button_container = ttk.Frame(button_frame)
        button_container.pack(expand=True)
        
        ttk.Button(button_container, text="확인", command=self.on_ok, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_container, text="취소", command=self.on_cancel, width=10).pack(side=tk.LEFT, padx=5)
    
    def browse_folder(self):
        """폴더 선택 대화상자"""
        folder = filedialog.askdirectory(
            title="엑셀 파일이 있는 폴더 선택",
            initialdir=self.config.output_dir
        )
        if folder:
            self.folder_path_var.set(folder)
            self.source_var.set("custom_folder")
    
    def browse_files(self):
        """파일 선택 대화상자"""
        files = filedialog.askopenfilenames(
            title="처리할 엑셀 파일 선택",
            initialdir=self.config.output_dir,
            filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")]
        )
        if files:
            self.selected_files = list(files)
            self.files_listbox.delete(0, tk.END)
            for file in self.selected_files:
                self.files_listbox.insert(tk.END, os.path.basename(file))
            self.files_label.config(text=f"선택된 파일: {len(self.selected_files)}개")
            self.source_var.set("selected_files")
    
    def on_ok(self):
        """확인 버튼 클릭"""
        source_type = self.source_var.get()
        
        if source_type == "default_folder":
            self.result = ("default_folder", self.config.output_dir)
            
        elif source_type == "custom_folder":
            folder_path = self.folder_path_var.get()
            if not folder_path:
                messagebox.showwarning("경고", "폴더를 선택해주세요.")
                return
            if not os.path.exists(folder_path):
                messagebox.showerror("오류", "선택한 폴더가 존재하지 않습니다.")
                return
            self.result = ("custom_folder", folder_path)
            
        elif source_type == "selected_files":
            if not self.selected_files:
                messagebox.showwarning("경고", "파일을 선택해주세요.")
                return
            self.result = ("selected_files", self.selected_files)
        
        self.dialog.destroy()
    
    def on_cancel(self):
        """취소 버튼 클릭"""
        self.result = None
        self.dialog.destroy()
    
    def center_window(self):
        """창을 화면 중앙에 배치"""
        self.dialog.update_idletasks()
        
        # 화면 크기 확인
        screen_width = self.dialog.winfo_screenwidth()
        screen_height = self.dialog.winfo_screenheight()
        
        # 창 크기 확인
        window_width = self.dialog.winfo_width()
        window_height = self.dialog.winfo_height()
        
        # 창이 화면보다 크면 조정
        if window_height > screen_height - 100:
            self.dialog.geometry(f"500x{screen_height - 100}")
            self.dialog.update_idletasks()
            window_height = self.dialog.winfo_height()
        
        x = (screen_width // 2) - (window_width // 2)
        y = (screen_height // 2) - (window_height // 2)
        self.dialog.geometry(f"+{x}+{y}")


# 스크래퍼 클래스
class BankScraper:
    def __init__(self, config, logger, driver_manager, progress_manager):
        self.config = config
        self.logger = logger
        self.driver_manager = driver_manager
        self.progress_manager = progress_manager
    
    def extract_date_information(self, driver):
        """웹페이지에서 공시 날짜 정보를 추출합니다. (개선된 버전)"""
        try:
            # 방법 1: 당기 데이터 우선 찾기
            current_period_elements = driver.find_elements(
                By.XPATH, 
                "//*[contains(text(), '당기') and contains(text(), '년') and contains(text(), '월')]"
            )
            
            if current_period_elements:
                for element in current_period_elements:
                    text = element.text
                    # 정규식으로 날짜 패턴 추출
                    date_pattern = re.compile(r'\d{4}년\d{1,2}월말?')
                    matches = date_pattern.findall(text)
                    
                    if matches:
                        # 가장 최근 연도 찾기
                        latest_date = max(matches, key=lambda x: int(x[:4]))
                        self.logger.log_message(f"당기 날짜 발견: {latest_date}", verbose=False)
                        return latest_date
            
            # 방법 2: 모든 날짜를 찾아서 가장 최근 것 선택
            all_date_elements = driver.find_elements(
                By.XPATH, 
                "//*[contains(text(), '년') and contains(text(), '월')]"
            )
            
            all_dates = []
            for element in all_date_elements:
                text = element.text
                # 정규식 패턴 개선 (월말이 없는 경우도 포함)
                date_pattern = re.compile(r'\d{4}년\d{1,2}월말?')
                matches = date_pattern.findall(text)
                all_dates.extend(matches)
            
            if all_dates:
                # 중복 제거 및 정렬
                unique_dates = list(set(all_dates))
                # 연도 기준으로 정렬하여 가장 최근 날짜 선택
                sorted_dates = sorted(unique_dates, key=lambda x: int(x[:4]), reverse=True)
                
                # 2025년 데이터가 있으면 우선 선택
                for date in sorted_dates:
                    if "2025년" in date:
                        self.logger.log_message(f"최신 날짜 선택: {date}", verbose=False)
                        return date
                
                # 2025년이 없으면 가장 최근 날짜 반환
                return sorted_dates[0]
            
            # 방법 3: JavaScript로 직접 추출 (더 정확함)
            js_script = """
            var dates = [];
            var allText = document.body.innerText;
            
            // 당기 관련 텍스트 우선 찾기
            var currentPeriodMatch = allText.match(/당기[^\\n]*?(\\d{4}년\\d{1,2}월말?)/);
            if (currentPeriodMatch) {
                return currentPeriodMatch[1];
            }
            
            // 모든 날짜 찾아서 최신 것 반환
            var allMatches = allText.match(/\\d{4}년\\d{1,2}월말?/g);
            if (allMatches) {
                // 연도별로 정렬
                allMatches.sort(function(a, b) {
                    return parseInt(b.substr(0, 4)) - parseInt(a.substr(0, 4));
                });
                
                // 2025년 우선
                for (var i = 0; i < allMatches.length; i++) {
                    if (allMatches[i].includes('2025년')) {
                        return allMatches[i];
                    }
                }
                
                return allMatches[0];
            }
            
            return '';
            """
            
            date_text = driver.execute_script(js_script)
            if date_text:
                self.logger.log_message(f"JavaScript로 추출한 날짜: {date_text}", verbose=False)
                return date_text
            
            return "날짜 정보 없음"
            
        except Exception as e:
            self.logger.log_message(f"날짜 정보 추출 오류: {str(e)}", verbose=False)
            return "날짜 추출 실패"
    
    def select_bank(self, driver, bank_name):
        """다양한 방법으로 은행을 선택합니다. (정확한 매칭 우선)"""
        try:
            # 메인 페이지로 접속
            driver.get(self.config.BASE_URL)
            
            # 페이지 로딩 완료 대기
            WaitUtils.wait_for_page_load(driver, self.config.PAGE_LOAD_TIMEOUT)
            WaitUtils.wait_with_random(0.5, 1)
            
            # 특수 케이스 처리를 위한 정확한 은행명 목록 (JT친애 추가)
            exact_bank_names = {
                "키움": ["키움", "키움저축은행"],
                "키움YES": ["키움YES", "키움YES저축은행"],
                "JT": ["JT", "JT저축은행"],
                "JT친애": ["JT친애", "JT친애저축은행", "친애", "친애저축은행"],  # JT친애 매핑 추가
                "상상인": ["상상인", "상상인저축은행"],
                "상상인플러스": ["상상인플러스", "상상인플러스저축은행"],
                "머스트삼일": ["머스트삼일", "머스트삼일저축은행"]
            }
            
            # 검색할 은행명 목록 결정
            search_names = exact_bank_names.get(bank_name, [bank_name, f"{bank_name}저축은행"])
            
            # 방법 1: JavaScript로 정확한 은행명 매칭 (개선된 버전)
            js_script = f"""
            var targetBankNames = {json.dumps(search_names)};
            var found = false;
            
            // 모든 테이블 셀과 링크를 검사
            var allElements = document.querySelectorAll('td, a');
            
            for(var i = 0; i < allElements.length; i++) {{
                var element = allElements[i];
                var elementText = element.textContent.trim();
                
                // 정확한 매칭 확인
                for(var j = 0; j < targetBankNames.length; j++) {{
                    if(elementText === targetBankNames[j]) {{
                        // 키움/키움YES, JT/JT친애 구분을 위한 추가 검증
                        if('{bank_name}' === '키움' && elementText.includes('YES')) {{
                            continue;  // 키움을 찾는데 키움YES가 나오면 건너뛰기
                        }}
                        if('{bank_name}' === 'JT' && elementText.includes('친애')) {{
                            continue;  // JT를 찾는데 JT친애가 나오면 건너뛰기
                        }}
                        
                        element.scrollIntoView({{block: 'center'}});
                        
                        // 링크가 있으면 링크 클릭, 없으면 셀 클릭
                        if(element.tagName === 'A') {{
                            element.click();
                            found = true;
                            break;
                        }} else {{
                            var link = element.querySelector('a');
                            if(link) {{
                                link.click();
                                found = true;
                                break;
                            }} else {{
                                element.click();
                                found = true;
                                break;
                            }}
                        }}
                    }}
                }}
                if(found) break;
            }}
            
            return found ? "정확한 매칭 성공" : false;
            """
            
            result = driver.execute_script(js_script)
            if result:
                self.logger.log_message(f"{bank_name} 은행: {result}", verbose=False)
                WaitUtils.wait_with_random(1, 1.5)  # 페이지 전환 대기 시간 증가
                
                # 페이지 전환 확인
                current_url = driver.current_url
                if current_url != self.config.BASE_URL:
                    return True
            
            # 방법 2: XPath로 정확한 텍스트 매칭 (보완)
            for search_name in search_names:
                # 추가 조건으로 더 정확한 매칭
                if bank_name == "키움":
                    xpath = f"//td[normalize-space(text())='{search_name}' and not(contains(text(), 'YES'))]"
                elif bank_name == "JT":
                    xpath = f"//td[normalize-space(text())='{search_name}' and not(contains(text(), '친애'))]"
                else:
                    xpath = f"//td[normalize-space(text())='{search_name}']"
                
                bank_elements = driver.find_elements(By.XPATH, xpath)
                
                if bank_elements:
                    for element in bank_elements:
                        try:
                            if element.is_displayed():
                                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
                                WaitUtils.wait_with_random(0.3, 0.5)
                                driver.execute_script("arguments[0].click();", element)
                                WaitUtils.wait_with_random(1, 1.5)
                                
                                # 페이지 전환 확인
                                if driver.current_url != self.config.BASE_URL:
                                    return True
                        except:
                            continue
            
            self.logger.log_message(f"{bank_name} 은행을 찾을 수 없습니다.")
            return False
            
        except Exception as e:
            self.logger.log_message(f"{bank_name} 은행 선택 실패: {str(e)}")
            return False
    
    def select_category(self, driver, category):
        """특정 카테고리 탭을 클릭합니다."""
        try:
            # 방법 1: 정확한 텍스트 매칭
            tab_xpaths = [
                f"//a[normalize-space(text())='{category}']",
                f"//a[contains(@class, 'tab') and contains(text(), '{category}')]",
                f"//li[contains(@class, 'tab') and contains(text(), '{category}')]",
                f"//span[contains(text(), '{category}')]",
                f"//button[contains(text(), '{category}')]"
            ]
            
            for xpath in tab_xpaths:
                elements = driver.find_elements(By.XPATH, xpath)
                for element in elements:
                    try:
                        if element.is_displayed():
                            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
                            WaitUtils.wait_with_random(0.3, 0.5)
                            driver.execute_script("arguments[0].click();", element)
                            WaitUtils.wait_with_random(0.5, 1)
                            return True
                    except:
                        continue
            
            # 방법 2: JavaScript로 카테고리 탭 클릭
            category_indices = {
                "영업개황": 0,
                "재무현황": 1,
                "손익현황": 2,
                "기타": 3
            }
            
            if category in category_indices:
                idx = category_indices[category]
                script = f"""
                // 모든 탭 관련 요소 찾기
                var tabContainers = document.querySelectorAll('ul.tabs, div.tab-container, nav, .tab-list, ul, div[role="tablist"]');
                
                // 정확한 텍스트 매칭
                var allElements = document.querySelectorAll('a, button, span, li, div');
                for (var k = 0; k < allElements.length; k++) {{
                    if (allElements[k].innerText.trim() === '{category}') {{
                        allElements[k].scrollIntoView({{block: 'center'}});
                        allElements[k].click();
                        return "exact_match";
                    }}
                }}
                
                // 탭 컨테이너에서 인덱스 기반 검색
                for (var i = 0; i < tabContainers.length; i++) {{
                    var tabs = tabContainers[i].querySelectorAll('a, li, button, div[role="tab"], span');
                    
                    // 먼저 텍스트로 찾기
                    for (var j = 0; j < tabs.length; j++) {{
                        if (tabs[j].innerText.includes('{category}')) {{
                            tabs[j].scrollIntoView({{block: 'center'}});
                            tabs[j].click();
                            return "text_match_in_container";
                        }}
                    }}
                    
                    // 인덱스로 찾기
                    if (tabs.length >= {idx + 1}) {{
                        tabs[{idx}].scrollIntoView({{block: 'center'}});
                        tabs[{idx}].click();
                        return "index_match";
                    }}
                }}
                
                // 모든 클릭 가능 요소에서 포함 문자열 검색
                var clickables = document.querySelectorAll('a, button, span, div, li');
                for (var j = 0; j < clickables.length; j++) {{
                    if (clickables[j].innerText.includes('{category}')) {{
                        clickables[j].scrollIntoView({{block: 'center'}});
                        clickables[j].click();
                        return "contains_match";
                    }}
                }}
                
                return false;
                """
                
                result = driver.execute_script(script)
                if result:
                    self.logger.log_message(f"{category} 탭: {result} 성공", verbose=False)
                    WaitUtils.wait_with_random(0.5, 1)
                    return True
            
            # 방법 3: 포함 문자열로 검색 (더 관대한 매칭)
            tab_broad_xpath = f"//*[contains(text(), '{category}')]"
            elements = driver.find_elements(By.XPATH, tab_broad_xpath)
            
            for element in elements:
                try:
                    if element.is_displayed() and (element.tag_name in ['a', 'li', 'span', 'button', 'div']):
                        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
                        WaitUtils.wait_with_random(0.3, 0.5)
                        driver.execute_script("arguments[0].click();", element)
                        WaitUtils.wait_with_random(0.5, 1)
                        return True
                except:
                    continue
            
            # 방법 4: CSS 선택자 시도
            tab_css = f"[role='tab'], .tab, .tab-item, .tabs li, .tabs a, nav a, ul li a"
            tabs = driver.find_elements(By.CSS_SELECTOR, tab_css)
            
            for tab in tabs:
                try:
                    if category in tab.text and tab.is_displayed():
                        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", tab)
                        WaitUtils.wait_with_random(0.3, 0.5)
                        driver.execute_script("arguments[0].click();", tab)
                        WaitUtils.wait_with_random(0.5, 1)
                        return True
                except:
                    continue
            
            self.logger.log_message(f"{category} 탭을 찾을 수 없습니다.", verbose=False)
            return False
            
        except Exception as e:
            self.logger.log_message(f"{category} 탭 클릭 실패: {str(e)}", verbose=False)
            return False
    
    def extract_tables_from_page(self, driver):
        """현재 페이지에서 모든 테이블을 추출합니다."""
        try:
            # 페이지가 완전히 로드될 때까지 대기
            WaitUtils.wait_for_page_load(driver, self.config.PAGE_LOAD_TIMEOUT)
            WaitUtils.wait_with_random(0.5, 1)
            
            # 방법 1: pandas로 테이블 추출
            try:
                html_source = driver.page_source
                dfs = pd.read_html(StringIO(html_source))
                
                if dfs:
                    valid_dfs = []
                    seen_shapes = set()
                    
                    for df in dfs:
                        # 비어있지 않은 테이블만 선택
                        if not df.empty and df.shape[0] > 0 and df.shape[1] > 0:
                            # MultiIndex 컬럼 처리
                            if isinstance(df.columns, pd.MultiIndex):
                                new_cols = []
                                for col in df.columns:
                                    if isinstance(col, tuple):
                                        clean_col = [str(c).strip() for c in col if str(c).strip() and str(c).lower() != 'nan']
                                        new_cols.append('_'.join(clean_col) if clean_col else f"Column_{len(new_cols)+1}")
                                    else:
                                        new_cols.append(str(col))
                                df.columns = new_cols
                            
                            # 중복 테이블 제거
                            try:
                                # 테이블 해시 생성 (중복 확인용)
                                shape_hash = f"{df.shape}"
                                headers_hash = f"{list(df.columns)}"
                                data_hash = ""
                                if len(df) > 0:
                                    # 데이터 첫 행의 값들을 해시 생성
                                    data_hash = f"{list(df.iloc[0].astype(str))}"
                                
                                table_hash = f"{shape_hash}_{headers_hash}_{data_hash}"
                                
                                if table_hash not in seen_shapes:
                                    valid_dfs.append(df)
                                    seen_shapes.add(table_hash)
                            except:
                                # 해시 생성 실패 시 그냥 추가
                                valid_dfs.append(df)
                    
                    return valid_dfs
            except Exception as e:
                self.logger.log_message(f"pandas 테이블 추출 실패: {str(e)}", verbose=False)
            
            # 방법 2: BeautifulSoup으로 테이블 추출 (pandas 실패 시)
            try:
                soup = BeautifulSoup(driver.page_source, 'html.parser')
                tables = soup.find_all('table')
                
                extracted_dfs = []
                table_hashes = set()  # 중복 테이블 제거용
                
                for table in tables:
                    try:
                        # 테이블 구조 파악
                        headers = []
                        rows = []
                        
                        # 헤더 추출
                        th_elements = table.select('thead th') or table.select('tr:first-child th')
                        if th_elements:
                            headers = [th.get_text(strip=True) for th in th_elements]
                        
                        # 헤더가 없으면 첫 번째 행의 td를 헤더로 사용
                        if not headers:
                            first_row_tds = table.select('tr:first-child td')
                            if first_row_tds:
                                headers = [td.get_text(strip=True) or f"Column_{i+1}" for i, td in enumerate(first_row_tds)]
                        
                        # 헤더가 없으면 기본 열 이름 생성
                        if not headers:
                            max_cols = max([len(row.select('td')) for row in table.select('tr')], default=0)
                            headers = [f'Column_{j+1}' for j in range(max_cols)]
                        
                        # 데이터 행 추출
                        header_rows = table.select('thead tr')
                        for tr in table.select('tbody tr') or table.select('tr')[1:]:
                            if tr in header_rows:
                                continue
                            
                            cells = tr.select('td')
                            if cells:
                                row_data = [td.get_text(strip=True) for td in cells]
                                if row_data and len(row_data) > 0:
                                    rows.append(row_data)
                        
                        # 행 데이터가 있으면 DataFrame 생성
                        if rows and headers:
                            # 열 개수 맞추기
                            for i, row in enumerate(rows):
                                if len(row) < len(headers):
                                    rows[i] = row + [''] * (len(headers) - len(row))
                                elif len(row) > len(headers):
                                    rows[i] = row[:len(headers)]
                            
                            df = pd.DataFrame(rows, columns=headers)
                            
                            if not df.empty:
                                # 테이블 해시 생성 (중복 확인용)
                                try:
                                    table_hash = f"{df.shape}_{hash(str(df.iloc[0].values) if len(df) > 0 else '')}"
                                    if table_hash not in table_hashes:
                                        extracted_dfs.append(df)
                                        table_hashes.add(table_hash)
                                except:
                                    extracted_dfs.append(df)
                    except Exception as e:
                        self.logger.log_message(f"개별 테이블 파싱 실패: {str(e)}", verbose=False)
                
                if extracted_dfs:
                    return extracted_dfs
            except Exception as e:
                self.logger.log_message(f"BeautifulSoup 테이블 추출 실패: {str(e)}", verbose=False)
            
            return []
            
        except Exception as e:
            self.logger.log_message(f"테이블 추출 실패: {str(e)}", verbose=False)
            return []
    
    def scrape_bank_data(self, bank_name, driver):
        """단일 은행의 데이터를 스크래핑합니다."""
        self.logger.log_message(f"[시작] {bank_name} 은행 스크래핑 시작")
        
        try:
            # 은행 선택
            if not self.select_bank(driver, bank_name):
                self.logger.log_message(f"{bank_name} 은행 선택 실패")
                return None
            
            # 현재 페이지 URL 저장
            try:
                base_bank_url = driver.current_url
                self.logger.log_message(f"{bank_name} 은행 페이지 접속 성공", verbose=False)
            except:
                self.logger.log_message(f"{bank_name} 은행 페이지 URL 획득 실패")
                return None
            
            # 날짜 정보 추출
            date_info = self.extract_date_information(driver)
            self.logger.log_message(f"{bank_name} 은행 날짜 정보: {date_info}", verbose=True)
            
            result_data = {'날짜정보': date_info}
            all_table_hashes = set()  # 중복 테이블 제거용
            
            # 각 카테고리 처리
            for category in self.config.CATEGORIES:
                try:
                    # 카테고리 탭 클릭
                    if not self.select_category(driver, category):
                        self.logger.log_message(f"{bank_name} 은행 {category} 탭 클릭 실패, 다음 카테고리로 진행")
                        continue
                    
                    # 테이블 추출
                    tables = self.extract_tables_from_page(driver)
                    if not tables:
                        self.logger.log_message(f"{bank_name} 은행 {category} 카테고리에서 테이블을 찾을 수 없습니다.")
                        continue
                    
                    # 중복 제거된 유효 테이블 저장
                    valid_tables = []
                    for df in tables:
                        # 테이블 해시 생성 (중복 확인용)
                        try:
                            shape_hash = f"{df.shape}"
                            headers_hash = f"{list(df.columns)}"
                            data_hash = ""
                            if len(df) > 0:
                                data_hash = f"{list(df.iloc[0].astype(str))}"
                            
                            table_hash = f"{shape_hash}_{headers_hash}_{data_hash}"
                            
                            if table_hash not in all_table_hashes:
                                valid_tables.append(df)
                                all_table_hashes.add(table_hash)
                        except:
                            valid_tables.append(df)
                    
                    # 유효한 테이블 저장
                    if valid_tables:
                        result_data[category] = valid_tables
                        self.logger.log_message(f"{bank_name} 은행 {category} 카테고리에서 {len(valid_tables)}개 테이블 추출")
                
                except Exception as e:
                    self.logger.log_message(f"{bank_name} 은행 {category} 카테고리 처리 실패: {str(e)}")
            
            # 데이터 수집 여부 확인
            if not any(isinstance(data, list) and data for key, data in result_data.items() if key != '날짜정보'):
                self.logger.log_message(f"{bank_name} 은행에서 데이터를 추출할 수 없습니다.")
                return None
            
            self.logger.log_message(f"[완료] {bank_name} 은행 데이터 수집 완료")
            return result_data
            
        except Exception as e:
            self.logger.log_message(f"{bank_name} 은행 처리 중 오류 발생: {str(e)}")
            return None
    
    def save_bank_data(self, bank_name, data_dict):
        """수집된 은행 데이터를 엑셀 파일로 저장합니다."""
        if not data_dict:
            return False
        
        try:
            # 날짜 정보 추출
            date_info = data_dict.get('날짜정보', '날짜정보없음')
            date_info = date_info.replace('/', '-').replace('\\', '-')  # 파일명에 사용할 수 없는 문자 제거
            
            # 파일명에 날짜 정보 포함
            excel_path = os.path.join(self.config.output_dir, f"{bank_name}_{date_info}.xlsx")
            
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                # 날짜 정보 시트 생성
                date_df = pd.DataFrame({
                    '은행명': [bank_name],
                    '공시 날짜': [date_info],
                    '추출 일시': [datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                    '스크래핑 시스템': [f'통일경영공시 자동 스크래퍼 v{self.config.VERSION}']
                })
                date_df.to_excel(writer, sheet_name='공시정보', index=False)
                
                # 각 카테고리별 데이터 저장
                for category, tables in data_dict.items():
                    if category == '날짜정보' or not tables:
                        continue
                    
                    # 각 카테고리의 테이블을 별도 시트로 저장
                    for i, df in enumerate(tables):
                        # 시트명 생성
                        if i == 0:
                            sheet_name = category
                        else:
                            sheet_name = f"{category}_{i+1}"
                        
                        # 시트명 길이 제한 (엑셀 제한: 31자)
                        if len(sheet_name) > 31:
                            sheet_name = sheet_name[:31]
                        
                        # MultiIndex 확인 및 처리
                        if isinstance(df.columns, pd.MultiIndex):
                            new_cols = []
                            for col in df.columns:
                                if isinstance(col, tuple):
                                    col_parts = [str(c).strip() for c in col if str(c).strip() and str(c).lower() != 'nan']
                                    new_cols.append('_'.join(col_parts) if col_parts else f"Column_{len(new_cols)+1}")
                                else:
                                    new_cols.append(str(col))
                            df.columns = new_cols
                        
                        # 데이터프레임 저장
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            self.logger.log_message(f"{bank_name} 은행 데이터 저장 완료: {excel_path}")
            return True
            
        except Exception as e:
            self.logger.log_message(f"{bank_name} 은행 데이터 저장 오류: {str(e)}")
            return False
    
    def save_bank_data_to_md(self, bank_name, data_dict, is_settlement=False):
        """수집된 은행 데이터를 마크다운 파일로 저장합니다."""
        if not data_dict:
            return False
        
        try:
            # 날짜 정보 추출
            date_info = data_dict.get('날짜정보', '날짜정보없음')
            date_info = date_info.replace('/', '-').replace('\\', '-')
            
            # 파일명 설정
            file_suffix = "결산" if is_settlement else "분기"
            md_path = os.path.join(self.config.output_dir, f"{bank_name}_{file_suffix}_{date_info}.md")
            
            with open(md_path, 'w', encoding='utf-8') as f:
                # 헤더 작성
                f.write(f"# {bank_name} 저축은행 {file_suffix}공시 데이터\n\n")
                f.write(f"## 기본 정보\n\n")
                f.write(f"- **은행명**: {bank_name}\n")
                f.write(f"- **공시 날짜**: {date_info}\n")
                f.write(f"- **추출 일시**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"- **스크래핑 시스템**: {file_suffix}공시 자동 스크래퍼 v{self.config.VERSION}\n\n")
                
                # 목차 생성
                f.write("## 목차\n\n")
                for category in data_dict.keys():
                    if category != '날짜정보' and data_dict[category]:
                        f.write(f"- [{category}](#{category.lower().replace(' ', '-')})\n")
                f.write("\n")
                
                # 각 카테고리별 데이터 작성
                for category, tables in data_dict.items():
                    if category == '날짜정보' or not tables:
                        continue
                    
                    f.write(f"## {category}\n\n")
                    
                    for i, df in enumerate(tables):
                        if i > 0:
                            f.write(f"### {category} - 테이블 {i+1}\n\n")
                        
                        # DataFrame을 마크다운 테이블로 변환
                        if not df.empty:
                            # 열 이름 정리
                            df_clean = df.copy()
                            df_clean.columns = [str(col).replace('\n', ' ').replace('|', '\\|').strip() for col in df_clean.columns]
                            
                            # 마크다운 테이블 헤더
                            headers = '| ' + ' | '.join(df_clean.columns) + ' |\n'
                            separator = '|' + '|'.join([' --- ' for _ in df_clean.columns]) + '|\n'
                            f.write(headers)
                            f.write(separator)
                            
                            # 데이터 행 (최대 50행까지만 표시)
                            max_rows = min(50, len(df_clean))
                            for idx in range(max_rows):
                                row = df_clean.iloc[idx]
                                row_data = []
                                for value in row:
                                    # 값 정리 (파이프 문자 이스케이프, 개행 제거)
                                    str_value = str(value).replace('|', '\\|').replace('\n', ' ').replace('\r', '').strip()
                                    if str_value == 'nan' or str_value == 'None':
                                        str_value = ''
                                    row_data.append(str_value)
                                f.write('| ' + ' | '.join(row_data) + ' |\n')
                            
                            if len(df_clean) > 50:
                                f.write(f"\n*({len(df_clean) - 50}개 행 더 있음...)*\n")
                            
                            f.write('\n')
                        else:
                            f.write("*데이터가 없습니다.*\n\n")
                
                # 푸터
                f.write("---\n")
                f.write(f"*이 문서는 {datetime.now().strftime('%Y년 %m월 %d일 %H:%M:%S')}에 자동 생성되었습니다.*\n")
            
            self.logger.log_message(f"{bank_name} 은행 MD 파일 저장 완료: {md_path}")
            return True
            
        except Exception as e:
            self.logger.log_message(f"{bank_name} 은행 MD 파일 저장 오류: {str(e)}")
            return False
    
    def worker_process_bank(self, bank_name, progress_callback=None, save_md=False):
        """단일 은행을 처리합니다."""
        driver = self.driver_manager.get_driver()
        
        try:
            # 최대 재시도 횟수만큼 시도
            for attempt in range(self.config.MAX_RETRIES):
                try:
                    # 진행 상황 업데이트
                    if progress_callback:
                        progress_callback(bank_name, "처리 중")
                    
                    # 은행 데이터 스크래핑
                    result_data = self.scrape_bank_data(bank_name, driver)
                    
                    if result_data:
                        # 엑셀 데이터 저장
                        excel_saved = self.save_bank_data(bank_name, result_data)
                        
                        # MD 데이터 저장 (옵션)
                        md_saved = True
                        if save_md:
                            md_saved = self.save_bank_data_to_md(bank_name, result_data, is_settlement=False)
                        
                        if excel_saved and md_saved:
                            self.progress_manager.mark_completed(bank_name)
                            self.driver_manager.return_driver(driver)
                            
                            # 진행 상황 업데이트
                            if progress_callback:
                                progress_callback(bank_name, "완료")
                            
                            return bank_name, True
                        else:
                            if attempt < self.config.MAX_RETRIES - 1:
                                self.logger.log_message(f"{bank_name} 은행 데이터 저장 실패, 재시도 {attempt+1}/{self.config.MAX_RETRIES}...")
                                
                                # 진행 상황 업데이트
                                if progress_callback:
                                    progress_callback(bank_name, f"저장 재시도 {attempt+1}")
                            else:
                                self.logger.log_message(f"{bank_name} 은행 데이터 저장 실패, 최대 시도 횟수 초과")
                                
                                # 진행 상황 업데이트
                                if progress_callback:
                                    progress_callback(bank_name, "저장 실패")
                    else:
                        if attempt < self.config.MAX_RETRIES - 1:
                            self.logger.log_message(f"{bank_name} 은행 데이터 스크래핑 실패, 재시도 {attempt+1}/{self.config.MAX_RETRIES}...")
                            WaitUtils.wait_with_random(1, 2)  # 재시도 전 잠시 대기
                            
                            # 진행 상황 업데이트
                            if progress_callback:
                                progress_callback(bank_name, f"스크래핑 재시도 {attempt+1}")
                        else:
                            self.logger.log_message(f"{bank_name} 은행 데이터 스크래핑 실패, 최대 시도 횟수 초과")
                            
                            # 진행 상황 업데이트
                            if progress_callback:
                                progress_callback(bank_name, "스크래핑 실패")
                
                except Exception as e:
                    if attempt < self.config.MAX_RETRIES - 1:
                        self.logger.log_message(f"{bank_name} 은행 처리 중 오류: {str(e)}, 재시도 {attempt+1}/{self.config.MAX_RETRIES}...")
                        WaitUtils.wait_with_random(1, 2)
                        
                        # 진행 상황 업데이트
                        if progress_callback:
                            progress_callback(bank_name, f"오류, 재시도 {attempt+1}")
                    else:
                        self.logger.log_message(f"{bank_name} 은행 처리 실패: {str(e)}, 최대 시도 횟수 초과")
                        
                        # 진행 상황 업데이트
                        if progress_callback:
                            progress_callback(bank_name, "실패")
            
            # 모든 시도 실패
            self.progress_manager.mark_failed(bank_name)
            self.driver_manager.return_driver(driver)
            
            # 진행 상황 업데이트
            if progress_callback:
                progress_callback(bank_name, "실패")
            
            return bank_name, False
            
        except Exception as e:
            self.logger.log_message(f"{bank_name} 은행 처리 중 예상치 못한 오류: {str(e)}")
            self.progress_manager.mark_failed(bank_name)
            self.driver_manager.return_driver(driver)
            
            # 진행 상황 업데이트
            if progress_callback:
                progress_callback(bank_name, "오류")
            
            return bank_name, False
    
    def process_banks(self, banks, progress_callback=None, save_md=False):
        """은행 목록을 병렬로 처리합니다."""
        with concurrent.futures.ThreadPoolExecutor(max_workers=self.config.MAX_WORKERS) as executor:
            # 작업 제출
            future_to_bank = {
                executor.submit(self.worker_process_bank, bank, progress_callback, save_md): bank
                for bank in banks
            }
            
            # 결과 수집
            results = []
            for future in concurrent.futures.as_completed(future_to_bank):
                bank = future_to_bank[future]
                try:
                    result = future.result()
                    results.append(result)
                except Exception as e:
                    self.logger.log_message(f"{bank} 은행 처리 중 예외 발생: {e}")
                    results.append((bank, False))
            
            return results
    
    def generate_summary_report(self):
        """스크래핑 결과 요약 보고서를 생성합니다."""
        try:
            # 완료된 은행과 실패한 은행 목록
            completed_banks = self.progress_manager.progress.get('completed', [])
            failed_banks = self.progress_manager.progress.get('failed', [])
            
            # 은행별 데이터 요약
            bank_summary = []
            
            # 완료된 은행 파일 검사
            for bank in self.config.BANKS:
                # 각 은행의 엑셀 파일 찾기
                bank_files = [f for f in os.listdir(self.config.output_dir) if f.startswith(f"{bank}_") and f.endswith(".xlsx")]
                
                if bank_files:
                    try:
                        # 가장 최근 파일 선택
                        latest_file = sorted(bank_files)[-1]
                        file_path = os.path.join(self.config.output_dir, latest_file)
                        
                        # 엑셀 파일 분석
                        xls = pd.ExcelFile(file_path)
                        sheet_count = len(xls.sheet_names)
                        
                        # 카테고리 추출
                        categories = []
                        for sheet in xls.sheet_names:
                            if sheet != '공시정보':
                                category = sheet.split('_')[0] if '_' in sheet else sheet
                                categories.append(category)
                        
                        # 중복 제거
                        categories = sorted(list(set(categories)))
                        
                        # 날짜 정보 추출
                        date_info = "날짜 정보 없음"
                        if '공시정보' in xls.sheet_names:
                            info_df = pd.read_excel(file_path, sheet_name='공시정보')
                            if '공시 날짜' in info_df.columns and not info_df['공시 날짜'].empty:
                                date_info = str(info_df['공시 날짜'].iloc[0])
                        
                        status = '완료' if set(categories) >= set(self.config.CATEGORIES) else '부분 완료'
                        
                        bank_summary.append({
                            '은행명': bank,
                            '스크래핑 상태': status,
                            '공시 날짜': date_info,
                            '시트 수': sheet_count - 1,  # 공시정보 시트 제외
                            '스크래핑된 카테고리': ', '.join(categories)
                        })
                    except Exception as e:
                        bank_summary.append({
                            '은행명': bank,
                            '스크래핑 상태': '파일 손상',
                            '공시 날짜': '확인 불가',
                            '시트 수': '확인 불가',
                            '스크래핑된 카테고리': f'오류: {str(e)}'
                        })
                else:
                    status = '실패' if bank in failed_banks else '미처리'
                    bank_summary.append({
                        '은행명': bank,
                        '스크래핑 상태': status,
                        '공시 날짜': '',
                        '시트 수': 0,
                        '스크래핑된 카테고리': ''
                    })
            
            # 요약 DataFrame 생성
            summary_df = pd.DataFrame(bank_summary)
            
            # 완료 상태별 정렬
            status_order = {'완료': 0, '부분 완료': 1, '파일 손상': 2, '실패': 3, '미처리': 4}
            summary_df['상태순서'] = summary_df['스크래핑 상태'].map(status_order)
            summary_df = summary_df.sort_values(['상태순서', '은행명']).drop('상태순서', axis=1)
            
            # 요약 저장
            summary_file = os.path.join(self.config.output_dir, f"저축은행_스크래핑_요약_{self.config.today}.xlsx")
            summary_df.to_excel(summary_file, index=False)
            
            # 통계 정보
            stats = {
                '전체 은행 수': len(self.config.BANKS),
                '완료 은행 수': len([r for r in bank_summary if r['스크래핑 상태'] == '완료']),
                '부분 완료 은행 수': len([r for r in bank_summary if r['스크래핑 상태'] == '부분 완료']),
                '실패 은행 수': len([r for r in bank_summary if r['스크래핑 상태'] in ['실패', '파일 손상']]),
                '성공률': f"{len([r for r in bank_summary if r['스크래핑 상태'] in ['완료', '부분 완료']]) / len(self.config.BANKS) * 100:.2f}%"
            }
            
            self.logger.log_message("\n===== 스크래핑 결과 요약 =====")
            for key, value in stats.items():
                self.logger.log_message(f"{key}: {value}")
            
            self.logger.log_message(f"요약 파일 저장 완료: {summary_file}")
            return summary_file, stats, summary_df
            
        except Exception as e:
            self.logger.log_message(f"요약 보고서 생성 오류: {str(e)}")
            return None, {}, None

    def create_consolidated_md_report(self):
        """모든 개별 MD 파일을 하나의 통합 MD 파일로 만듭니다."""
        try:
            self.logger.log_message("\n===== 통합 MD 보고서 생성 시작 =====")
            
            # MD 파일 찾기
            md_files = []
            for file in os.listdir(self.config.output_dir):
                if file.endswith('.md') and not any(keyword in file for keyword in ['통합', '요약', 'consolidated', 'summary']):
                    md_files.append(os.path.join(self.config.output_dir, file))
            
            if not md_files:
                self.logger.log_message("통합할 MD 파일이 없습니다.")
                return None
            
            # 통합 MD 파일 경로
            consolidated_md_path = os.path.join(self.config.output_dir, 
                                               f'저축은행_분기공시_통합보고서_{self.config.today}.md')
            
            with open(consolidated_md_path, 'w', encoding='utf-8') as f_out:
                # 헤더 작성
                f_out.write(f"# 📊 저축은행 분기공시 통합 보고서\n\n")
                f_out.write(f"**생성일시**: {datetime.now().strftime('%Y년 %m월 %d일 %H:%M:%S')}\n")
                f_out.write(f"**총 {len(md_files)}개 은행 데이터 통합**\n\n")
                f_out.write("---\n\n")
                
                # 목차 생성
                f_out.write("## 📑 목차\n\n")
                bank_names = []
                for md_file in sorted(md_files):
                    bank_name = os.path.basename(md_file).split('_')[0]
                    bank_names.append(bank_name)
                    f_out.write(f"- [{bank_name} 저축은행](#{bank_name.lower()}-저축은행)\n")
                f_out.write("\n---\n\n")
                
                # 각 MD 파일 내용 통합
                for idx, md_file in enumerate(sorted(md_files), 1):
                    bank_name = os.path.basename(md_file).split('_')[0]
                    
                    f_out.write(f"## {bank_name} 저축은행\n\n")
                    
                    try:
                        with open(md_file, 'r', encoding='utf-8') as f_in:
                            content = f_in.read()
                            
                            # 기존 헤더 레벨 조정 (# -> ###, ## -> ####)
                            content = re.sub(r'^# ', '### ', content, flags=re.MULTILINE)
                            content = re.sub(r'^## ', '#### ', content, flags=re.MULTILINE)
                            
                            # 기본 정보 섹션만 추출하거나 전체 내용 포함
                            lines = content.split('\n')
                            filtered_lines = []
                            
                            for line in lines:
                                if '### ' in line and '저축은행' in line:
                                    continue  # 중복 제목 제거
                                filtered_lines.append(line)
                            
                            f_out.write('\n'.join(filtered_lines))
                            
                    except Exception as e:
                        self.logger.log_message(f"{bank_name} MD 파일 읽기 오류: {str(e)}", verbose=False)
                        f_out.write(f"*{bank_name} 데이터 읽기 실패*\n\n")
                    
                    if idx < len(md_files):
                        f_out.write("\n\n---\n\n")
                
                # 푸터
                f_out.write("\n---\n")
                f_out.write(f"*이 통합 보고서는 {len(md_files)}개 은행의 개별 MD 파일을 자동으로 통합하여 생성되었습니다.*\n")
                f_out.write(f"*저축은행 분기공시 자동 스크래퍼 v{self.config.VERSION}*\n")
            
            self.logger.log_message(f"통합 MD 보고서 생성 완료: {consolidated_md_path}")
            return consolidated_md_path
            
        except Exception as e:
            self.logger.log_message(f"통합 MD 보고서 생성 오류: {str(e)}")
            return None    
    
    def generate_summary_report_md(self):
        """스크래핑 결과 요약을 마크다운으로 생성합니다."""
        try:
            # 기존 요약 보고서 생성
            summary_file, stats, summary_df = self.generate_summary_report()
            
            if summary_df is None:
                return None
            
            # MD 파일 경로
            md_summary_file = os.path.join(self.config.output_dir, 
                                          f"저축은행_분기공시_스크래핑_요약_{self.config.today}.md")
            
            with open(md_summary_file, 'w', encoding='utf-8') as f:
                f.write(f"# 저축은행 분기공시 스크래핑 결과 요약\n\n")
                f.write(f"📅 **보고서 생성일**: {datetime.now().strftime('%Y년 %m월 %d일 %H:%M:%S')}\n\n")
                
                # 전체 통계
                f.write("## 📊 전체 통계\n\n")
                for key, value in stats.items():
                    icon = "✅" if "성공" in key else "📈" if "률" in key else "🏦"
                    f.write(f"{icon} **{key}**: {value}\n")
                f.write("\n")
                
                # 상태별 은행 분류
                f.write("## 📋 상태별 은행 현황\n\n")
                
                status_groups = summary_df.groupby('스크래핑 상태')
                for status, group in status_groups:
                    emoji_map = {
                        '완료': '✅',
                        '부분 완료': '⚠️', 
                        '실패': '❌',
                        '파일 손상': '💥',
                        '미처리': '⏳'
                    }
                    emoji = emoji_map.get(status, '📋')
                    
                    f.write(f"### {emoji} {status} ({len(group)}개 은행)\n\n")
                    
                    if len(group) > 0:
                        for _, row in group.iterrows():
                            f.write(f"- **{row['은행명']}**")
                            if row['공시 날짜'] and row['공시 날짜'] != '':
                                f.write(f" (📅 {row['공시 날짜']})")
                            if row['스크래핑된 카테고리'] and row['스크래핑된 카테고리'] != '':
                                f.write(f" - 📂 {row['스크래핑된 카테고리']}")
                            f.write("\n")
                    f.write("\n")
                
                # 상세 테이블
                f.write("## 📄 상세 현황표\n\n")
                
                # 마크다운 테이블 생성
                headers = '| ' + ' | '.join(summary_df.columns) + ' |\n'
                separator = '|' + '|'.join([' --- ' for _ in summary_df.columns]) + '|\n'
                f.write(headers)
                f.write(separator)
                
                for _, row in summary_df.iterrows():
                    row_data = []
                    for value in row:
                        str_value = str(value).replace('|', '\\|').replace('\n', ' ')
                        if str_value == 'nan' or str_value == 'None':
                            str_value = ''
                        row_data.append(str_value)
                    f.write('| ' + ' | '.join(row_data) + ' |\n')
                
                f.write('\n')
                
                # 권장사항
                f.write("## 💡 권장사항\n\n")
                
                failed_count = len(summary_df[summary_df['스크래핑 상태'].isin(['실패', '파일 손상'])])
                if failed_count > 0:
                    f.write(f"🔧 {failed_count}개 은행의 데이터 수집에 실패했습니다. 수동으로 재시도를 고려해보세요.\n\n")
                
                partial_count = len(summary_df[summary_df['스크래핑 상태'] == '부분 완료'])
                if partial_count > 0:
                    f.write(f"⚠️ {partial_count}개 은행의 데이터가 부분적으로만 수집되었습니다. 누락된 카테고리를 확인해보세요.\n\n")
                
                success_rate = float(stats['성공률'].replace('%', ''))
                if success_rate < 90:
                    f.write("🔍 전체 성공률이 90% 미만입니다. 네트워크 상태나 웹사이트 변경사항을 확인해보세요.\n\n")
                else:
                    f.write("🎉 스크래핑이 성공적으로 완료되었습니다!\n\n")
                
                # 파일 목록
                f.write("## 📁 생성된 파일들\n\n")
                f.write("다음 파일들이 생성되었습니다:\n\n")
                f.write("- 📊 엑셀 파일: 각 은행별 재무 데이터\n")
                f.write("- 📝 MD 파일: 마크다운 형식 데이터 (선택한 경우)\n")
                f.write("- 📋 요약 보고서: 전체 스크래핑 결과\n")
                f.write("- 📈 통합 재무 보고서: 모든 은행 데이터 통합\n\n")
                
                # 푸터
                f.write("---\n")
                f.write(f"*이 요약 보고서는 저축은행 통합 데이터 스크래퍼 v{self.config.VERSION}에 의해 자동 생성되었습니다.*\n")
                f.write(f"*생성 시간: {datetime.now().strftime('%Y년 %m월 %d일 %H:%M:%S')}*\n")
            
            self.logger.log_message(f"MD 요약 보고서 저장 완료: {md_summary_file}")
            return md_summary_file
            
        except Exception as e:
            self.logger.log_message(f"MD 요약 보고서 생성 오류: {str(e)}")
            return None
    
    def extract_financial_data_from_excel(self, excel_path, bank_name):
        """개별 은행 엑셀 파일에서 주요 재무 데이터를 추출합니다. (당기/전년동기 비교 포함)"""
        try:
            # 엑셀 파일 열기
            xls = pd.ExcelFile(excel_path)
            
            # 기본 정보 추출 - 확장된 구조
            financial_data = {
                '은행명': bank_name,
                '재무정보 날짜': '데이터 없음',
                '분기': '데이터 없음',
                
                # 당기 데이터
                '당기_총자산': None,
                '당기_자기자본': None,
                '당기_총여신': None,
                '당기_총수신': None,
                '당기_수익합계': None,
                '당기_비용합계': None,
                '당기_당기순이익': None,
                '당기_고정이하여신비율(%)': None,
                '당기_위험가중자산에_대한_자기자본비율(%)': None,
                
                # 전년동기 데이터
                '전년동기_총자산': None,
                '전년동기_자기자본': None,
                '전년동기_총여신': None,
                '전년동기_총수신': None,
                '전년동기_수익합계': None,
                '전년동기_비용합계': None,
                '전년동기_당기순이익': None,
                '전년동기_고정이하여신비율(%)': None,
                '전년동기_위험가중자산에_대한_자기자본비율(%)': None,
                
                # 증감 데이터 (나중에 계산)
                '증감_총자산': None,
                '증감_자기자본': None,
                '증감_총여신': None,
                '증감_총수신': None,
                '증감_수익합계': None,
                '증감_비용합계': None,
                '증감_당기순이익': None,
                '증감_고정이하여신비율(%)': None,
                '증감_위험가중자산에_대한_자기자본비율(%)': None
            }
            
            # 공시정보 시트에서 날짜 정보 추출
            if '공시정보' in xls.sheet_names:
                info_df = pd.read_excel(excel_path, sheet_name='공시정보')
                if '공시 날짜' in info_df.columns and not info_df['공시 날짜'].empty:
                    date_info = str(info_df['공시 날짜'].iloc[0])
                    financial_data['재무정보 날짜'] = date_info
                    
                    # 날짜에서 연도와 월 추출하여 분기 계산
                    import re
                    date_match = re.search(r'(\d{4})년(\d{1,2})월', date_info)
                    if date_match:
                        year = int(date_match.group(1))
                        month = int(date_match.group(2))
                        quarter = (month - 1) // 3 + 1
                        financial_data['분기'] = f"{year}년 {quarter}분기"
            
            # 각 시트에서 재무 데이터 추출 (당기/전년동기 구분)
            for sheet_name in xls.sheet_names:
                if sheet_name == '공시정보':
                    continue
                    
                try:
                    df = pd.read_excel(excel_path, sheet_name=sheet_name)
                    
                    # 데이터프레임이 비어있지 않은 경우에만 처리
                    if df.empty:
                        continue
                    
                    # 당기와 전년동기 데이터를 구분하여 추출
                    self._extract_period_data(df, financial_data, bank_name, sheet_name)
                    
                except Exception as e:
                    self.logger.log_message(f"{bank_name} 은행 {sheet_name} 시트 처리 중 오류: {str(e)}", verbose=False)
                    continue
            
            # 증감 계산
            self._calculate_financial_changes(financial_data)
            
            # 디버깅용 로그
            current_data = [financial_data[f'당기_{key}'] for key in ['총자산', '자기자본'] if financial_data.get(f'당기_{key}')]
            previous_data = [financial_data[f'전년동기_{key}'] for key in ['총자산', '자기자본'] if financial_data.get(f'전년동기_{key}')]
            
            if current_data:
                self.logger.log_message(f"{bank_name} - 당기 데이터 {len(current_data)}개, 전년동기 데이터 {len(previous_data)}개 추출", verbose=False)
            
            return financial_data
            
        except Exception as e:
            self.logger.log_message(f"{bank_name} 은행 재무 데이터 추출 오류: {str(e)}")
            return None
    
    def _extract_period_data(self, df, financial_data, bank_name, sheet_name):
        """데이터프레임에서 당기/전년동기 구분하여 재무 데이터 추출"""
        try:
            # 컬럼에서 당기/전년동기 구분
            current_period_cols = []
            previous_period_cols = []
            
            for col in df.columns:
                col_str = str(col).strip()
                if any(keyword in col_str for keyword in ['당기', '현재', '이번']):
                    current_period_cols.append(col)
                elif any(keyword in col_str for keyword in ['전년', '작년', '이전', '전기']):
                    previous_period_cols.append(col)
            
            # 재무현황 관련 데이터 추출
            if '재무현황' in sheet_name or '재무' in sheet_name:
                self._extract_balance_sheet_data(df, financial_data, current_period_cols, previous_period_cols, bank_name)
            
            # 영업개황 관련 데이터 추출
            elif '영업개황' in sheet_name or '영업' in sheet_name:
                self._extract_business_data(df, financial_data, current_period_cols, previous_period_cols, bank_name)
            
            # 손익현황 관련 데이터 추출
            elif '손익현황' in sheet_name or '손익' in sheet_name:
                self._extract_income_statement_data(df, financial_data, current_period_cols, previous_period_cols, bank_name)
            
            # 기타 비율 데이터 추출
            elif '기타' in sheet_name or True:  # 모든 시트에서 비율 찾기
                self._extract_ratio_data(df, financial_data, current_period_cols, previous_period_cols, bank_name)
                
        except Exception as e:
            self.logger.log_message(f"{bank_name} 시트 {sheet_name} 기간별 데이터 추출 오류: {str(e)}", verbose=False)
    
    def _extract_balance_sheet_data(self, df, financial_data, current_cols, previous_cols, bank_name):
        """대차대조표 데이터 추출 (총자산, 자기자본)"""
        try:
            for idx in range(len(df)):
                for col in df.columns:
                    try:
                        cell_value = str(df.iloc[idx][col]).strip()
                        
                        # 총자산 찾기
                        if '총자산' in cell_value or '자산총계' in cell_value:
                            # 당기 데이터
                            if financial_data['당기_총자산'] is None:
                                for target_col in current_cols or df.columns:
                                    if target_col != col:
                                        try:
                                            value = pd.to_numeric(str(df.iloc[idx][target_col]).replace(',', ''), errors='coerce')
                                            if pd.notna(value) and value > 0:
                                                financial_data['당기_총자산'] = value
                                                break
                                        except:
                                            pass
                            
                            # 전년동기 데이터
                            if financial_data['전년동기_총자산'] is None:
                                for target_col in previous_cols or []:
                                    if target_col != col:
                                        try:
                                            value = pd.to_numeric(str(df.iloc[idx][target_col]).replace(',', ''), errors='coerce')
                                            if pd.notna(value) and value > 0:
                                                financial_data['전년동기_총자산'] = value
                                                break
                                        except:
                                            pass
                        
                        # 자기자본 찾기
                        if ('자기자본' in cell_value or '자본총계' in cell_value or '자본합계' in cell_value) and '자산' not in cell_value:
                            # 당기 데이터
                            if financial_data['당기_자기자본'] is None:
                                for target_col in current_cols or df.columns:
                                    if target_col != col:
                                        try:
                                            value = pd.to_numeric(str(df.iloc[idx][target_col]).replace(',', ''), errors='coerce')
                                            if pd.notna(value) and value > 0:
                                                if financial_data['당기_총자산'] is None or value != financial_data['당기_총자산']:
                                                    financial_data['당기_자기자본'] = value
                                                    break
                                        except:
                                            pass
                            
                            # 전년동기 데이터
                            if financial_data['전년동기_자기자본'] is None:
                                for target_col in previous_cols or []:
                                    if target_col != col:
                                        try:
                                            value = pd.to_numeric(str(df.iloc[idx][target_col]).replace(',', ''), errors='coerce')
                                            if pd.notna(value) and value > 0:
                                                if financial_data['전년동기_총자산'] is None or value != financial_data['전년동기_총자산']:
                                                    financial_data['전년동기_자기자본'] = value
                                                    break
                                        except:
                                            pass
                    except:
                        pass
                        
        except Exception as e:
            self.logger.log_message(f"{bank_name} 대차대조표 데이터 추출 오류: {str(e)}", verbose=False)
    
    def _extract_business_data(self, df, financial_data, current_cols, previous_cols, bank_name):
        """영업개황 데이터 추출 (총여신, 총수신)"""
        try:
            for idx in range(len(df)):
                for col in df.columns:
                    try:
                        cell_value = str(df.iloc[idx][col]).strip()
                        
                        # 총여신 찾기
                        if '총여신' in cell_value or '여신총계' in cell_value or '대출채권' in cell_value:
                            # 당기 데이터
                            if financial_data['당기_총여신'] is None:
                                for target_col in current_cols or df.columns:
                                    if target_col != col:
                                        try:
                                            value = pd.to_numeric(str(df.iloc[idx][target_col]).replace(',', ''), errors='coerce')
                                            if pd.notna(value) and value > 0:
                                                financial_data['당기_총여신'] = value
                                                break
                                        except:
                                            pass
                            
                            # 전년동기 데이터
                            if financial_data['전년동기_총여신'] is None:
                                for target_col in previous_cols or []:
                                    if target_col != col:
                                        try:
                                            value = pd.to_numeric(str(df.iloc[idx][target_col]).replace(',', ''), errors='coerce')
                                            if pd.notna(value) and value > 0:
                                                financial_data['전년동기_총여신'] = value
                                                break
                                        except:
                                            pass
                        
                        # 총수신 찾기
                        if '총수신' in cell_value or '수신총계' in cell_value or '예수금' in cell_value:
                            # 당기 데이터
                            if financial_data['당기_총수신'] is None:
                                for target_col in current_cols or df.columns:
                                    if target_col != col:
                                        try:
                                            value = pd.to_numeric(str(df.iloc[idx][target_col]).replace(',', ''), errors='coerce')
                                            if pd.notna(value) and value > 0:
                                                financial_data['당기_총수신'] = value
                                                break
                                        except:
                                            pass
                            
                            # 전년동기 데이터
                            if financial_data['전년동기_총수신'] is None:
                                for target_col in previous_cols or []:
                                    if target_col != col:
                                        try:
                                            value = pd.to_numeric(str(df.iloc[idx][target_col]).replace(',', ''), errors='coerce')
                                            if pd.notna(value) and value > 0:
                                                financial_data['전년동기_총수신'] = value
                                                break
                                        except:
                                            pass
                    except:
                        pass
                        
        except Exception as e:
            self.logger.log_message(f"{bank_name} 영업개황 데이터 추출 오류: {str(e)}", verbose=False)
    
    def _extract_income_statement_data(self, df, financial_data, current_cols, previous_cols, bank_name):
        """손익계산서 데이터 추출 (수익, 비용, 순이익) - 누적 데이터 우선"""
        try:
            for idx in range(len(df)):
                for col in df.columns:
                    try:
                        cell_value = str(df.iloc[idx][col]).strip()
                        
                        # 수익합계 찾기
                        if '수익합계' in cell_value or '영업수익' in cell_value or '총수익' in cell_value:
                            # 당기 데이터
                            if financial_data['당기_수익합계'] is None:
                                for target_col in current_cols or df.columns:
                                    if target_col != col:
                                        try:
                                            value = pd.to_numeric(str(df.iloc[idx][target_col]).replace(',', ''), errors='coerce')
                                            if pd.notna(value) and value > 0:
                                                financial_data['당기_수익합계'] = value
                                                break
                                        except:
                                            pass
                            
                            # 전년동기 데이터
                            if financial_data['전년동기_수익합계'] is None:
                                for target_col in previous_cols or []:
                                    if target_col != col:
                                        try:
                                            value = pd.to_numeric(str(df.iloc[idx][target_col]).replace(',', ''), errors='coerce')
                                            if pd.notna(value) and value > 0:
                                                financial_data['전년동기_수익합계'] = value
                                                break
                                        except:
                                            pass
                        
                        # 비용합계 찾기
                        if '비용합계' in cell_value or '영업비용' in cell_value or '총비용' in cell_value:
                            # 당기 데이터
                            if financial_data['당기_비용합계'] is None:
                                for target_col in current_cols or df.columns:
                                    if target_col != col:
                                        try:
                                            value = pd.to_numeric(str(df.iloc[idx][target_col]).replace(',', ''), errors='coerce')
                                            if pd.notna(value) and value > 0:
                                                financial_data['당기_비용합계'] = value
                                                break
                                        except:
                                            pass
                            
                            # 전년동기 데이터
                            if financial_data['전년동기_비용합계'] is None:
                                for target_col in previous_cols or []:
                                    if target_col != col:
                                        try:
                                            value = pd.to_numeric(str(df.iloc[idx][target_col]).replace(',', ''), errors='coerce')
                                            if pd.notna(value) and value > 0:
                                                financial_data['전년동기_비용합계'] = value
                                                break
                                        except:
                                            pass
                        
                        # 당기순이익 찾기 - 누적 데이터 우선
                        if '당기순이익' in cell_value or '순이익' in cell_value:
                            # 누적 표시가 있는지 확인
                            is_cumulative = False
                            
                            # 같은 행 또는 인접 행에서 '누적', '누계', '합계' 등의 키워드 찾기
                            for check_idx in range(max(0, idx-2), min(len(df), idx+3)):
                                for check_col in df.columns:
                                    check_value = str(df.iloc[check_idx][check_col]).strip().lower()
                                    if any(keyword in check_value for keyword in ['누적', '누계', '합계', 'ytd', 'cumulative', '1월~', '연초~']):
                                        is_cumulative = True
                                        break
                                if is_cumulative:
                                    break
                            
                            # 분기 표시가 있는지 확인
                            is_quarterly = False
                            for check_col in df.columns:
                                check_value = str(df.iloc[idx][check_col]).strip().lower()
                                if any(keyword in check_value for keyword in ['분기', 'quarter', 'q1', 'q2', 'q3', 'q4', '3개월']):
                                    is_quarterly = True
                                    break
                            
                            # 당기 데이터 - 누적 우선, 분기는 나중에
                            if financial_data['당기_당기순이익'] is None or (not is_quarterly and is_cumulative):
                                for target_col in current_cols or df.columns:
                                    if target_col != col:
                                        try:
                                            value = pd.to_numeric(str(df.iloc[idx][target_col]).replace(',', ''), errors='coerce')
                                            if pd.notna(value):
                                                # 누적 데이터이거나 분기 표시가 없으면 저장
                                                if is_cumulative or not is_quarterly:
                                                    financial_data['당기_당기순이익'] = value
                                                    self.logger.log_message(f"{bank_name} - 당기순이익(누적) 발견: {value}", verbose=False)
                                                    break
                                                # 분기 데이터인 경우 기존 값이 없을 때만 저장
                                                elif financial_data['당기_당기순이익'] is None:
                                                    financial_data['당기_당기순이익'] = value
                                                    self.logger.log_message(f"{bank_name} - 당기순이익(분기) 발견: {value}", verbose=False)
                                                    break
                                        except:
                                            pass
                            
                            # 전년동기 데이터도 동일한 로직 적용
                            if financial_data['전년동기_당기순이익'] is None or (not is_quarterly and is_cumulative):
                                for target_col in previous_cols or []:
                                    if target_col != col:
                                        try:
                                            value = pd.to_numeric(str(df.iloc[idx][target_col]).replace(',', ''), errors='coerce')
                                            if pd.notna(value):
                                                if is_cumulative or not is_quarterly:
                                                    financial_data['전년동기_당기순이익'] = value
                                                    self.logger.log_message(f"{bank_name} - 전년동기 순이익(누적) 발견: {value}", verbose=False)
                                                    break
                                                elif financial_data['전년동기_당기순이익'] is None:
                                                    financial_data['전년동기_당기순이익'] = value
                                                    self.logger.log_message(f"{bank_name} - 전년동기 순이익(분기) 발견: {value}", verbose=False)
                                                    break
                                        except:
                                            pass
                    except:
                        pass
                        
        except Exception as e:
            self.logger.log_message(f"{bank_name} 손익계산서 데이터 추출 오류: {str(e)}", verbose=False)
            
    def _extract_ratio_data(self, df, financial_data, current_cols, previous_cols, bank_name):
        """비율 데이터 추출 (고정이하여신비율, BIS비율)"""
        try:
            for idx in range(len(df)):
                for col in df.columns:
                    try:
                        cell_value = str(df.iloc[idx][col]).strip()
                        
                        # 고정이하여신비율 찾기
                        if '고정이하여신비율' in cell_value or 'NPL' in cell_value:
                            # 당기 데이터
                            if financial_data['당기_고정이하여신비율(%)'] is None:
                                for target_col in current_cols or df.columns:
                                    if target_col != col:
                                        try:
                                            value_str = str(df.iloc[idx][target_col]).replace('%', '').replace(',', '').strip()
                                            value = pd.to_numeric(value_str, errors='coerce')
                                            if pd.notna(value) and 0 <= value <= 100:
                                                financial_data['당기_고정이하여신비율(%)'] = value
                                                break
                                        except:
                                            pass
                            
                            # 전년동기 데이터
                            if financial_data['전년동기_고정이하여신비율(%)'] is None:
                                for target_col in previous_cols or []:
                                    if target_col != col:
                                        try:
                                            value_str = str(df.iloc[idx][target_col]).replace('%', '').replace(',', '').strip()
                                            value = pd.to_numeric(value_str, errors='coerce')
                                            if pd.notna(value) and 0 <= value <= 100:
                                                financial_data['전년동기_고정이하여신비율(%)'] = value
                                                break
                                        except:
                                            pass
                        
                        # 위험가중자산에 대한 자기자본비율 찾기
                        if 'BIS' in cell_value or '자기자본비율' in cell_value or '위험가중자산' in cell_value:
                            # 당기 데이터
                            if financial_data['당기_위험가중자산에_대한_자기자본비율(%)'] is None:
                                for target_col in current_cols or df.columns:
                                    if target_col != col:
                                        try:
                                            value_str = str(df.iloc[idx][target_col]).replace('%', '').replace(',', '').strip()
                                            value = pd.to_numeric(value_str, errors='coerce')
                                            if pd.notna(value) and 0 <= value <= 100:
                                                financial_data['당기_위험가중자산에_대한_자기자본비율(%)'] = value
                                                break
                                        except:
                                            pass
                            
                            # 전년동기 데이터
                            if financial_data['전년동기_위험가중자산에_대한_자기자본비율(%)'] is None:
                                for target_col in previous_cols or []:
                                    if target_col != col:
                                        try:
                                            value_str = str(df.iloc[idx][target_col]).replace('%', '').replace(',', '').strip()
                                            value = pd.to_numeric(value_str, errors='coerce')
                                            if pd.notna(value) and 0 <= value <= 100:
                                                financial_data['전년동기_위험가중자산에_대한_자기자본비율(%)'] = value
                                                break
                                        except:
                                            pass
                    except:
                        pass
                        
        except Exception as e:
            self.logger.log_message(f"{bank_name} 비율 데이터 추출 오류: {str(e)}", verbose=False)
    
    def _calculate_financial_changes(self, financial_data):
        """당기와 전년동기 데이터 간의 증감을 계산합니다."""
        try:
            # 계산할 항목 목록
            items = [
                '총자산', '자기자본', '총여신', '총수신',
                '수익합계', '비용합계', '당기순이익',
                '고정이하여신비율(%)', '위험가중자산에_대한_자기자본비율(%)'
            ]
            
            for item in items:
                current_key = f'당기_{item}'
                previous_key = f'전년동기_{item}'
                change_key = f'증감_{item}'
                
                current_value = financial_data.get(current_key)
                previous_value = financial_data.get(previous_key)
                
                if current_value is not None and previous_value is not None:
                    try:
                        # 절대 증감 계산
                        absolute_change = float(current_value) - float(previous_value)
                        
                        # 증감률 계산 (전년동기 대비)
                        if float(previous_value) != 0:
                            change_rate = (absolute_change / float(previous_value)) * 100
                            financial_data[change_key] = {
                                '절대증감': absolute_change,
                                '증감률(%)': change_rate
                            }
                        else:
                            financial_data[change_key] = {
                                '절대증감': absolute_change,
                                '증감률(%)': 'N/A'  # 0으로 나누기 방지
                            }
                    except (ValueError, ZeroDivisionError):
                        financial_data[change_key] = None
                else:
                    financial_data[change_key] = None
                    
        except Exception as e:
            self.logger.log_message(f"증감 계산 오류: {str(e)}", verbose=False)
    
    def create_consolidated_financial_report(self):
        """79개 은행의 재무 데이터를 통합한 엑셀 보고서를 생성합니다."""
        try:
            self.logger.log_message("\n===== 재무 데이터 통합 보고서 생성 시작 =====")
            
            # 모든 은행의 재무 데이터를 저장할 리스트
            all_financial_data = []
            
            # 각 은행의 엑셀 파일에서 데이터 추출
            for bank in self.config.BANKS:
                # 해당 은행의 가장 최근 엑셀 파일 찾기
                bank_files = [f for f in os.listdir(self.config.output_dir) 
                             if f.startswith(f"{bank}_") and f.endswith(".xlsx")]
                
                if bank_files:
                    # 가장 최근 파일 선택
                    latest_file = sorted(bank_files)[-1]
                    file_path = os.path.join(self.config.output_dir, latest_file)
                    
                    # 재무 데이터 추출
                    financial_data = self.extract_financial_data_from_excel(file_path, bank)
                    
                    if financial_data:
                        all_financial_data.append(financial_data)
                        self.logger.log_message(f"{bank} 은행 재무 데이터 추출 완료")
                    else:
                        # 데이터가 없는 경우에도 은행명만 포함
                        all_financial_data.append(self._create_empty_financial_data(bank))
                else:
                    # 파일이 없는 경우
                    all_financial_data.append(self._create_empty_financial_data(bank, file_missing=True))
            
            # 나머지 처리는 _process_financial_data 메서드로 위임
            return self._process_financial_data(all_financial_data, self.config.output_dir)
            
        except Exception as e:
            self.logger.log_message(f"재무 데이터 통합 보고서 생성 오류: {str(e)}")
            import traceback
            self.logger.log_message(traceback.format_exc())
            return None, None
    
    def _create_empty_financial_data(self, bank_name, file_missing=False):
        """빈 재무 데이터 구조 생성"""
        status = '파일 없음' if file_missing else '데이터 없음'
        
        return {
            '은행명': bank_name,
            '재무정보 날짜': status,
            '분기': status,
            
            # 당기 데이터
            '당기_총자산': None,
            '당기_자기자본': None,
            '당기_총여신': None,
            '당기_총수신': None,
            '당기_수익합계': None,
            '당기_비용합계': None,
            '당기_당기순이익': None,
            '당기_고정이하여신비율(%)': None,
            '당기_위험가중자산에_대한_자기자본비율(%)': None,
            
            # 전년동기 데이터
            '전년동기_총자산': None,
            '전년동기_자기자본': None,
            '전년동기_총여신': None,
            '전년동기_총수신': None,
            '전년동기_수익합계': None,
            '전년동기_비용합계': None,
            '전년동기_당기순이익': None,
            '전년동기_고정이하여신비율(%)': None,
            '전년동기_위험가중자산에_대한_자기자본비율(%)': None,
            
            # 증감 데이터
            '증감_총자산': None,
            '증감_자기자본': None,
            '증감_총여신': None,
            '증감_총수신': None,
            '증감_수익합계': None,
            '증감_비용합계': None,
            '증감_당기순이익': None,
            '증감_고정이하여신비율(%)': None,
            '증감_위험가중자산에_대한_자기자본비율(%)': None
        }
    
    def create_consolidated_financial_report_from_folder(self, folder_path):
        """지정된 폴더에서 재무 데이터를 통합한 엑셀 보고서를 생성합니다."""
        try:
            self.logger.log_message(f"\n===== 재무 데이터 통합 보고서 생성 시작 (폴더: {folder_path}) =====")
            
            # 폴더 내 모든 엑셀 파일 찾기
            excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
            self.logger.log_message(f"발견된 엑셀 파일: {len(excel_files)}개")
            
            # 모든 은행의 재무 데이터를 저장할 리스트
            all_financial_data = []
            
            # 각 엑셀 파일에서 데이터 추출
            for excel_file in excel_files:
                file_path = os.path.join(folder_path, excel_file)
                
                # 파일명에서 은행명 추출 시도
                bank_name = "알 수 없음"
                for bank in self.config.BANKS:
                    if excel_file.startswith(f"{bank}_"):
                        bank_name = bank
                        break
                
                if bank_name == "알 수 없음":
                    # 파일명의 첫 부분을 은행명으로 사용
                    bank_name = excel_file.split('_')[0] if '_' in excel_file else excel_file.replace('.xlsx', '')
                
                # 재무 데이터 추출
                financial_data = self.extract_financial_data_from_excel(file_path, bank_name)
                
                if financial_data:
                    all_financial_data.append(financial_data)
                    self.logger.log_message(f"{bank_name} 은행 재무 데이터 추출 완료")
            
            # 나머지 처리는 기존과 동일
            return self._process_financial_data(all_financial_data, folder_path)
            
        except Exception as e:
            self.logger.log_message(f"재무 데이터 통합 보고서 생성 오류: {str(e)}")
            return None, None
    
    def create_consolidated_financial_report_from_files(self, file_list):
        """선택된 파일들에서 재무 데이터를 통합한 엑셀 보고서를 생성합니다."""
        try:
            self.logger.log_message(f"\n===== 재무 데이터 통합 보고서 생성 시작 (파일 {len(file_list)}개) =====")
            
            # 모든 은행의 재무 데이터를 저장할 리스트
            all_financial_data = []
            
            # 각 선택된 파일에서 데이터 추출
            for file_path in file_list:
                excel_file = os.path.basename(file_path)
                
                # 파일명에서 은행명 추출 시도
                bank_name = "알 수 없음"
                for bank in self.config.BANKS:
                    if excel_file.startswith(f"{bank}_"):
                        bank_name = bank
                        break
                
                if bank_name == "알 수 없음":
                    # 파일명의 첫 부분을 은행명으로 사용
                    bank_name = excel_file.split('_')[0] if '_' in excel_file else excel_file.replace('.xlsx', '')
                
                # 재무 데이터 추출
                financial_data = self.extract_financial_data_from_excel(file_path, bank_name)
                
                if financial_data:
                    all_financial_data.append(financial_data)
                    self.logger.log_message(f"{bank_name} 은행 재무 데이터 추출 완료 ({excel_file})")
            
            # 출력 폴더 결정 (첫 번째 파일의 폴더 사용)
            output_folder = os.path.dirname(file_list[0]) if file_list else self.config.output_dir
            
            # 나머지 처리는 기존과 동일
            return self._process_financial_data(all_financial_data, output_folder)
            
        except Exception as e:
            self.logger.log_message(f"재무 데이터 통합 보고서 생성 오류: {str(e)}")
            return None, None
    
    def create_consolidated_financial_report_md(self, all_financial_data, output_folder, is_settlement=False):
        """통합 재무 데이터를 마크다운 파일로 저장합니다. (당기/전년동기 비교 포함)"""
        try:
            file_suffix = "결산" if is_settlement else "분기"
            md_file = os.path.join(output_folder, f'저축은행_{file_suffix}_재무데이터_통합_비교_{self.config.today}.md')
            
            # DataFrame 생성 - 확장된 구조
            consolidated_df = self._create_consolidated_dataframe(all_financial_data)
            
            with open(md_file, 'w', encoding='utf-8') as f:
                # 헤더
                f.write(f"# 저축은행 {file_suffix}공시 재무데이터 통합 비교 보고서\n\n")
                f.write(f"## 요약 정보\n\n")
                
                # 통계 계산
                total_banks = len(consolidated_df)
                banks_with_current_data = len(consolidated_df[consolidated_df['당기_총자산'].notna()])
                banks_with_previous_data = len(consolidated_df[consolidated_df['전년동기_총자산'].notna()])
                banks_with_both_data = len(consolidated_df[
                    (consolidated_df['당기_총자산'].notna()) & 
                    (consolidated_df['전년동기_총자산'].notna())
                ])
                
                f.write(f"- **보고서 생성일**: {datetime.now().strftime('%Y년 %m월 %d일 %H:%M:%S')}\n")
                f.write(f"- **전체 은행 수**: {total_banks}개\n")
                f.write(f"- **당기 데이터 보유 은행**: {banks_with_current_data}개\n")
                f.write(f"- **전년동기 데이터 보유 은행**: {banks_with_previous_data}개\n")
                f.write(f"- **비교 분석 가능 은행**: {banks_with_both_data}개\n")
                
                # 평균 지표
                if '당기_위험가중자산에_대한_자기자본비율(%)' in consolidated_df.columns:
                    avg_bis_current = consolidated_df['당기_위험가중자산에_대한_자기자본비율(%)'].mean()
                    avg_bis_previous = consolidated_df['전년동기_위험가중자산에_대한_자기자본비율(%)'].mean()
                    if not pd.isna(avg_bis_current):
                        f.write(f"- **당기 평균 자기자본비율**: {avg_bis_current:.2f}%\n")
                    if not pd.isna(avg_bis_previous):
                        f.write(f"- **전년동기 평균 자기자본비율**: {avg_bis_previous:.2f}%\n")
                
                f.write(f"\n## 📊 상위 성과 은행 (당기 기준)\n\n")
                
                # 총자산 상위 5개 은행
                if '당기_총자산' in consolidated_df.columns:
                    asset_data = consolidated_df[consolidated_df['당기_총자산'].notna()]
                    if not asset_data.empty:
                        asset_sorted = asset_data.sort_values('당기_총자산', ascending=False)
                        
                        f.write("### 총자산 상위 5개 은행\n\n")
                        f.write("| 순위 | 은행명 | 당기 총자산 | 전년동기 총자산 | 증감 |\n")
                        f.write("| --- | --- | --- | --- | --- |\n")
                        for i, (_, row) in enumerate(asset_sorted.head(5).iterrows()):
                            current_asset = f"{int(row['당기_총자산']):,}" if pd.notna(row['당기_총자산']) else 'N/A'
                            previous_asset = f"{int(row['전년동기_총자산']):,}" if pd.notna(row['전년동기_총자산']) else 'N/A'
                            
                            # 증감 표시
                            change_info = "N/A"
                            if pd.notna(row['당기_총자산']) and pd.notna(row['전년동기_총자산']):
                                change = row['당기_총자산'] - row['전년동기_총자산']
                                change_rate = (change / row['전년동기_총자산']) * 100
                                change_symbol = "📈" if change > 0 else "📉" if change < 0 else "➡️"
                                change_info = f"{change_symbol} {change_rate:+.1f}%"
                            
                            f.write(f"| {i+1} | {row['은행명']} | {current_asset} | {previous_asset} | {change_info} |\n")
                        f.write("\n")
                
                # 자기자본비율 상위 5개 은행
                if '당기_위험가중자산에_대한_자기자본비율(%)' in consolidated_df.columns:
                    bis_data = consolidated_df[consolidated_df['당기_위험가중자산에_대한_자기자본비율(%)'].notna()]
                    if not bis_data.empty:
                        bis_sorted = bis_data.sort_values('당기_위험가중자산에_대한_자기자본비율(%)', ascending=False)
                        
                        f.write("### 자기자본비율 상위 5개 은행\n\n")
                        f.write("| 순위 | 은행명 | 당기 BIS비율 | 전년동기 BIS비율 | 변동 |\n")
                        f.write("| --- | --- | --- | --- | --- |\n")
                        for i, (_, row) in enumerate(bis_sorted.head(5).iterrows()):
                            current_bis = f"{row['당기_위험가중자산에_대한_자기자본비율(%)']:.2f}%" if pd.notna(row['당기_위험가중자산에_대한_자기자본비율(%)']) else 'N/A'
                            previous_bis = f"{row['전년동기_위험가중자산에_대한_자기자본비율(%)']:.2f}%" if pd.notna(row['전년동기_위험가중자산에_대한_자기자본비율(%)']) else 'N/A'
                            
                            # 변동 표시
                            change_info = "N/A"
                            if pd.notna(row['당기_위험가중자산에_대한_자기자본비율(%)']) and pd.notna(row['전년동기_위험가중자산에_대한_자기자본비율(%)']):
                                change = row['당기_위험가중자산에_대한_자기자본비율(%)'] - row['전년동기_위험가중자산에_대한_자기자본비율(%)']
                                change_symbol = "📈" if change > 0 else "📉" if change < 0 else "➡️"
                                change_info = f"{change_symbol} {change:+.2f}%p"
                            
                            f.write(f"| {i+1} | {row['은행명']} | {current_bis} | {previous_bis} | {change_info} |\n")
                        f.write("\n")
                
                f.write(f"## 📈 전체 재무현황 비교표\n\n")
                f.write("*주요 항목만 표시됩니다. 전체 데이터는 엑셀 파일을 확인하세요.*\n\n")
                
                # 간소화된 마크다운 테이블 생성 (주요 항목만)
                key_columns = [
                    '은행명', '재무정보 날짜', '당기_총자산', '전년동기_총자산',
                    '당기_자기자본', '전년동기_자기자본',
                    '당기_위험가중자산에_대한_자기자본비율(%)', '전년동기_위험가중자산에_대한_자기자본비율(%)'
                ]
                
                # 존재하는 컬럼만 선택
                available_columns = [col for col in key_columns if col in consolidated_df.columns]
                display_df = consolidated_df[available_columns]
                
                # 컬럼명 한글화
                column_mapping = {
                    '은행명': '은행명',
                    '재무정보 날짜': '공시날짜',
                    '당기_총자산': '당기총자산',
                    '전년동기_총자산': '전년총자산',
                    '당기_자기자본': '당기자기자본',
                    '전년동기_자기자본': '전년자기자본',
                    '당기_위험가중자산에_대한_자기자본비율(%)': '당기BIS',
                    '전년동기_위험가중자산에_대한_자기자본비율(%)': '전년BIS'
                }
                
                headers = [column_mapping.get(col, col) for col in available_columns]
                header_line = '| ' + ' | '.join(headers) + ' |\n'
                separator_line = '|' + '|'.join([' --- ' for _ in headers]) + '|\n'
                f.write(header_line)
                f.write(separator_line)
                
                # 데이터 행 (최대 20개 은행만 표시)
                max_rows = min(20, len(display_df))
                for idx in range(max_rows):
                    row = display_df.iloc[idx]
                    row_data = []
                    for col, value in zip(available_columns, row):
                        if pd.isna(value):
                            formatted_value = ''
                        elif col in ['당기_총자산', '전년동기_총자산', '당기_자기자본', '전년동기_자기자본']:
                            # 숫자 포맷팅 (천단위 구분)
                            try:
                                formatted_value = f"{int(value):,}"
                            except:
                                formatted_value = str(value)
                        elif col in ['당기_위험가중자산에_대한_자기자본비율(%)', '전년동기_위험가중자산에_대한_자기자본비율(%)']:
                            # 소수점 2자리
                            try:
                                formatted_value = f"{float(value):.2f}%"
                            except:
                                formatted_value = str(value)
                        else:
                            formatted_value = str(value)
                        
                        # 파이프 문자 이스케이프
                        formatted_value = formatted_value.replace('|', '\\|')
                        row_data.append(formatted_value)
                    
                    f.write('| ' + ' | '.join(row_data) + ' |\n')
                
                if len(display_df) > 20:
                    f.write(f"\n*({len(display_df) - 20}개 은행 더 있음...)*\n")
                
                f.write('\n')
                
                # 분석 요약
                f.write("## 💡 주요 분석 결과\n\n")
                f.write("### 전체 시장 동향\n\n")
                
                # 총자산 증감 분석
                if banks_with_both_data > 0:
                    asset_growth_banks = len(consolidated_df[
                        (consolidated_df['당기_총자산'].notna()) & 
                        (consolidated_df['전년동기_총자산'].notna()) &
                        (consolidated_df['당기_총자산'] > consolidated_df['전년동기_총자산'])
                    ])
                    asset_growth_rate = (asset_growth_banks / banks_with_both_data) * 100
                    
                    f.write(f"- **총자산 증가 은행**: {asset_growth_banks}/{banks_with_both_data}개 ({asset_growth_rate:.1f}%)\n")
                    
                    # 자기자본 증감 분석
                    capital_growth_banks = len(consolidated_df[
                        (consolidated_df['당기_자기자본'].notna()) & 
                        (consolidated_df['전년동기_자기자본'].notna()) &
                        (consolidated_df['당기_자기자본'] > consolidated_df['전년동기_자기자본'])
                    ])
                    if capital_growth_banks > 0:
                        capital_growth_rate = (capital_growth_banks / banks_with_both_data) * 100
                        f.write(f"- **자기자본 증가 은행**: {capital_growth_banks}/{banks_with_both_data}개 ({capital_growth_rate:.1f}%)\n")
                
                f.write("\n### 권장사항\n\n")
                f.write("- 당기와 전년동기 데이터를 비교하여 각 은행의 성장성을 평가하세요.\n")
                f.write("- 자기자본비율 변동을 통해 재무 안정성 변화를 모니터링하세요.\n")
                f.write("- 총자산 대비 증가율이 높은 은행들의 성장 전략을 분석해보세요.\n\n")
                
                # 푸터
                f.write("---\n")
                f.write(f"*이 보고서는 {datetime.now().strftime('%Y년 %m월 %d일 %H:%M:%S')}에 자동 생성되었습니다.*\n")
                f.write(f"*전체 상세 데이터는 동일한 폴더의 엑셀 파일을 확인하세요.*\n")
            
            self.logger.log_message(f"MD 통합 비교 보고서 저장 완료: {md_file}")
            return md_file
            
        except Exception as e:
            self.logger.log_message(f"MD 통합 비교 보고서 생성 오류: {str(e)}")
            return None
    
    def _create_consolidated_dataframe(self, all_financial_data):
        """통합 재무 데이터를 DataFrame으로 변환"""
        try:
            # 증감 데이터를 평면화하여 컬럼으로 분리
            flattened_data = []
            
            for data in all_financial_data:
                flattened_row = {}
                
                for key, value in data.items():
                    if key.startswith('증감_') and isinstance(value, dict):
                        # 증감 데이터를 개별 컬럼으로 분리
                        base_key = key.replace('증감_', '')
                        flattened_row[f'{base_key}_절대증감'] = value.get('절대증감')
                        flattened_row[f'{base_key}_증감률(%)'] = value.get('증감률(%)')
                    else:
                        flattened_row[key] = value
                
                flattened_data.append(flattened_row)
            
            # DataFrame 생성
            consolidated_df = pd.DataFrame(flattened_data)
            
            # 컬럼 순서 정렬
            basic_columns = ['은행명', '재무정보 날짜', '분기']
            current_columns = [col for col in consolidated_df.columns if col.startswith('당기_')]
            previous_columns = [col for col in consolidated_df.columns if col.startswith('전년동기_')]
            change_columns = [col for col in consolidated_df.columns if ('절대증감' in col or '증감률' in col)]
            
            # 존재하는 컬럼만 선택
            column_order = []
            for col_list in [basic_columns, current_columns, previous_columns, change_columns]:
                column_order.extend([col for col in col_list if col in consolidated_df.columns])
            
            # 누락된 컬럼 추가
            remaining_columns = [col for col in consolidated_df.columns if col not in column_order]
            column_order.extend(remaining_columns)
            
            consolidated_df = consolidated_df[column_order]
            
            return consolidated_df
            
        except Exception as e:
            self.logger.log_message(f"DataFrame 생성 오류: {str(e)}")
            return pd.DataFrame(all_financial_data)
    
    def _process_financial_data(self, all_financial_data, output_folder):
        """수집된 재무 데이터를 처리하여 통합 보고서 생성 (공통 로직) - 확장된 버전"""
        try:
            # DataFrame 생성
            consolidated_df = self._create_consolidated_dataframe(all_financial_data)
            
            # 엑셀 파일로 저장
            output_file = os.path.join(output_folder, 
                                      f'저축은행_재무데이터_통합_비교_{self.config.today}.xlsx')
            
            # 엑셀 저장 (개선된 버전)
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # 전체 데이터 시트
                consolidated_df.to_excel(writer, sheet_name='전체_재무현황_비교', index=False)
                
                # 요약 통계 시트 생성
                self._create_summary_sheet(writer, consolidated_df)
                
                # 당기만 데이터 시트
                current_only_df = self._create_current_only_dataframe(consolidated_df)
                current_only_df.to_excel(writer, sheet_name='당기_재무현황', index=False)
                
                # 증감 분석 시트
                change_analysis_df = self._create_change_analysis_dataframe(consolidated_df)
                if not change_analysis_df.empty:
                    change_analysis_df.to_excel(writer, sheet_name='증감_분석', index=False)
                
                # 서식 설정
                self._format_excel_sheets(writer, consolidated_df)
            
            # MD 파일로도 저장
            md_output_file = self.create_consolidated_financial_report_md(
                all_financial_data, output_folder, is_settlement=False
            )
            
            # 통계 요약
            stats = {
                '처리된 파일': len(all_financial_data),
                '당기 데이터 있는 은행': len([d for d in all_financial_data if d.get('당기_총자산')]),
                '전년동기 데이터 있는 은행': len([d for d in all_financial_data if d.get('전년동기_총자산')]),
                '비교 분석 가능 은행': len([d for d in all_financial_data if d.get('당기_총자산') and d.get('전년동기_총자산')]),
                '평균 당기 자기자본비율': consolidated_df['당기_위험가중자산에_대한_자기자본비율(%)'].mean() if '당기_위험가중자산에_대한_자기자본비율(%)' in consolidated_df else 0,
                '평균 전년동기 자기자본비율': consolidated_df['전년동기_위험가중자산에_대한_자기자본비율(%)'].mean() if '전년동기_위험가중자산에_대한_자기자본비율(%)' in consolidated_df else 0
            }
            
            self.logger.log_message(f"\n통합 비교 보고서 저장 완료: {output_file}")
            if md_output_file:
                self.logger.log_message(f"MD 통합 비교 보고서 저장 완료: {md_output_file}")
            self.logger.log_message(f"처리된 파일: {stats['처리된 파일']}개")
            self.logger.log_message(f"비교 분석 가능 은행: {stats['비교 분석 가능 은행']}개")
            
            return output_file, consolidated_df
            
        except Exception as e:
            self.logger.log_message(f"재무 데이터 처리 오류: {str(e)}")
            import traceback
            self.logger.log_message(traceback.format_exc())
            return None, None
    
    def _create_summary_sheet(self, writer, consolidated_df):
        """요약 통계 시트 생성"""
        try:
            summary_data = []
            
            # 기본 통계
            total_banks = len(consolidated_df)
            banks_with_current = len(consolidated_df[consolidated_df['당기_총자산'].notna()])
            banks_with_previous = len(consolidated_df[consolidated_df['전년동기_총자산'].notna()])
            banks_with_both = len(consolidated_df[
                (consolidated_df['당기_총자산'].notna()) & 
                (consolidated_df['전년동기_총자산'].notna())
            ])
            
            summary_data.extend([
                ['항목', '값'],
                ['전체 은행 수', total_banks],
                ['당기 데이터 보유 은행', banks_with_current],
                ['전년동기 데이터 보유 은행', banks_with_previous],
                ['비교 분석 가능 은행', banks_with_both],
                ['', ''],
                ['평균 지표 (당기)', ''],
            ])
            
            # 평균 지표 계산
            if '당기_총자산' in consolidated_df.columns:
                avg_current_assets = consolidated_df['당기_총자산'].mean()
                if not pd.isna(avg_current_assets):
                    summary_data.append(['평균 총자산 (당기)', f'{avg_current_assets:,.0f}'])
            
            if '당기_위험가중자산에_대한_자기자본비율(%)' in consolidated_df.columns:
                avg_current_bis = consolidated_df['당기_위험가중자산에_대한_자기자본비율(%)'].mean()
                if not pd.isna(avg_current_bis):
                    summary_data.append(['평균 자기자본비율 (당기)', f'{avg_current_bis:.2f}%'])
            
            summary_data.extend([
                ['', ''],
                ['평균 지표 (전년동기)', ''],
            ])
            
            if '전년동기_총자산' in consolidated_df.columns:
                avg_previous_assets = consolidated_df['전년동기_총자산'].mean()
                if not pd.isna(avg_previous_assets):
                    summary_data.append(['평균 총자산 (전년동기)', f'{avg_previous_assets:,.0f}'])
            
            if '전년동기_위험가중자산에_대한_자기자본비율(%)' in consolidated_df.columns:
                avg_previous_bis = consolidated_df['전년동기_위험가중자산에_대한_자기자본비율(%)'].mean()
                if not pd.isna(avg_previous_bis):
                    summary_data.append(['평균 자기자본비율 (전년동기)', f'{avg_previous_bis:.2f}%'])
            
            # DataFrame으로 변환하여 저장
            summary_df = pd.DataFrame(summary_data[1:], columns=summary_data[0])
            summary_df.to_excel(writer, sheet_name='요약_통계', index=False)
            
        except Exception as e:
            self.logger.log_message(f"요약 시트 생성 오류: {str(e)}", verbose=False)
    
    def _create_current_only_dataframe(self, consolidated_df):
        """당기 데이터만 포함하는 DataFrame 생성"""
        try:
            # 당기 데이터 컬럼만 선택
            basic_columns = ['은행명', '재무정보 날짜', '분기']
            current_columns = [col for col in consolidated_df.columns if col.startswith('당기_')]
            
            # 컬럼명 변경 (당기_ 접두사 제거)
            current_df = consolidated_df[basic_columns + current_columns].copy()
            
            # 컬럼명 정리
            rename_mapping = {}
            for col in current_columns:
                new_name = col.replace('당기_', '')
                rename_mapping[col] = new_name
            
            current_df.rename(columns=rename_mapping, inplace=True)
            
            return current_df
            
        except Exception as e:
            self.logger.log_message(f"당기 데이터 DataFrame 생성 오류: {str(e)}", verbose=False)
            return pd.DataFrame()
    
    def _create_change_analysis_dataframe(self, consolidated_df):
        """증감 분석 DataFrame 생성"""
        try:
            # 비교 가능한 은행만 선택
            comparison_df = consolidated_df[
                (consolidated_df['당기_총자산'].notna()) & 
                (consolidated_df['전년동기_총자산'].notna())
            ].copy()
            
            if comparison_df.empty:
                return pd.DataFrame()
            
            # 증감 관련 컬럼만 선택
            basic_columns = ['은행명', '재무정보 날짜']
            change_columns = [col for col in comparison_df.columns if ('절대증감' in col or '증감률' in col)]
            
            result_columns = basic_columns + change_columns
            change_df = comparison_df[result_columns].copy()
            
            # 총자산 증감률로 정렬
            if '총자산_증감률(%)' in change_df.columns:
                change_df = change_df.sort_values('총자산_증감률(%)', ascending=False)
            
            return change_df
            
        except Exception as e:
            self.logger.log_message(f"증감 분석 DataFrame 생성 오류: {str(e)}", verbose=False)
            return pd.DataFrame()
    
    def _format_excel_sheets(self, writer, consolidated_df):
        """엑셀 시트들의 서식 설정"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            workbook = writer.book
            
            # 전체 데이터 시트 서식
            if '전체_재무현황_비교' in writer.sheets:
                worksheet = writer.sheets['전체_재무현황_비교']
                self._format_main_sheet(worksheet, consolidated_df)
            
            # 요약 통계 시트 서식
            if '요약_통계' in writer.sheets:
                worksheet = writer.sheets['요약_통계']
                self._format_summary_sheet(worksheet)
            
            # 당기 데이터 시트 서식
            if '당기_재무현황' in writer.sheets:
                worksheet = writer.sheets['당기_재무현황']
                self._format_current_sheet(worksheet)
            
            # 증감 분석 시트 서식
            if '증감_분석' in writer.sheets:
                worksheet = writer.sheets['증감_분석']
                self._format_change_sheet(worksheet)
                
        except Exception as e:
            self.logger.log_message(f"엑셀 서식 설정 오류: {str(e)}", verbose=False)
    
    def _format_main_sheet(self, worksheet, consolidated_df):
        """메인 시트 서식 설정"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # 헤더 스타일
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center")
            
            # 헤더 서식 적용
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
            
            # 컬럼 너비 자동 조정
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # 최대 50자로 제한
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # 숫자 포맷 설정
            self._apply_number_formats(worksheet, consolidated_df)
            
        except Exception as e:
            self.logger.log_message(f"메인 시트 서식 오류: {str(e)}", verbose=False)
    
    def _format_summary_sheet(self, worksheet):
        """요약 시트 서식 설정"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # 헤더 서식
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # 컬럼 너비 조정
            worksheet.column_dimensions['A'].width = 30
            worksheet.column_dimensions['B'].width = 20
            
        except Exception as e:
            self.logger.log_message(f"요약 시트 서식 오류: {str(e)}", verbose=False)
    
    def _format_current_sheet(self, worksheet):
        """당기 데이터 시트 서식 설정"""
        try:
            from openpyxl.styles import Font, PatternFill
            
            # 헤더 서식
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
        except Exception as e:
            self.logger.log_message(f"당기 시트 서식 오류: {str(e)}", verbose=False)
    
    def _format_change_sheet(self, worksheet):
        """증감 분석 시트 서식 설정"""
        try:
            from openpyxl.styles import Font, PatternFill
            
            # 헤더 서식
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
        except Exception as e:
            self.logger.log_message(f"증감 시트 서식 오류: {str(e)}", verbose=False)
    
    def _apply_number_formats(self, worksheet, consolidated_df):
        """숫자 포맷 적용"""
        try:
            # 컬럼 타입별 포맷 매핑
            for col_idx, column_name in enumerate(consolidated_df.columns, 1):
                column_letter = worksheet.cell(row=1, column=col_idx).column_letter
                
                # 금액 컬럼 (천원 단위, 천단위 구분)
                if any(keyword in column_name for keyword in ['총자산', '자기자본', '총여신', '총수신', '수익합계', '비용합계', '당기순이익', '절대증감']):
                    for row in range(2, len(consolidated_df) + 2):
                        cell = worksheet.cell(row=row, column=col_idx)
                        if cell.value is not None and str(cell.value).replace('.', '').replace('-', '').isdigit():
                            cell.number_format = '#,##0'
                
                # 비율 컬럼 (소수점 2자리)
                elif any(keyword in column_name for keyword in ['비율', '증감률']):
                    for row in range(2, len(consolidated_df) + 2):
                        cell = worksheet.cell(row=row, column=col_idx)
                        if cell.value is not None:
                            if '증감률' in column_name:
                                cell.number_format = '+0.00;-0.00;0.00'  # 증감률 표시
                            else:
                                cell.number_format = '0.00'
                                
        except Exception as e:
            self.logger.log_message(f"숫자 포맷 적용 오류: {str(e)}", verbose=False)
    
    def create_zip_file(self):
        """결과 디렉토리를 ZIP 파일로 압축합니다."""
        try:
            self.logger.log_message("\n데이터 압축 중...")
            zip_filename = os.path.join(os.path.dirname(self.config.output_dir), 
                                      f'저축은행_통일경영공시_데이터_{self.config.today}.zip')
            
            with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for root, dirs, files in os.walk(self.config.output_dir):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arcname = os.path.relpath(file_path, os.path.dirname(self.config.output_dir))
                        zipf.write(file_path, arcname)
            
            self.logger.log_message(f"압축 파일 생성 완료: {zip_filename}")
            return zip_filename
        except Exception as e:
            self.logger.log_message(f"압축 파일 생성 오류: {str(e)}")
            return None


# GUI 클래스 (탭용으로 수정된 버전)
class QuarterlyScraperTab:
    def __init__(self, parent):
        self.parent = parent
        self.frame = ttk.Frame(parent)  # 탭용 프레임 생성
        
        # 구성 설정
        self.config = Config()
        self.logger = Logger(self.config, self)
        self.progress_status = {}  # 은행별 진행 상태 저장
        
        # 실행 상태 변수
        self.running = False
        self.scraper = None
        self.driver_manager = None
        self.progress_manager = None
        
        # 자동 압축 변수 초기화
        self.auto_zip_var = tk.BooleanVar(value=self.config.auto_zip)
        
        # MD 생성 옵션 변수 초기화
        self.save_md_var = tk.BooleanVar(value=False)
        
        # 메인 프레임
        self.create_widgets()
        
        # 초기 은행 목록 로드
        self.load_bank_list()
        
        # 설정 로드
        self.load_settings()
    
    def create_widgets(self):
        """GUI 위젯을 생성합니다."""
        # 메인 프레임 설정 - self.frame을 부모로 사용
        self.main_frame = ttk.Frame(self.frame, padding="10")
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 상단 프레임 (설정)
        self.settings_frame = ttk.LabelFrame(self.main_frame, text="설정", padding="5")
        self.settings_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # 출력 디렉토리 선택
        ttk.Label(self.settings_frame, text="출력 디렉토리:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.output_dir_var = tk.StringVar(value=self.config.output_dir)
        ttk.Entry(self.settings_frame, textvariable=self.output_dir_var, width=50).grid(row=0, column=1, sticky=tk.W+tk.E, padx=5, pady=5)
        ttk.Button(self.settings_frame, text="찾아보기", command=self.browse_output_dir).grid(row=0, column=2, padx=5, pady=5)
        
        # ChromeDriver 경로 선택
        ttk.Label(self.settings_frame, text="ChromeDriver 경로:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        self.chrome_driver_path_var = tk.StringVar(value=self.config.chrome_driver_path if self.config.chrome_driver_path else "")
        ttk.Entry(self.settings_frame, textvariable=self.chrome_driver_path_var, width=50).grid(row=1, column=1, sticky=tk.W+tk.E, padx=5, pady=5)
        ttk.Button(self.settings_frame, text="찾아보기", command=self.browse_chrome_driver_path).grid(row=1, column=2, padx=5, pady=5)
        
        # 작업자 수 설정
        ttk.Label(self.settings_frame, text="병렬 작업자 수:").grid(row=2, column=0, sticky=tk.W, padx=5, pady=5)
        self.workers_var = tk.IntVar(value=self.config.MAX_WORKERS)
        worker_spinbox = ttk.Spinbox(self.settings_frame, from_=1, to=10, textvariable=self.workers_var, width=5)
        worker_spinbox.grid(row=2, column=1, sticky=tk.W, padx=5, pady=5)
        
        # 자동 압축 옵션
        self.auto_zip_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(self.settings_frame, text="스크래핑 완료 후 자동 압축", variable=self.auto_zip_var).grid(row=2, column=2, sticky=tk.W, padx=5, pady=5)
        
        # MD 생성 옵션 추가
        self.save_md_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(self.settings_frame, text="📝 MD 파일도 함께 생성", 
                       variable=self.save_md_var).grid(row=3, column=0, sticky=tk.W, padx=5, pady=5)
        
        # 중앙 프레임 (은행 선택)
        self.bank_frame = ttk.LabelFrame(self.main_frame, text="은행 선택", padding="5")
        self.bank_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 은행 선택 버튼 및 체크박스
        self.select_all_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(self.bank_frame, text="전체 선택", variable=self.select_all_var, command=self.toggle_all_banks).pack(anchor=tk.W, padx=5, pady=5)
        
        # 은행 목록 표시 영역 (스크롤 가능)
        self.bank_list_frame = ttk.Frame(self.bank_frame)
        self.bank_list_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 스크롤바
        scrollbar = ttk.Scrollbar(self.bank_list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 트리뷰 (은행 목록) - 수정된 부분
        self.bank_tree = ttk.Treeview(self.bank_list_frame, columns=("bank", "status"), show="headings", yscrollcommand=scrollbar.set)
        self.bank_tree.heading("bank", text="은행명")
        self.bank_tree.heading("status", text="상태")
        self.bank_tree.column("bank", width=200)
        self.bank_tree.column("status", width=100, anchor=tk.CENTER)
        self.bank_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar.config(command=self.bank_tree.yview)
        
        # 하단 프레임 (로그 및 버튼)
        self.bottom_frame = ttk.Frame(self.main_frame)
        self.bottom_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 로그 표시 영역
        log_frame = ttk.LabelFrame(self.bottom_frame, text="로그")
        log_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, wrap=tk.WORD, width=80, height=10)
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.log_text.config(state=tk.DISABLED)
        
        # 버튼 프레임 - 수정된 버튼 배치
        button_frame = ttk.Frame(self.main_frame)
        button_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # 첫 번째 줄 버튼들
        button_row1 = ttk.Frame(button_frame)
        button_row1.pack(fill=tk.X, pady=2)
        
        self.start_button = ttk.Button(button_row1, text="스크래핑 시작", command=self.start_scraping)
        self.start_button.pack(side=tk.LEFT, padx=5)
        
        self.stop_button = ttk.Button(button_row1, text="중지", command=self.stop_scraping, state=tk.DISABLED)
        self.stop_button.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(button_row1, text="설정 저장", command=self.save_settings).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_row1, text="결과 폴더 열기", command=self.open_output_folder).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_row1, text="요약 보고서 생성", command=self.generate_report).pack(side=tk.LEFT, padx=5)
        
        # 두 번째 줄 버튼들 - 확장된 기능들
        button_row2 = ttk.Frame(button_frame)
        button_row2.pack(fill=tk.X, pady=2)
        
        ttk.Button(button_row2, text="📊 통합 재무 보고서 (당기/전년 비교)", 
                   command=self.create_financial_consolidation_with_selection).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(button_row2, text="📝 MD 요약 보고서", 
                   command=self.generate_md_summary_report).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(button_row2, text="🗜️ 데이터 압축", command=self.compress_and_download).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_row2, text="🔄 진행 상태 초기화", command=self.reset_progress).pack(side=tk.LEFT, padx=5)
        
        # 상태바는 제거 (메인 윈도우에서 관리)
    
    # 나머지 메서드들은 동일하므로 생략...
    # (browse_chrome_driver_path, save_settings, load_bank_list, load_settings 등은 기존과 동일)
    
    def browse_chrome_driver_path(self):
        """ChromeDriver 파일을 선택합니다."""
        filetypes = []
        if sys.platform == 'win32':
            filetypes = [("ChromeDriver", "*.exe"), ("모든 파일", "*.*")]
        else:
            filetypes = [("ChromeDriver", "*"), ("모든 파일", "*.*")]
            
        file_path = filedialog.askopenfilename(
            title="ChromeDriver 선택",
            filetypes=filetypes
        )
        
        if file_path:
            self.chrome_driver_path_var.set(file_path)
            self.config.update_chrome_driver_path(file_path)
            self.logger.log_message(f"ChromeDriver 경로 변경: {file_path}")
    
    def save_settings(self):
        """설정을 저장합니다."""
        # 출력 디렉토리 업데이트
        try:
            self.config.update_output_dir(self.output_dir_var.get())
        except Exception as e:
            messagebox.showerror("오류", f"출력 디렉토리 설정 실패: {str(e)}")
            return
        
        # ChromeDriver 경로 업데이트
        chrome_driver_path = self.chrome_driver_path_var.get()
        if chrome_driver_path and os.path.exists(chrome_driver_path):
            self.config.update_chrome_driver_path(chrome_driver_path)
        elif not chrome_driver_path:
            self.config.update_chrome_driver_path(None)
        
        # 작업자 수 업데이트
        self.config.MAX_WORKERS = self.workers_var.get()
        
        # 자동 압축 옵션 업데이트
        self.config.update_auto_zip(self.auto_zip_var.get())
        
        # 설정 저장
        self.config.save_settings()
        
        messagebox.showinfo("설정", "설정이 저장되었습니다.")
        self.logger.log_message("설정 저장 완료")
    
    def load_bank_list(self):
        """은행 목록을 로드하고 UI에 표시합니다."""
        # 트리뷰 항목 삭제
        for item in self.bank_tree.get_children():
            self.bank_tree.delete(item)
        
        # 은행 추가 - v2.3에서 수정된 트리뷰 구조 사용
        for bank in self.config.BANKS:
            # insert 메서드에서 values 파라미터 사용
            self.bank_tree.insert("", tk.END, iid=bank, values=(bank, "대기 중"))
            self.progress_status[bank] = "대기 중"
        
        self.update_log("은행 목록 로드 완료")
    
    def load_settings(self):
        """설정을 로드합니다."""
        # 이미 완료된 은행의 상태 업데이트
        self.progress_manager = ProgressManager(self.config, self.logger)
        completed_banks = self.progress_manager.progress.get('completed', [])
        failed_banks = self.progress_manager.progress.get('failed', [])
        
        for bank in completed_banks:
            self.update_bank_status(bank, "완료")
        
        for bank in failed_banks:
            self.update_bank_status(bank, "실패")
        
        self.update_log(f"설정 로드 완료. 출력 디렉토리: {self.config.output_dir}")
        self.update_log(f"완료된 은행: {len(completed_banks)}개, 실패한 은행: {len(failed_banks)}개")
    
    def browse_output_dir(self):
        """출력 디렉토리를 선택합니다."""
        directory = filedialog.askdirectory(initialdir=self.config.output_dir)
        if directory:
            self.output_dir_var.set(directory)
            try:
                self.config.update_output_dir(directory)
                self.logger.log_message(f"출력 디렉토리 변경: {directory}")
            except Exception as e:
                messagebox.showerror("오류", f"디렉토리 변경 실패: {str(e)}")
    
    def toggle_all_banks(self):
        """모든 은행을 선택하거나 선택 해제합니다."""
        select_all = self.select_all_var.get()
        for item in self.bank_tree.get_children():
            if select_all:
                self.bank_tree.selection_add(item)
            else:
                self.bank_tree.selection_remove(item)
    
    def update_log(self, message):
        """로그 메시지를 추가합니다."""
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, f"{message}\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
        self.parent.update_idletasks()  # self.root → self.parent로 변경
    
    def update_bank_status(self, bank_name, status):
        """은행의 상태를 업데이트합니다."""
        # 상태 저장
        self.progress_status[bank_name] = status
        
        # UI 업데이트 - v2.3에서 수정된 트리뷰 구조 사용
        try:
            if self.bank_tree.exists(bank_name):
                self.bank_tree.item(bank_name, values=(bank_name, status))
                self.parent.update_idletasks()  # self.root → self.parent로 변경
        except Exception as e:
            print(f"상태 업데이트 오류: {e}")
    
    def update_progress_callback(self, bank_name, status):
        """스크래핑 진행 상태 콜백"""
        self.update_bank_status(bank_name, status)
    
    def start_scraping(self):
        """스크래핑을 시작합니다."""
        if self.running:
            messagebox.showwarning("경고", "이미 스크래핑이 실행 중입니다.")
            return
        
        # 설정 업데이트
        try:
            self.config.update_output_dir(self.output_dir_var.get())
        except Exception as e:
            messagebox.showerror("오류", f"출력 디렉토리 설정 실패: {str(e)}")
            return
            
        self.config.MAX_WORKERS = self.workers_var.get()
        
        chrome_driver_path = self.chrome_driver_path_var.get()
        if chrome_driver_path:
            if os.path.exists(chrome_driver_path):
                self.config.update_chrome_driver_path(chrome_driver_path)
            else:
                messagebox.showwarning("경고", f"지정한 ChromeDriver 경로가 존재하지 않습니다: {chrome_driver_path}\n자동으로 찾습니다.")
                self.config.update_chrome_driver_path(None)
        else:
            self.config.update_chrome_driver_path(None)
        
        # 선택된 은행 확인
        selected_banks = list(self.bank_tree.selection())
        if not selected_banks:
            messagebox.showwarning("경고", "스크래핑할 은행을 선택하세요.")
            return
        
        # UI 업데이트
        self.start_button.config(state=tk.DISABLED)
        self.stop_button.config(state=tk.NORMAL)
        self.running = True
        
        # 스크래핑 스레드 시작
        self.scraping_thread = threading.Thread(target=self.run_scraping, args=(selected_banks,))
        self.scraping_thread.daemon = True
        self.scraping_thread.start()
    
    def run_scraping(self, selected_banks):
        """별도 스레드에서 스크래핑을 실행합니다."""
        try:
            # 드라이버 및 진행 관리자 초기화
            self.driver_manager = DriverManager(self.config, self.logger)
            self.progress_manager = ProgressManager(self.config, self.logger)
            self.scraper = BankScraper(self.config, self.logger, self.driver_manager, self.progress_manager)
            
            # 드라이버 초기화
            self.driver_manager.initialize_drivers()
            
            # 스크래핑 실행
            self.logger.log_message(f"\n===== 저축은행 중앙회 통일경영공시 데이터 스크래핑 시작 [{self.config.today}] =====\n")
            self.logger.log_message(f"선택된 은행 수: {len(selected_banks)}")
            
            # 스크래핑 시작 전 상태 초기화
            for bank in selected_banks:
                self.update_bank_status(bank, "대기 중")
            
            # 스크래핑 실행 (MD 옵션 포함)
            start_time = time.time()
            save_md = self.save_md_var.get()  # MD 저장 옵션 확인
            results = self.scraper.process_banks(selected_banks, self.update_progress_callback, save_md)
            
            # 결과 처리
            successful_banks = [r[0] for r in results if r[1]]
            failed_banks = [r[0] for r in results if not r[1]]
            
            # 실행 시간 계산
            end_time = time.time()
            total_duration = end_time - start_time
            minutes, seconds = divmod(total_duration, 60)
            
            # 요약 메시지
            self.logger.log_message(f"\n스크래핑 완료!")
            self.logger.log_message(f"성공: {len(successful_banks)}개, 실패: {len(failed_banks)}개")
            self.logger.log_message(f"총 실행 시간: {int(minutes)}분 {int(seconds)}초")
            
            # 요약 보고서 생성 (MD 포함)
            summary_file, stats, _ = self.scraper.generate_summary_report()
            if save_md:
                md_summary_file = self.scraper.generate_summary_report_md()
                if md_summary_file:
                    self.logger.log_message(f"MD 요약 보고서 생성 완료: {md_summary_file}")

           # MD 파일 통합 추가
                consolidated_md = self.scraper.create_consolidated_md_report()
                if consolidated_md:
                    self.logger.log_message(f"📄 통합 MD 보고서 생성 완료: {consolidated_md}")         
            
            # UI 업데이트
            self.parent.after(0, self.on_scraping_complete)  # self.root → self.parent로 변경
            
        except Exception as e:
            self.logger.log_message(f"스크래핑 중 오류 발생: {str(e)}")
            import traceback
            self.logger.log_message(traceback.format_exc())
            
            # UI 업데이트
            self.parent.after(0, self.on_scraping_error)  # self.root → self.parent로 변경
        finally:
            # 드라이버 종료
            if self.driver_manager:
                self.driver_manager.close_all()
    
    def on_scraping_complete(self):
        """스크래핑 완료 후 UI 업데이트"""
        self.running = False
        self.start_button.config(state=tk.NORMAL)
        self.stop_button.config(state=tk.DISABLED)
        messagebox.showinfo("완료", "저축은행 데이터 스크래핑이 완료되었습니다.")
        
        # 자동 압축 옵션이 활성화된 경우
        if self.auto_zip_var.get():
            self.parent.after(500, self.compress_and_download)  # 약간 지연 후 압축 시작
    
    def on_scraping_error(self):
        """스크래핑 오류 발생 시 UI 업데이트"""
        self.running = False
        self.start_button.config(state=tk.NORMAL)
        self.stop_button.config(state=tk.DISABLED)
        messagebox.showerror("오류", "스크래핑 중 오류가 발생했습니다. 로그를 확인하세요.")
    
    def stop_scraping(self):
        """스크래핑을 중지합니다."""
        if not self.running:
            return
        
        if messagebox.askyesno("중지 확인", "스크래핑을 중지하시겠습니까? 현재 진행 중인 작업이 완료된 후 중지됩니다."):
            self.running = False
            
            # 드라이버 종료
            if self.driver_manager:
                self.driver_manager.close_all()
            
            self.start_button.config(state=tk.NORMAL)
            self.stop_button.config(state=tk.DISABLED)
    
    def open_output_folder(self):
        """출력 폴더를 엽니다."""
        output_dir = self.output_dir_var.get()
        if os.path.exists(output_dir):
            try:
                if sys.platform == 'win32':
                    os.startfile(output_dir)
                elif sys.platform == 'darwin':  # macOS
                    import subprocess
                    subprocess.Popen(['open', output_dir])
                else:  # Linux
                    import subprocess
                    subprocess.Popen(['xdg-open', output_dir])
            except Exception as e:
                messagebox.showerror("오류", f"폴더를 열 수 없습니다: {str(e)}")
        else:
            messagebox.showerror("오류", f"폴더가 존재하지 않습니다: {output_dir}")
    
    def generate_report(self):
        """요약 보고서를 생성합니다."""
        try:
            # 스크래퍼 초기화
            if self.scraper is None:
                self.progress_manager = ProgressManager(self.config, self.logger)
                self.scraper = BankScraper(self.config, self.logger, None, self.progress_manager)
            
            # 보고서 생성
            summary_file, stats, summary_df = self.scraper.generate_summary_report()
            
            if summary_file and os.path.exists(summary_file):
                messagebox.showinfo("완료", f"요약 보고서가 생성되었습니다: {summary_file}")
                
                # 요약 창 표시
                self.show_summary_window(stats, summary_df)
            else:
                messagebox.showerror("오류", "요약 보고서 생성에 실패했습니다.")
        
        except Exception as e:
            messagebox.showerror("오류", f"요약 보고서 생성 중 오류 발생: {str(e)}")
    
    def generate_md_summary_report(self):
        """MD 요약 보고서를 생성합니다."""
        try:
            if self.scraper is None:
                self.progress_manager = ProgressManager(self.config, self.logger)
                self.scraper = BankScraper(self.config, self.logger, None, self.progress_manager)
            
            md_summary_file = self.scraper.generate_summary_report_md()
            
            if md_summary_file and os.path.exists(md_summary_file):
                messagebox.showinfo("완료", f"📝 MD 요약 보고서가 생성되었습니다!\n\n{os.path.basename(md_summary_file)}")
                
                if messagebox.askyesno("파일 열기", "생성된 MD 파일을 열어보시겠습니까?"):
                    self.open_md_file(md_summary_file)
            else:
                messagebox.showerror("오류", "MD 요약 보고서 생성에 실패했습니다.")
        
        except Exception as e:
            messagebox.showerror("오류", f"MD 요약 보고서 생성 중 오류 발생: {str(e)}")

    def open_md_file(self, file_path):
        """마크다운 파일을 엽니다."""
        if os.path.exists(file_path):
            try:
                if sys.platform == 'win32':
                    os.startfile(file_path)
                elif sys.platform == 'darwin':  # macOS
                    import subprocess
                    subprocess.Popen(['open', file_path])
                else:  # Linux
                    import subprocess
                    subprocess.Popen(['xdg-open', file_path])
            except Exception as e:
                messagebox.showerror("오류", f"파일을 열 수 없습니다: {str(e)}")
        else:
            messagebox.showerror("오류", f"파일이 존재하지 않습니다: {file_path}")
    
    def show_summary_window(self, stats, summary_df):
        """요약 통계를 보여주는 창을 표시합니다."""
        if summary_df is None:
            return
        
        summary_window = tk.Toplevel(self.parent)  # self.root → self.parent로 변경
        summary_window.title("스크래핑 결과 요약")
        summary_window.geometry("600x400")
        
        # 통계 프레임
        stats_frame = ttk.LabelFrame(summary_window, text="통계")
        stats_frame.pack(fill=tk.X, padx=10, pady=10)
        
        row = 0
        for key, value in stats.items():
            ttk.Label(stats_frame, text=key).grid(row=row, column=0, sticky=tk.W, padx=5, pady=2)
            ttk.Label(stats_frame, text=str(value)).grid(row=row, column=1, sticky=tk.W, padx=5, pady=2)
            row += 1
        
        # 은행별 상태 프레임
        status_frame = ttk.LabelFrame(summary_window, text="은행별 상태")
        status_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 트리뷰
        tree = ttk.Treeview(status_frame, columns=list(summary_df.columns), show="headings")
        for col in summary_df.columns:
            tree.heading(col, text=col)
            tree.column(col, width=100)
        
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # 스크롤바
        scrollbar = ttk.Scrollbar(status_frame, orient=tk.VERTICAL, command=tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        tree.configure(yscrollcommand=scrollbar.set)
        
        # 데이터 추가
        for _, row in summary_df.iterrows():
            values = list(row.values)
            tree.insert("", tk.END, values=values)
        
        # 닫기 버튼
        ttk.Button(summary_window, text="닫기", command=summary_window.destroy).pack(pady=10)
    
    def create_financial_consolidation_with_selection(self):
        """재무 데이터 통합 보고서를 생성합니다. (폴더/파일 선택 가능) - 확장된 버전"""
        try:
            # 스크래퍼 초기화
            if self.scraper is None:
                self.progress_manager = ProgressManager(self.config, self.logger)
                self.scraper = BankScraper(self.config, self.logger, None, self.progress_manager)
            
            # 확장된 안내 메시지
            result = messagebox.askyesno(
                "📊 통합 재무 보고서 (당기/전년동기 비교)", 
                "당기와 전년동기 데이터를 비교하는 통합 재무 보고서를 생성하시겠습니까?\n\n"
                "✨ 새로운 기능:\n"
                "• 당기와 전년동기 데이터 동시 표시\n"
                "• 증감률 및 절대증감 자동 계산\n"
                "• 엑셀과 MD 파일 동시 생성\n"
                "• 다양한 분석 시트 제공\n\n"
                "데이터 소스를 선택하는 창이 나타납니다."
            )
            
            if not result:
                return
            
            # 데이터 소스 선택 대화상자 표시
            source_dialog = FinancialDataSourceDialog(self.parent, self.config)  # self.root → self.parent로 변경
            self.parent.wait_window(source_dialog.dialog)  # self.root → self.parent로 변경
            
            if source_dialog.result:
                source_type, source_path = source_dialog.result
                
                # 상태 업데이트
                self.update_log(f"📊 통합 재무 데이터 분석 시작: {source_type}")
                self.update_log(f"당기/전년동기 비교 분석 및 엑셀+MD 생성 중...")
                
                # 별도 스레드에서 실행
                threading.Thread(
                    target=self._run_financial_consolidation_flexible, 
                    args=(source_type, source_path),
                    daemon=True
                ).start()
            
        except Exception as e:
            messagebox.showerror("오류", f"통합 재무 보고서 생성 중 오류 발생: {str(e)}")
    
    def _run_financial_consolidation_flexible(self, source_type, source_path):
        """유연한 재무 데이터 통합 실행 - 확장된 버전"""
        try:
            # 소스 타입에 따른 처리
            if source_type == "default_folder":
                # 기존 방식: 기본 출력 폴더에서 모든 파일 읽기
                output_file, consolidated_df = self.scraper.create_consolidated_financial_report()
                
            elif source_type == "custom_folder":
                # 사용자 지정 폴더에서 파일 읽기
                output_file, consolidated_df = self.scraper.create_consolidated_financial_report_from_folder(source_path)
                
            elif source_type == "selected_files":
                # 선택한 파일들만 처리
                output_file, consolidated_df = self.scraper.create_consolidated_financial_report_from_files(source_path)
            
            if output_file and os.path.exists(output_file):
                # 성공 메시지 - 확장된 정보
                self.parent.after(0, lambda: self.update_log(f"✅ 통합 재무 보고서 생성 완료!"))
                self.parent.after(0, lambda: self.update_log(f"📊 엑셀 파일: {os.path.basename(output_file)}"))
                
                # MD 파일 확인
                md_file = output_file.replace('.xlsx', '_비교.md').replace('_통합_', '_통합_비교_')
                if os.path.exists(md_file):
                    self.parent.after(0, lambda: self.update_log(f"📝 MD 파일: {os.path.basename(md_file)}"))
                
                self.parent.after(0, lambda: messagebox.showinfo(
                    "✅ 완료", 
                    f"📊 당기/전년동기 비교 분석 보고서가 생성되었습니다!\n\n"
                    f"📁 엑셀 파일: {os.path.basename(output_file)}\n"
                    f"📝 MD 파일: {os.path.basename(md_file) if os.path.exists(md_file) else '생성 안됨'}\n\n"
                    f"💡 새로운 기능:\n"
                    f"• 당기와 전년동기 데이터 비교\n"
                    f"• 증감률 자동 계산\n"
                    f"• 다양한 분석 시트 제공\n"
                    f"• MD 형식 보고서 동시 생성"
                ))
                
                # 통합 결과 창 표시 - 확장된 기능
                if consolidated_df is not None:
                    self.parent.after(0, lambda: self.show_enhanced_financial_consolidation_window(consolidated_df, output_file))
            else:
                self.parent.after(0, lambda: messagebox.showerror("오류", "통합 재무 보고서 생성에 실패했습니다."))
                
        except Exception as e:
            self.parent.after(0, lambda: messagebox.showerror("오류", f"통합 재무 보고서 생성 중 오류 발생: {str(e)}"))
    
    def show_enhanced_financial_consolidation_window(self, consolidated_df, output_file):
        """확장된 재무 데이터 통합 결과 창 표시"""
        if consolidated_df is None:
            return
        
        # 새 창 생성
        consolidation_window = tk.Toplevel(self.parent)
        consolidation_window.title("📊 당기/전년동기 비교 분석 결과")
        consolidation_window.geometry("1400x700")
        
        # 메인 프레임
        main_frame = ttk.Frame(consolidation_window, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 제목 프레임
        title_frame = ttk.Frame(main_frame)
        title_frame.pack(fill=tk.X, pady=5)
        
        title_label = ttk.Label(title_frame, text="📊 저축은행 당기/전년동기 비교 분석 결과", 
                               font=("", 14, "bold"))
        title_label.pack()
        
        # 요약 통계 프레임 - 확장된 버전
        stats_frame = ttk.LabelFrame(main_frame, text="📈 요약 통계", padding="10")
        stats_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # 통계 계산 - 확장된 버전
        total_banks = len(consolidated_df)
        banks_with_current = len(consolidated_df[consolidated_df['당기_총자산'].notna()])
        banks_with_previous = len(consolidated_df[consolidated_df['전년동기_총자산'].notna()])
        banks_with_both = len(consolidated_df[
            (consolidated_df['당기_총자산'].notna()) & 
            (consolidated_df['전년동기_총자산'].notna())
        ])
        
        # 성장률 분석
        asset_growth_banks = 0
        if banks_with_both > 0:
            asset_growth_banks = len(consolidated_df[
                (consolidated_df['당기_총자산'].notna()) & 
                (consolidated_df['전년동기_총자산'].notna()) &
                (consolidated_df['당기_총자산'] > consolidated_df['전년동기_총자산'])
            ])
        
        # 통계 표시 - 2열 레이아웃
        stats_left = ttk.Frame(stats_frame)
        stats_left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        stats_right = ttk.Frame(stats_frame)
        stats_right.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        
        # 왼쪽 통계
        ttk.Label(stats_left, text=f"📊 전체 은행 수: {total_banks}개", font=("", 10, "bold")).pack(anchor=tk.W)
        ttk.Label(stats_left, text=f"✅ 당기 데이터 보유: {banks_with_current}개").pack(anchor=tk.W)
        ttk.Label(stats_left, text=f"📅 전년동기 데이터 보유: {banks_with_previous}개").pack(anchor=tk.W)
        ttk.Label(stats_left, text=f"🔍 비교 분석 가능: {banks_with_both}개", font=("", 10, "bold")).pack(anchor=tk.W)
        
        # 오른쪽 통계
        if banks_with_both > 0:
            growth_rate = (asset_growth_banks / banks_with_both) * 100
            ttk.Label(stats_right, text=f"📈 총자산 증가 은행: {asset_growth_banks}개 ({growth_rate:.1f}%)", 
                     font=("", 10, "bold")).pack(anchor=tk.W)
        
        # 평균 BIS 비율
        if '당기_위험가중자산에_대한_자기자본비율(%)' in consolidated_df.columns:
            avg_bis_current = consolidated_df['당기_위험가중자산에_대한_자기자본비율(%)'].mean()
            avg_bis_previous = consolidated_df['전년동기_위험가중자산에_대한_자기자본비율(%)'].mean()
            if not pd.isna(avg_bis_current):
                ttk.Label(stats_right, text=f"💪 당기 평균 BIS 비율: {avg_bis_current:.2f}%").pack(anchor=tk.W)
            if not pd.isna(avg_bis_previous):
                ttk.Label(stats_right, text=f"📊 전년 평균 BIS 비율: {avg_bis_previous:.2f}%").pack(anchor=tk.W)
        
        # 탭 노트북 생성
        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # 탭 1: 요약 보기 (주요 지표만)
        summary_frame = ttk.Frame(notebook)
        notebook.add(summary_frame, text="📋 요약 보기")
        
        self._create_summary_tab(summary_frame, consolidated_df)
        
        # 탭 2: 상세 데이터
        detail_frame = ttk.Frame(notebook)
        notebook.add(detail_frame, text="📊 상세 데이터")
        
        self._create_detail_tab(detail_frame, consolidated_df)
        
        # 탭 3: 증감 분석
        if banks_with_both > 0:
            change_frame = ttk.Frame(notebook)
            notebook.add(change_frame, text="📈 증감 분석")
            
            self._create_change_tab(change_frame, consolidated_df)
        
        # 버튼 프레임
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(button_frame, text="📊 엑셀 파일 열기", 
                  command=lambda: self.open_excel_file(output_file)).pack(side=tk.LEFT, padx=5)
        
        # MD 파일 열기 버튼
        md_file = output_file.replace('.xlsx', '_비교.md').replace('_통합_', '_통합_비교_')
        if os.path.exists(md_file):
            ttk.Button(button_frame, text="📝 MD 파일 열기", 
                      command=lambda: self.open_md_file(md_file)).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(button_frame, text="📁 폴더 열기", 
                  command=lambda: self.open_output_folder()).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(button_frame, text="❌ 닫기", command=consolidation_window.destroy).pack(side=tk.RIGHT, padx=5)
    
    def _create_summary_tab(self, parent_frame, consolidated_df):
        """요약 탭 생성"""
        # 스크롤 가능한 프레임
        canvas = tk.Canvas(parent_frame)
        scrollbar = ttk.Scrollbar(parent_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # 주요 지표만 선택
        key_columns = ['은행명', '당기_총자산', '전년동기_총자산', '당기_자기자본', '전년동기_자기자본',
                      '당기_위험가중자산에_대한_자기자본비율(%)', '전년동기_위험가중자산에_대한_자기자본비율(%)']
        
        available_columns = [col for col in key_columns if col in consolidated_df.columns]
        display_df = consolidated_df[available_columns].head(20)  # 상위 20개만 표시
        
        # 트리뷰 생성
        tree = ttk.Treeview(scrollable_frame, columns=list(display_df.columns), show="headings", height=15)
        
        # 컬럼 설정
        column_widths = {
            '은행명': 100,
            '당기_총자산': 120,
            '전년동기_총자산': 120,
            '당기_자기자본': 120,
            '전년동기_자기자본': 120,
            '당기_위험가중자산에_대한_자기자본비율(%)': 150,
            '전년동기_위험가중자산에_대한_자기자본비율(%)': 150
        }
        
        for col in display_df.columns:
            # 컬럼명 단순화
            display_name = col.replace('당기_', '당기 ').replace('전년동기_', '전년 ').replace('위험가중자산에_대한_', '')
            tree.heading(col, text=display_name)
            width = column_widths.get(col, 100)
            tree.column(col, width=width, anchor=tk.CENTER)
        
        tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 데이터 추가
        for _, row in display_df.iterrows():
            values = []
            for col, value in zip(display_df.columns, row):
                if pd.isna(value):
                    formatted_value = ''
                elif col in ['당기_총자산', '전년동기_총자산', '당기_자기자본', '전년동기_자기자본']:
                    try:
                        formatted_value = f"{int(value):,}"
                    except:
                        formatted_value = str(value)
                elif col in ['당기_위험가중자산에_대한_자기자본비율(%)', '전년동기_위험가중자산에_대한_자기자본비율(%)']:
                    try:
                        formatted_value = f"{float(value):.2f}%"
                    except:
                        formatted_value = str(value)
                else:
                    formatted_value = str(value)
                values.append(formatted_value)
            
            tree.insert("", tk.END, values=values)
    
    def _create_detail_tab(self, parent_frame, consolidated_df):
        """상세 데이터 탭 생성"""
        # 트리뷰와 스크롤바
        tree_frame = ttk.Frame(parent_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        tree = ttk.Treeview(tree_frame, columns=list(consolidated_df.columns), show="headings")
        
        # 스크롤바
        v_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # 컬럼 설정
        for col in consolidated_df.columns:
            display_name = col.replace('당기_', '당기 ').replace('전년동기_', '전년 ').replace('위험가중자산에_대한_', '')
            tree.heading(col, text=display_name)
            tree.column(col, width=120, anchor=tk.CENTER)
        
        # 배치
        tree.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # 데이터 추가 (최대 50개)
        max_rows = min(50, len(consolidated_df))
        for idx in range(max_rows):
            row = consolidated_df.iloc[idx]
            values = []
            for col, value in zip(consolidated_df.columns, row):
                if pd.isna(value):
                    values.append('')
                elif isinstance(value, dict):  # 증감 데이터
                    if '절대증감' in value:
                        values.append(f"{value['절대증감']:+,.0f}")
                    elif '증감률(%)' in value:
                        values.append(f"{value['증감률(%)']:+.1f}%")
                    else:
                        values.append(str(value))
                elif col in ['당기_총자산', '전년동기_총자산', '당기_자기자본', '전년동기_자기자본']:
                    try:
                        values.append(f"{int(value):,}")
                    except:
                        values.append(str(value))
                elif col in ['당기_위험가중자산에_대한_자기자본비율(%)', '전년동기_위험가중자산에_대한_자기자본비율(%)']:
                    try:
                        values.append(f"{float(value):.2f}%")
                    except:
                        values.append(str(value))
                else:
                    values.append(str(value))
            
            tree.insert("", tk.END, values=values)
        
        if len(consolidated_df) > 50:
            info_label = ttk.Label(parent_frame, text=f"({len(consolidated_df) - 50}개 은행 더 있음... 전체는 엑셀 파일에서 확인하세요)")
            info_label.pack(pady=5)
    
    def _create_change_tab(self, parent_frame, consolidated_df):
        """증감 분석 탭 생성"""
        # 비교 가능한 은행만 선택
        comparison_df = consolidated_df[
            (consolidated_df['당기_총자산'].notna()) & 
            (consolidated_df['전년동기_총자산'].notna())
        ].copy()
        
        if comparison_df.empty:
            ttk.Label(parent_frame, text="비교 분석 가능한 데이터가 없습니다.", 
                     font=("", 12)).pack(expand=True)
            return
        
        # 증감률 계산 (임시)
        comparison_df['총자산_증감률'] = ((comparison_df['당기_총자산'] - comparison_df['전년동기_총자산']) / 
                                    comparison_df['전년동기_총자산'] * 100)
        
        if '당기_자기자본' in comparison_df.columns and '전년동기_자기자본' in comparison_df.columns:
            comparison_df['자기자본_증감률'] = ((comparison_df['당기_자기자본'] - comparison_df['전년동기_자기자본']) / 
                                         comparison_df['전년동기_자기자본'] * 100)
        
        # 총자산 증감률로 정렬
        comparison_df = comparison_df.sort_values('총자산_증감률', ascending=False)
        
        # 트리뷰 생성
        display_columns = ['은행명', '당기_총자산', '전년동기_총자산', '총자산_증감률']
        if '자기자본_증감률' in comparison_df.columns:
            display_columns.extend(['당기_자기자본', '전년동기_자기자본', '자기자본_증감률'])
        
        tree = ttk.Treeview(parent_frame, columns=display_columns, show="headings", height=20)
        
        # 컬럼 설정
        column_names = {
            '은행명': '은행명',
            '당기_총자산': '당기 총자산',
            '전년동기_총자산': '전년 총자산',
            '총자산_증감률': '총자산 증감률',
            '당기_자기자본': '당기 자기자본',
            '전년동기_자기자본': '전년 자기자본',
            '자기자본_증감률': '자기자본 증감률'
        }
        
        for col in display_columns:
            tree.heading(col, text=column_names.get(col, col))
            if '증감률' in col:
                tree.column(col, width=120, anchor=tk.CENTER)
            else:
                tree.column(col, width=150, anchor=tk.CENTER)
        
        tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 데이터 추가
        for _, row in comparison_df.head(30).iterrows():  # 상위 30개만
            values = []
            for col in display_columns:
                value = row.get(col, '')
                if pd.isna(value):
                    values.append('')
                elif '증감률' in col:
                    try:
                        symbol = "📈" if float(value) > 0 else "📉" if float(value) < 0 else "➡️"
                        values.append(f"{symbol} {float(value):+.1f}%")
                    except:
                        values.append(str(value))
                elif col in ['당기_총자산', '전년동기_총자산', '당기_자기자본', '전년동기_자기자본']:
                    try:
                        values.append(f"{int(value):,}")
                    except:
                        values.append(str(value))
                else:
                    values.append(str(value))
            
            tree.insert("", tk.END, values=values)
    
    def open_excel_file(self, file_path):
        """엑셀 파일을 엽니다."""
        if os.path.exists(file_path):
            try:
                if sys.platform == 'win32':
                    os.startfile(file_path)
                elif sys.platform == 'darwin':  # macOS
                    import subprocess
                    subprocess.Popen(['open', file_path])
                else:  # Linux
                    import subprocess
                    subprocess.Popen(['xdg-open', file_path])
            except Exception as e:
                messagebox.showerror("오류", f"파일을 열 수 없습니다: {str(e)}")
        else:
            messagebox.showerror("오류", f"파일이 존재하지 않습니다: {file_path}")
    
    def compress_and_download(self):
        """결과를 압축 파일로 만들고 원하는 위치에 저장합니다."""
        try:
            # 스크래퍼 초기화
            if self.scraper is None:
                self.progress_manager = ProgressManager(self.config, self.logger)
                self.scraper = BankScraper(self.config, self.logger, None, self.progress_manager)
            
            # 기본 저장 경로 및 파일명 설정
            default_filename = f'저축은행_통일경영공시_데이터_{self.config.today}.zip'
            default_dir = os.path.dirname(self.config.output_dir)
            
            # 파일 저장 대화상자
            save_path = filedialog.asksaveasfilename(
                title="압축 파일 저장",
                initialdir=default_dir,
                initialfile=default_filename,
                defaultextension=".zip",
                filetypes=[("ZIP 파일", "*.zip"), ("모든 파일", "*.*")]
            )
            
            if not save_path:
                return  # 사용자가 취소함
            
            # 압축 작업은 별도 스레드에서 실행 (UI 응답성 유지)
            threading.Thread(target=self._create_zip_file, args=(save_path,), daemon=True).start()
            
        except Exception as e:
            messagebox.showerror("오류", f"압축 파일 생성 중 오류 발생: {str(e)}")
    
    def _create_zip_file(self, save_path):
        """별도 스레드에서 실행되는 압축 파일 생성 함수"""
        try:
            # 압축 작업 수행
            with zipfile.ZipFile(save_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                # 진행 상황 모니터링 변수
                total_files = 0
                for root, _, files in os.walk(self.config.output_dir):
                    total_files += len(files)
                
                files_processed = 0
                
                # 파일 압축
                for root, dirs, files in os.walk(self.config.output_dir):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arcname = os.path.relpath(file_path, os.path.dirname(self.config.output_dir))
                        zipf.write(file_path, arcname)
                        
                        # 진행 상황 업데이트
                        files_processed += 1
                        progress = int(files_processed / total_files * 100)
                        self.parent.after(0, lambda p=progress: self.update_log(f"압축 중... {p}%"))
            
            # 완료 메시지
            self.parent.after(0, lambda: self.logger.log_message(f"압축 파일 생성 완료: {save_path}"))
            self.parent.after(0, lambda: messagebox.showinfo("완료", f"압축 파일이 생성되었습니다:\n{save_path}"))
            
            # 파일 탐색기에서 압축 파일 열기 (선택적)
            self.parent.after(0, lambda: self._open_file_location(save_path))
            
        except Exception as e:
            self.parent.after(0, lambda: self.logger.log_message(f"압축 파일 생성 오류: {str(e)}"))
            self.parent.after(0, lambda: messagebox.showerror("오류", f"압축 파일 생성 중 오류 발생: {str(e)}"))
    
    def _open_file_location(self, file_path):
        """파일이 위치한 폴더를 엽니다."""
        try:
            directory = os.path.dirname(file_path)
            if sys.platform == 'win32':
                os.startfile(directory)
            elif sys.platform == 'darwin':  # macOS
                import subprocess
                subprocess.Popen(['open', directory])
            else:  # Linux
                import subprocess
                subprocess.Popen(['xdg-open', directory])
        except Exception as e:
            self.logger.log_message(f"폴더 열기 실패: {str(e)}", verbose=False)
    
    def reset_progress(self):
        """진행 상태를 초기화합니다."""
        if messagebox.askyesno("초기화 확인", "진행 상태를 초기화하시겠습니까? 이 작업은 되돌릴 수 없습니다."):
            try:
                # 진행 관리자 초기화
                self.progress_manager = ProgressManager(self.config, self.logger)
                self.progress_manager.reset_progress()
                
                # 상태 업데이트
                for bank in self.config.BANKS:
                    self.update_bank_status(bank, "대기 중")
                
                messagebox.showinfo("완료", "진행 상태가 초기화되었습니다.")
                self.logger.log_message("진행 상태 초기화 완료")
            
            except Exception as e:
                messagebox.showerror("오류", f"진행 상태 초기화 중 오류 발생: {str(e)}")


# 메인 함수와 실행 코드는 주석 처리
# def main():
#     """프로그램의 메인 진입점"""
#     try:
#         # Tkinter 루트 윈도우 생성
#         root = tk.Tk()
#         
#         # 앱 인스턴스 생성
#         app = QuarterlyScraperTab(root)
#         
#         # 이벤트 루프 시작
#         root.mainloop()
#         
#     except Exception as e:
#         print(f"프로그램 실행 중 오류 발생: {str(e)}")
#         import traceback
#         traceback.print_exc()
# 
# 
# # 프로그램 진입점
# if __name__ == "__main__":
#     main()
