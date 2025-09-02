"""
저축은행 중앙회 결산공시 데이터 자동 스크래핑 도구
목적: 79개 저축은행의 결산 재무정보를 빠르고 효율적으로 스크래핑
작성일: 2025-01-15
버전: 2.6 (MD 통합 기능 추가)
개선사항:
- MD 파일 통합 재무보고서 기능 추가 (v2.6)
- 개별 MD 파일들을 하나로 통합하는 기능 추가 (v2.6)
- MD 파일 선택 및 통합을 위한 대화상자 추가 (v2.6)
- 당기/전년동기 비교 분석 기능 추가 (v2.5)
- 증감률 및 절대증감 자동 계산 기능 추가 (v2.5)
- MD 파일 당기/전년동기 비교 분석 보고서 생성 (v2.5)
- 확장된 엑셀 보고서 (4개 시트: 전체비교/요약통계/당기현황/증감분석) (v2.5)
- 시각적 증감 표시 (📈/📉/➡️) 및 순위 분석 기능 (v2.5)
- 결산공시 사이트 전용으로 수정 (v2.3)
- 재무데이터 통합 대화상자 버튼 표시 문제 수정 (v2.3)
- "친애" 은행을 "JT친애"로 수정 (v2.3)
- 날짜 추출 로직 개선 - 당기 데이터 우선 추출 (v2.3)
- 파일 경로 오류 수정 및 경로 처리 강화 (v2.2)
- 은행명 정확한 매칭 기능 강화 (v2.1)
- 재무 데이터 통합 보고서 기능 추가 (v2.1)
- 재무 데이터 통합 시 폴더/파일 선택 기능 추가 (v2.1)
- 자기자본/총자산 구분 추출 개선 (v2.1)
- MD 파일 생성 기능 추가 (v2.3-MD)
"""

import os
import sys
import time
import random
import json
import re
import concurrent.futures
import zipfile
from datetime import datetime
from io import StringIO
import io
from contextlib import contextmanager
from pathlib import Path
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from tqdm import tqdm
import selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
from bs4 import BeautifulSoup
import pandas as pd
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)  # 웹드라이버 관련 경고 무시
warnings.filterwarnings("ignore", category=UserWarning)  # 기타 경고 무시

# 표준 에러 출력을 억제하는 컨텍스트 매니저
@contextmanager
def suppress_stderr():
    """표준 에러 출력을 임시로 억제합니다."""
    original_stderr = sys.stderr
    sys.stderr = io.StringIO()
    try:
        yield
    finally:
        sys.stderr = original_stderr

# 상수 및 기본 설정
class Config:
    """프로그램 설정을 관리하는 클래스"""
    # 기본 설정 값
    VERSION = "2.6"  # 버전 업데이트
    BASE_URL = "https://www.fsb.or.kr/busmagesett_0100.act"  # 결산공시 URL (변경됨)
    MAX_RETRIES = 2  # 재시도 횟수
    PAGE_LOAD_TIMEOUT = 8  # 페이지 로드 타임아웃
    WAIT_TIMEOUT = 4  # 대기 시간
    MAX_WORKERS = 3  # 병렬 처리 워커 수 (기본값 감소)
    
    # 전체 79개 저축은행 목록 (친애 → JT친애로 수정)
    BANKS = [
        "다올", "대신", "더케이", "민국", "바로", "스카이", "신한", "애큐온", "예가람", "웰컴",
        "유안타", "조은", "키움YES", "푸른", "하나", "DB", "HB", "JT", "JT친애", "KB",  # "친애" → "JT친애"로 변경
        "NH", "OK", "OSB", "SBI", "금화", "남양", "모아", "부림", "삼정", "상상인",
        "세람", "안국", "안양", "영진", "융창", "인성", "인천", "키움", "페퍼", "평택",
        "한국투자", "한화", "고려", "국제", "동원제일", "솔브레인", "에스앤티", "우리", "조흥", "진주",
        "흥국", "BNK", "DH", "IBK", "대백", "대아", "대원", "드림", "라온", "머스트삼일",
        "엠에스", "오성", "유니온", "참", "CK", "대한", "더블", "동양", "삼호",
        "센트럴", "스마트", "스타", "대명", "상상인플러스", "아산", "오투", "우리금융", "청주", "한성"
    ]
    
    # 카테고리 목록 - 결산공시에 맞게 수정 (필요시 사이트 확인 후 변경)
    CATEGORIES = ["영업개황", "재무현황", "손익현황", "기타"]
    
    def __init__(self):
        """Config 초기화 - 경로 처리 개선"""
        self.today = datetime.now().strftime("%Y%m%d")
        
        # 설정 파일 경로 (먼저 설정)
        self.config_dir = os.path.join(os.path.expanduser("~"), ".bank_scraper")
        self.config_file = os.path.join(self.config_dir, "settings_settlement.json")  # 결산공시용 설정 파일
        
        # 기본값 설정
        self.chrome_driver_path = None
        self.auto_zip = True
        
        # 기본 출력 디렉토리 설정
        self.output_dir = self._get_default_output_dir()
        
        # 설정 로드 (있으면 덮어쓰기)
        self.load_settings()
        
        # 경로 정규화 및 검증
        self._validate_and_create_paths()
    
    def _get_default_output_dir(self):
        """기본 출력 디렉토리 경로 생성"""
        # 사용자 홈 디렉토리 사용
        base_dir = os.path.expanduser("~")
        # Downloads 폴더가 있으면 사용, 없으면 홈 디렉토리 사용
        downloads_dir = os.path.join(base_dir, "Downloads")
        if os.path.exists(downloads_dir):
            base_dir = downloads_dir
        
        # 오늘 날짜로 폴더명 생성 (결산공시용으로 변경)
        return os.path.join(base_dir, f"저축은행_결산공시_데이터_{self.today}")
    
    def _validate_and_create_paths(self):
        """모든 경로 검증 및 생성"""
        # 경로 정규화 (슬래시 통일)
        self.output_dir = os.path.normpath(self.output_dir)
        
        # 출력 디렉토리 생성
        try:
            os.makedirs(self.output_dir, exist_ok=True)
            print(f"✅ 출력 디렉토리 준비 완료: {self.output_dir}")
        except Exception as e:
            # 권한 문제 등으로 실패하면 임시 디렉토리 사용
            import tempfile
            self.output_dir = tempfile.mkdtemp(prefix=f"저축은행_결산공시_데이터_{self.today}_")
            print(f"⚠️ 기본 경로 생성 실패, 임시 경로 사용: {self.output_dir}")
        
        # 파일 경로 설정 (정규화된 경로 사용)
        self.progress_file = os.path.join(self.output_dir, 'bank_settlement_scraping_progress.json')
        self.log_file = os.path.join(self.output_dir, f'bank_settlement_scraping_log_{self.today}.txt')
    
    def update_output_dir(self, new_dir):
        """출력 디렉토리를 업데이트합니다."""
        # 경로 정규화
        new_dir = os.path.normpath(new_dir)
        
        # 디렉토리 존재 확인 및 생성
        try:
            os.makedirs(new_dir, exist_ok=True)
            self.output_dir = new_dir
            self.progress_file = os.path.join(self.output_dir, 'bank_settlement_scraping_progress.json')
            self.log_file = os.path.join(self.output_dir, f'bank_settlement_scraping_log_{self.today}.txt')
            self.save_settings()
            print(f"✅ 출력 디렉토리 변경 완료: {self.output_dir}")
        except Exception as e:
            print(f"❌ 출력 디렉토리 변경 실패: {str(e)}")
            raise
    
    def update_chrome_driver_path(self, new_path):
        """ChromeDriver 경로를 업데이트합니다."""
        if new_path:
            new_path = os.path.normpath(new_path)
        self.chrome_driver_path = new_path
        self.save_settings()
    
    def update_auto_zip(self, auto_zip):
        """자동 압축 옵션을 업데이트합니다."""
        self.auto_zip = auto_zip
        self.save_settings()
    
    def load_settings(self):
        """설정 파일에서 설정을 로드합니다."""
        if not os.path.exists(self.config_file):
            return
        
        try:
            with open(self.config_file, 'r', encoding='utf-8') as f:
                settings = json.load(f)
                
                # output_dir 처리 - 날짜 확인
                if 'output_dir' in settings:
                    saved_dir = settings['output_dir']
                    # 저장된 경로에 오늘 날짜가 포함되어 있는지 확인
                    if self.today in os.path.basename(saved_dir):
                        # 오늘 날짜가 포함된 경로면 그대로 사용
                        self.output_dir = os.path.normpath(saved_dir)
                    else:
                        # 날짜가 다르면 새 디렉토리 생성
                        print(f"📅 날짜가 변경되어 새 출력 폴더를 생성합니다.")
                        self.output_dir = self._get_default_output_dir()
                
                # 기타 설정 로드
                if 'chrome_driver_path' in settings:
                    path = settings['chrome_driver_path']
                    if path and os.path.exists(path):
                        self.chrome_driver_path = os.path.normpath(path)
                
                if 'max_workers' in settings:
                    self.MAX_WORKERS = int(settings['max_workers'])
                    
                if 'auto_zip' in settings:
                    self.auto_zip = bool(settings['auto_zip'])
                    
        except Exception as e:
            print(f"⚠️ 설정 파일 로드 중 오류 (기본값 사용): {str(e)}")
    
    def save_settings(self):
        """설정을 파일에 저장합니다."""
        try:
            os.makedirs(self.config_dir, exist_ok=True)
            
            settings = {
                'output_dir': self.output_dir,
                'chrome_driver_path': self.chrome_driver_path,
                'max_workers': self.MAX_WORKERS,
                'auto_zip': self.auto_zip,
                'last_saved': datetime.now().isoformat()
            }
            
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(settings, f, ensure_ascii=False, indent=2)
                
        except Exception as e:
            print(f"⚠️ 설정 저장 실패 (프로그램은 계속 실행됩니다): {str(e)}")


# 로그 관리 클래스
class Logger:
    def __init__(self, config, gui=None):
        """Logger 초기화 - 에러 처리 강화"""
        self.config = config
        self.gui = gui
        self.fallback_log_file = None  # 대체 로그 파일 경로
        
        # 로그 파일 초기화
        self._initialize_log_file()
    
    def _initialize_log_file(self):
        """로그 파일 초기화 - 실패 시 대체 경로 사용"""
        try:
            # 디렉토리 확인 및 생성
            log_dir = os.path.dirname(self.config.log_file)
            if log_dir and not os.path.exists(log_dir):
                os.makedirs(log_dir, exist_ok=True)
            
            # 로그 파일 생성/초기화
            with open(self.config.log_file, 'w', encoding='utf-8') as f:
                f.write(f"=== 저축은행 결산공시 스크래핑 로그 시작 ===\n")
                f.write(f"시작 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"출력 폴더: {self.config.output_dir}\n")
                f.write(f"=" * 50 + "\n\n")
            
            print(f"✅ 로그 파일 초기화 완료: {self.config.log_file}")
            
        except Exception as e:
            # 기본 경로 실패 시 대체 경로 사용
            print(f"⚠️ 기본 로그 파일 생성 실패: {str(e)}")
            
            # 대체 경로 시도 (현재 작업 디렉토리)
            self.fallback_log_file = os.path.join(
                os.getcwd(), 
                f'bank_settlement_scraping_log_{self.config.today}.txt'
            )
            
            try:
                with open(self.fallback_log_file, 'w', encoding='utf-8') as f:
                    f.write(f"=== 저축은행 결산공시 스크래핑 로그 (대체 경로) ===\n")
                    f.write(f"시작 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write(f"원래 경로: {self.config.log_file}\n")
                    f.write(f"대체 경로: {self.fallback_log_file}\n")
                    f.write(f"=" * 50 + "\n\n")
                
                print(f"✅ 대체 로그 파일 사용: {self.fallback_log_file}")
                
            except Exception as e2:
                # 모든 파일 쓰기 실패 시
                print(f"❌ 로그 파일 생성 완전 실패: {str(e2)}")
                print("⚠️ 로그가 콘솔에만 출력됩니다.")
    
    def log_message(self, message, print_to_console=True, verbose=True):
        """로그 메시지를 파일에 기록하고 필요한 경우 콘솔에 출력합니다."""
        if not verbose:
            return
        
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_entry = f"[{timestamp}] {message}"
        
        # 로그 파일에 기록 시도
        log_files = [self.config.log_file]
        if self.fallback_log_file:
            log_files.append(self.fallback_log_file)
        
        file_written = False
        for log_file in log_files:
            try:
                # 파일이 있는 디렉토리 확인
                log_dir = os.path.dirname(log_file)
                if log_dir and not os.path.exists(log_dir):
                    os.makedirs(log_dir, exist_ok=True)
                
                # 로그 추가
                with open(log_file, 'a', encoding='utf-8') as f:
                    f.write(log_entry + '\n')
                    f.flush()  # 즉시 디스크에 쓰기
                
                file_written = True
                break  # 성공하면 중단
                
            except Exception as e:
                continue  # 다음 파일 시도
        
        # GUI에 출력
        if self.gui:
            try:
                self.gui.update_log(message)
            except:
                pass  # GUI 업데이트 실패는 무시
        
        # 콘솔에 출력
        if print_to_console:
            if file_written:
                print(f"📝 {message}")
            else:
                print(f"⚠️ [로그 파일 쓰기 실패] {message}")
    
    def get_log_location(self):
        """현재 사용 중인 로그 파일 경로 반환"""
        if self.fallback_log_file and os.path.exists(self.fallback_log_file):
            return self.fallback_log_file
        return self.config.log_file


# 웹드라이버 관리 클래스
class DriverManager:
    def __init__(self, config, logger):
        self.config = config
        self.logger = logger
        self.max_drivers = config.MAX_WORKERS
        self.drivers = []
        self.available_drivers = []
        
    def initialize_drivers(self):
        """드라이버 풀 초기화"""
        self.logger.log_message(f"{self.max_drivers}개의 드라이버 초기화 중...")
        
        # 환경 변수 설정 (경고 메시지 억제)
        os.environ['WDM_LOG_LEVEL'] = '0'  # 로깅 레벨 최소화

        # stdout 및 stderr 임시 리다이렉션
        original_stdout = sys.stdout
        original_stderr = sys.stderr
        sys.stdout = io.StringIO()
        sys.stderr = io.StringIO()
        
        try:
            # 첫 번째 드라이버 생성 시도
            try:
                first_driver = self.create_driver()
                self.drivers.append(first_driver)
                self.available_drivers.append(first_driver)
                
                # 나머지 드라이버 생성
                for _ in range(self.max_drivers - 1):
                    driver = self.create_driver()
                    self.drivers.append(driver)
                    self.available_drivers.append(driver)
                    
            except Exception as e:
                self.logger.log_message(f"드라이버 초기화 중 오류 발생: {str(e)}")
                # 이미 생성된 드라이버가 있다면 모두 종료
                self.close_all()
                raise
        finally:
            # 원래 stdout 및 stderr 복원
            sys.stdout = original_stdout
            sys.stderr = original_stderr
    
    def create_driver(self):
        """최적화된 Chrome 웹드라이버를 생성합니다."""
        # stderr 출력 억제 (ChromeDriver 경고 메시지 숨기기)
        with suppress_stderr():
            options = webdriver.ChromeOptions()
            options.add_argument('--headless=new')
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('--disable-gpu')
            options.add_argument('--window-size=1280,800')
            options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36')
            
            # 로그 레벨 설정 (경고 숨기기)
            options.add_argument('--log-level=3')  # 오류만 표시
            options.add_argument('--silent')       # 메시지 억제
            
            # 최적화 옵션
            options.add_argument('--disable-extensions')
            options.add_argument('--disable-browser-side-navigation')
            options.add_argument('--disable-infobars')
            options.add_argument('--disable-notifications')
            options.add_argument('--disable-popup-blocking')
            
            # 이미지 로딩 활성화 (필요시)
            prefs = {
                'profile.default_content_setting_values': {
                    'images': 1,      # 이미지 로딩 활성화 (1=허용)
                    'plugins': 2,     # 플러그인 차단
                    'javascript': 1,  # JavaScript 허용 (필요)
                    'notifications': 2  # 알림 차단
                },
                'disk-cache-size': 4096,
            }
            options.add_experimental_option('prefs', prefs)
            
            # 브라우저 자동 종료 방지를 위한 경고 무시
            options.add_experimental_option("detach", True)
            options.add_experimental_option('excludeSwitches', ['enable-logging'])  # 콘솔 로깅 비활성화
            
            try:
                # 수동으로 지정된 ChromeDriver 경로 사용
                if self.config.chrome_driver_path and os.path.exists(self.config.chrome_driver_path):
                    from selenium.webdriver.chrome.service import Service
                    service = Service(executable_path=self.config.chrome_driver_path)
                    service.log_path = os.devnull  # 로그 비활성화
                    driver = webdriver.Chrome(service=service, options=options)
                    self.logger.log_message(f"지정된 ChromeDriver 사용: {self.config.chrome_driver_path}", verbose=False)
                else:
                    # ChromeDriver 자동 다운로드 오류 방지
                    from selenium.webdriver.chrome.service import Service
                    from webdriver_manager.chrome import ChromeDriverManager
                    
                    # 캐시 모드 설정 (오프라인 사용)
                    os.environ['WDM_LOCAL_CACHE_DIR'] = os.path.join(os.path.expanduser("~"), ".wdm", "drivers")
                    os.environ['WDM_OFFLINE'] = "true"  # 오프라인 모드로 설정
                    os.environ['WDM_LOG_LEVEL'] = '0'   # 로깅 레벨 최소화
                    
                    # 캐시된 드라이버 사용
                    service = Service(ChromeDriverManager().install())
                    
                    # 서비스 로그 수준 설정
                    service.log_path = os.devnull  # 로그 출력을 /dev/null로 리다이렉션
                    
                    driver = webdriver.Chrome(service=service, options=options)
            except Exception as e:
                self.logger.log_message(f"ChromeDriver 자동 설치 실패, 기본 방식으로 시도: {str(e)}", verbose=False)
                # 자동 설치 실패 시 기본 방식으로 시도
                driver = webdriver.Chrome(options=options)
            
            driver.set_page_load_timeout(self.config.PAGE_LOAD_TIMEOUT)
            return driver
    
    def get_driver(self):
        """사용 가능한 드라이버를 가져옵니다."""
        if not self.available_drivers:
            self.logger.log_message("모든 드라이버가 사용 중입니다. 대기 중...", verbose=False)
            time.sleep(1)  # 잠시 대기 후 재시도
            return self.get_driver()
        
        driver = self.available_drivers.pop(0)
        return driver
    
    def return_driver(self, driver):
        """드라이버를 풀에 반환합니다."""
        if driver in self.drivers and driver not in self.available_drivers:
            try:
                # 드라이버 상태 확인
                driver.current_url  # 접근 가능한지 확인
                self.available_drivers.append(driver)
            except:
                # 오류 발생 시 해당 드라이버를 종료하고 새 드라이버 생성
                try:
                    driver.quit()
                except:
                    pass
                
                self.drivers.remove(driver)
                new_driver = self.create_driver()
                self.drivers.append(new_driver)
                self.available_drivers.append(new_driver)
    
    def close_all(self):
        """모든 드라이버를 종료합니다."""
        for driver in self.drivers:
            try:
                driver.quit()
            except:
                pass
        self.drivers = []
        self.available_drivers = []


# 진행 상황 관리 클래스
class ProgressManager:
    def __init__(self, config, logger):
        self.config = config
        self.logger = logger
        self.file_path = config.progress_file
        self.progress = self.load()
    
    def load(self):
        """저장된 진행 상황을 로드합니다."""
        if os.path.exists(self.file_path):
            try:
                with open(self.file_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except json.JSONDecodeError:
                self.logger.log_message(f"진행 파일 손상: {self.file_path}, 새로 생성합니다.")
            except Exception as e:
                self.logger.log_message(f"진행 파일 로드 실패: {str(e)}")
        
        return {
            'completed': [],
            'failed': [],
            'stats': {
                'last_run': None,
                'success_count': 0,
                'failure_count': 0
            }
        }
    
    def is_completed(self, bank_name):
        """특정 은행의 스크래핑이 완료되었는지 확인합니다."""
        return bank_name in self.progress.get('completed', [])
    
    def mark_completed(self, bank_name):
        """은행을 완료 목록에 추가합니다."""
        if bank_name not in self.progress.get('completed', []):
            self.progress.setdefault('completed', []).append(bank_name)
            self.progress['stats']['success_count'] = len(self.progress.get('completed', []))
        
        # 실패 목록에서 제거 (재시도 후 성공한 경우)
        if bank_name in self.progress.get('failed', []):
            self.progress['failed'].remove(bank_name)
        
        self.save()
    
    def mark_failed(self, bank_name):
        """은행을 실패 목록에 추가합니다."""
        if bank_name not in self.progress.get('failed', []) and bank_name not in self.progress.get('completed', []):
            self.progress.setdefault('failed', []).append(bank_name)
            self.progress['stats']['failure_count'] = len(self.progress.get('failed', []))
            self.save()
    
    def save(self):
        """진행 상황을 파일에 저장합니다."""
        try:
            self.progress['stats']['last_run'] = datetime.now().isoformat()
            # 디렉토리 확인
            progress_dir = os.path.dirname(self.file_path)
            if progress_dir and not os.path.exists(progress_dir):
                os.makedirs(progress_dir, exist_ok=True)
            
            with open(self.file_path, 'w', encoding='utf-8') as f:
                json.dump(self.progress, f, ensure_ascii=False, indent=2)
        except Exception as e:
            self.logger.log_message(f"진행 상황 저장 실패: {str(e)}")
    
    def get_pending_banks(self, all_banks=None):
        """처리할 은행 목록을 반환합니다."""
        if all_banks is None:
            all_banks = self.config.BANKS
        completed = set(self.progress.get('completed', []))
        return [bank for bank in all_banks if bank not in completed]
    
    def reset_progress(self):
        """진행 상황을 초기화합니다."""
        self.progress = {
            'completed': [],
            'failed': [],
            'stats': {
                'last_run': None,
                'success_count': 0,
                'failure_count': 0
            }
        }
        self.save()


# 명시적 웨이팅 유틸리티 클래스
class WaitUtils:
    @staticmethod
    def wait_for_element(driver, locator, timeout):
        """요소가 나타날 때까지 명시적으로 대기합니다."""
        try:
            element = WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located(locator)
            )
            return element
        except TimeoutException:
            return None
    
    @staticmethod
    def wait_for_clickable(driver, locator, timeout):
        """요소가 클릭 가능할 때까지 명시적으로 대기합니다."""
        try:
            element = WebDriverWait(driver, timeout).until(
                EC.element_to_be_clickable(locator)
            )
            return element
        except TimeoutException:
            return None
    
    @staticmethod
    def wait_for_page_load(driver, timeout):
        """페이지가 완전히 로드될 때까지 대기합니다."""
        try:
            WebDriverWait(driver, timeout).until(
                lambda d: d.execute_script('return document.readyState') == 'complete'
            )
            return True
        except TimeoutException:
            return False
    
    @staticmethod
    def wait_with_random(min_time=0.3, max_time=0.7):
        """무작위 시간 동안 대기합니다."""
        time.sleep(random.uniform(min_time, max_time))


# 재무 데이터 소스 선택 대화상자 (수정된 버전)
class FinancialDataSourceDialog:
    """재무 데이터 소스 선택 대화상자"""
    
    def __init__(self, parent, config):
        self.parent = parent
        self.config = config
        self.result = None
        
        # 대화상자 창 생성
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("재무 데이터 소스 선택")
        # 창 크기를 늘려서 버튼이 보이도록 수정
        self.dialog.geometry("500x480")  # 높이를 480으로 증가
        self.dialog.resizable(False, False)
        
        # 모달 대화상자로 설정
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # 위젯 생성
        self.create_widgets()
        
        # 창 가운데 정렬
        self.center_window()
        
        # 키보드 단축키 바인딩
        self.dialog.bind('<Return>', lambda e: self.on_ok())
        self.dialog.bind('<Escape>', lambda e: self.on_cancel())
        
        # 창 크기 자동 조정
        self.dialog.update_idletasks()
        self.dialog.minsize(self.dialog.winfo_reqwidth(), self.dialog.winfo_reqheight())
    
    def create_widgets(self):
        """대화상자 위젯 생성"""
        # 메인 프레임
        main_frame = ttk.Frame(self.dialog, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 설명 라벨
        ttk.Label(main_frame, text="재무 데이터를 추출할 소스를 선택하세요:", 
                  font=("", 10, "bold")).pack(pady=(0, 15))
        
        # 라디오 버튼 변수
        self.source_var = tk.StringVar(value="default_folder")
        
        # 옵션 1: 기본 출력 폴더
        option1_frame = ttk.LabelFrame(main_frame, text="옵션 1: 기본 출력 폴더", padding="10")
        option1_frame.pack(fill=tk.X, pady=5)
        
        ttk.Radiobutton(option1_frame, text="현재 설정된 출력 폴더의 모든 엑셀 파일", 
                       variable=self.source_var, value="default_folder").pack(anchor=tk.W)
        ttk.Label(option1_frame, text=f"경로: {self.config.output_dir}", 
                 font=("", 9), foreground="gray").pack(anchor=tk.W, padx=20)
        
        # 옵션 2: 다른 폴더 선택
        option2_frame = ttk.LabelFrame(main_frame, text="옵션 2: 다른 폴더 선택", padding="10")
        option2_frame.pack(fill=tk.X, pady=5)
        
        ttk.Radiobutton(option2_frame, text="다른 폴더에서 엑셀 파일 읽기", 
                       variable=self.source_var, value="custom_folder").pack(anchor=tk.W)
        
        folder_frame = ttk.Frame(option2_frame)
        folder_frame.pack(fill=tk.X, padx=20, pady=5)
        
        self.folder_path_var = tk.StringVar()
        ttk.Entry(folder_frame, textvariable=self.folder_path_var, width=40).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(folder_frame, text="폴더 선택", command=self.browse_folder).pack(side=tk.LEFT)
        
        # 옵션 3: 개별 파일 선택
        option3_frame = ttk.LabelFrame(main_frame, text="옵션 3: 개별 파일 선택", padding="10")
        option3_frame.pack(fill=tk.X, pady=5)
        
        ttk.Radiobutton(option3_frame, text="특정 엑셀 파일들만 선택", 
                       variable=self.source_var, value="selected_files").pack(anchor=tk.W)
        
        files_frame = ttk.Frame(option3_frame)
        files_frame.pack(fill=tk.X, padx=20, pady=5)
        
        self.selected_files = []
        self.files_label = ttk.Label(files_frame, text="선택된 파일: 0개", foreground="gray")
        self.files_label.pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(files_frame, text="파일 선택", command=self.browse_files).pack(side=tk.LEFT)
        
        # 선택된 파일 목록 표시 (스크롤 가능)
        list_frame = ttk.Frame(option3_frame)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=20)
        
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.files_listbox = tk.Listbox(list_frame, height=3, yscrollcommand=scrollbar.set)
        self.files_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.files_listbox.yview)
        
        # 버튼 프레임 - 중앙 정렬로 변경
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, side=tk.BOTTOM, pady=(20, 10))
        
        # 버튼을 중앙에 배치
        button_container = ttk.Frame(button_frame)
        button_container.pack(expand=True)
        
        ttk.Button(button_container, text="확인", command=self.on_ok, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_container, text="취소", command=self.on_cancel, width=10).pack(side=tk.LEFT, padx=5)
    
    def browse_folder(self):
        """폴더 선택 대화상자"""
        folder = filedialog.askdirectory(
            title="엑셀 파일이 있는 폴더 선택",
            initialdir=self.config.output_dir
        )
        if folder:
            self.folder_path_var.set(folder)
            self.source_var.set("custom_folder")
    
    def browse_files(self):
        """파일 선택 대화상자"""
        files = filedialog.askopenfilenames(
            title="처리할 엑셀 파일 선택",
            initialdir=self.config.output_dir,
            filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")]
        )
        if files:
            self.selected_files = list(files)
            self.files_listbox.delete(0, tk.END)
            for file in self.selected_files:
                self.files_listbox.insert(tk.END, os.path.basename(file))
            self.files_label.config(text=f"선택된 파일: {len(self.selected_files)}개")
            self.source_var.set("selected_files")
    
    def on_ok(self):
        """확인 버튼 클릭"""
        source_type = self.source_var.get()
        
        if source_type == "default_folder":
            self.result = ("default_folder", self.config.output_dir)
            
        elif source_type == "custom_folder":
            folder_path = self.folder_path_var.get()
            if not folder_path:
                messagebox.showwarning("경고", "폴더를 선택해주세요.")
                return
            if not os.path.exists(folder_path):
                messagebox.showerror("오류", "선택한 폴더가 존재하지 않습니다.")
                return
            self.result = ("custom_folder", folder_path)
            
        elif source_type == "selected_files":
            if not self.selected_files:
                messagebox.showwarning("경고", "파일을 선택해주세요.")
                return
            self.result = ("selected_files", self.selected_files)
        
        self.dialog.destroy()
    
    def on_cancel(self):
        """취소 버튼 클릭"""
        self.result = None
        self.dialog.destroy()
    
    def center_window(self):
        """창을 화면 중앙에 배치"""
        self.dialog.update_idletasks()
        
        # 화면 크기 확인
        screen_width = self.dialog.winfo_screenwidth()
        screen_height = self.dialog.winfo_screenheight()
        
        # 창 크기 확인
        window_width = self.dialog.winfo_width()
        window_height = self.dialog.winfo_height()
        
        # 창이 화면보다 크면 조정
        if window_height > screen_height - 100:
            self.dialog.geometry(f"500x{screen_height - 100}")
            self.dialog.update_idletasks()
            window_height = self.dialog.winfo_height()
        
        x = (screen_width // 2) - (window_width // 2)
        y = (screen_height // 2) - (window_height // 2)
        self.dialog.geometry(f"+{x}+{y}")


# MD 파일 소스 선택 대화상자 (v2.6 - 새로 추가)
class MDSourceDialog:
    """MD 파일 소스 선택 대화상자"""
    
    def __init__(self, parent, config):
        self.parent = parent
        self.config = config
        self.result = None
        
        # 대화상자 창 생성
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("📝 MD 파일 통합 소스 선택")
        self.dialog.geometry("500x480")
        self.dialog.resizable(False, False)
        
        # 모달 대화상자로 설정
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # 위젯 생성
        self.create_widgets()
        
        # 창 가운데 정렬
        self.center_window()
        
        # 키보드 단축키 바인딩
        self.dialog.bind('<Return>', lambda e: self.on_ok())
        self.dialog.bind('<Escape>', lambda e: self.on_cancel())
    
    def create_widgets(self):
        """대화상자 위젯 생성"""
        # 메인 프레임
        main_frame = ttk.Frame(self.dialog, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 설명 라벨
        ttk.Label(main_frame, text="📝 MD 파일을 통합할 소스를 선택하세요:", 
                  font=("", 10, "bold")).pack(pady=(0, 15))
        
        # 라디오 버튼 변수
        self.source_var = tk.StringVar(value="default_folder")
        
        # 옵션 1: 기본 출력 폴더
        option1_frame = ttk.LabelFrame(main_frame, text="옵션 1: 기본 출력 폴더", padding="10")
        option1_frame.pack(fill=tk.X, pady=5)
        
        ttk.Radiobutton(option1_frame, text="현재 설정된 출력 폴더의 모든 MD 파일", 
                       variable=self.source_var, value="default_folder").pack(anchor=tk.W)
        ttk.Label(option1_frame, text=f"경로: {self.config.output_dir}", 
                 font=("", 9), foreground="gray").pack(anchor=tk.W, padx=20)
        
        # 옵션 2: 다른 폴더 선택
        option2_frame = ttk.LabelFrame(main_frame, text="옵션 2: 다른 폴더 선택", padding="10")
        option2_frame.pack(fill=tk.X, pady=5)
        
        ttk.Radiobutton(option2_frame, text="다른 폴더에서 MD 파일 읽기", 
                       variable=self.source_var, value="custom_folder").pack(anchor=tk.W)
        
        folder_frame = ttk.Frame(option2_frame)
        folder_frame.pack(fill=tk.X, padx=20, pady=5)
        
        self.folder_path_var = tk.StringVar()
        ttk.Entry(folder_frame, textvariable=self.folder_path_var, width=40).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(folder_frame, text="폴더 선택", command=self.browse_folder).pack(side=tk.LEFT)
        
        # 옵션 3: 개별 파일 선택
        option3_frame = ttk.LabelFrame(main_frame, text="옵션 3: 개별 파일 선택", padding="10")
        option3_frame.pack(fill=tk.X, pady=5)
        
        ttk.Radiobutton(option3_frame, text="특정 MD 파일들만 선택", 
                       variable=self.source_var, value="selected_files").pack(anchor=tk.W)
        
        files_frame = ttk.Frame(option3_frame)
        files_frame.pack(fill=tk.X, padx=20, pady=5)
        
        self.selected_files = []
        self.files_label = ttk.Label(files_frame, text="선택된 파일: 0개", foreground="gray")
        self.files_label.pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(files_frame, text="파일 선택", command=self.browse_files).pack(side=tk.LEFT)
        
        # 선택된 파일 목록 표시 (스크롤 가능)
        list_frame = ttk.Frame(option3_frame)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=20)
        
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.files_listbox = tk.Listbox(list_frame, height=3, yscrollcommand=scrollbar.set)
        self.files_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.files_listbox.yview)
        
        # 버튼 프레임
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, side=tk.BOTTOM, pady=(20, 10))
        
        # 버튼을 중앙에 배치
        button_container = ttk.Frame(button_frame)
        button_container.pack(expand=True)
        
        ttk.Button(button_container, text="확인", command=self.on_ok, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_container, text="취소", command=self.on_cancel, width=10).pack(side=tk.LEFT, padx=5)
    
    def browse_folder(self):
        """폴더 선택 대화상자"""
        folder = filedialog.askdirectory(
            title="MD 파일이 있는 폴더 선택",
            initialdir=self.config.output_dir
        )
        if folder:
            self.folder_path_var.set(folder)
            self.source_var.set("custom_folder")
    
    def browse_files(self):
        """파일 선택 대화상자"""
        files = filedialog.askopenfilenames(
            title="처리할 MD 파일 선택",
            initialdir=self.config.output_dir,
            filetypes=[("Markdown 파일", "*.md"), ("모든 파일", "*.*")]
        )
        if files:
            self.selected_files = list(files)
            self.files_listbox.delete(0, tk.END)
            for file in self.selected_files:
                self.files_listbox.insert(tk.END, os.path.basename(file))
            self.files_label.config(text=f"선택된 파일: {len(self.selected_files)}개")
            self.source_var.set("selected_files")
    
    def on_ok(self):
        """확인 버튼 클릭"""
        source_type = self.source_var.get()
        
        if source_type == "default_folder":
            self.result = ("default_folder", self.config.output_dir)
            
        elif source_type == "custom_folder":
            folder_path = self.folder_path_var.get()
            if not folder_path:
                messagebox.showwarning("경고", "폴더를 선택해주세요.")
                return
            if not os.path.exists(folder_path):
                messagebox.showerror("오류", "선택한 폴더가 존재하지 않습니다.")
                return
            self.result = ("custom_folder", folder_path)
            
        elif source_type == "selected_files":
            if not self.selected_files:
                messagebox.showwarning("경고", "파일을 선택해주세요.")
                return
            self.result = ("selected_files", self.selected_files)
        
        self.dialog.destroy()
    
    def on_cancel(self):
        """취소 버튼 클릭"""
        self.result = None
        self.dialog.destroy()
    
    def center_window(self):
        """창을 화면 중앙에 배치"""
        self.dialog.update_idletasks()
        
        # 화면 크기 확인
        screen_width = self.dialog.winfo_screenwidth()
        screen_height = self.dialog.winfo_screenheight()
        
        # 창 크기 확인
        window_width = self.dialog.winfo_width()
        window_height = self.dialog.winfo_height()
        
        # 창이 화면보다 크면 조정
        if window_height > screen_height - 100:
            self.dialog.geometry(f"500x{screen_height - 100}")
            self.dialog.update_idletasks()
            window_height = self.dialog.winfo_height()
        
        x = (screen_width // 2) - (window_width // 2)
        y = (screen_height // 2) - (window_height // 2)
        self.dialog.geometry(f"+{x}+{y}")


# 스크래퍼 클래스
class BankScraper:
    def __init__(self, config, logger, driver_manager, progress_manager):
        self.config = config
        self.logger = logger
        self.driver_manager = driver_manager
        self.progress_manager = progress_manager
    
    def extract_date_information(self, driver):
        """웹페이지에서 공시 날짜 정보를 추출합니다. (개선된 버전)"""
        try:
            # 방법 1: 당기 데이터 우선 찾기
            current_period_elements = driver.find_elements(
                By.XPATH, 
                "//*[contains(text(), '당기') and contains(text(), '년') and contains(text(), '월')]"
            )
            
            if current_period_elements:
                for element in current_period_elements:
                    text = element.text
                    # 정규식으로 날짜 패턴 추출
                    date_pattern = re.compile(r'\d{4}년\d{1,2}월말?')
                    matches = date_pattern.findall(text)
                    
                    if matches:
                        # 가장 최근 연도 찾기
                        latest_date = max(matches, key=lambda x: int(x[:4]))
                        self.logger.log_message(f"당기 날짜 발견: {latest_date}", verbose=False)
                        return latest_date
            
            # 방법 2: 모든 날짜를 찾아서 가장 최근 것 선택
            all_date_elements = driver.find_elements(
                By.XPATH, 
                "//*[contains(text(), '년') and contains(text(), '월')]"
            )
            
            all_dates = []
            for element in all_date_elements:
                text = element.text
                # 정규식 패턴 개선 (월말이 없는 경우도 포함)
                date_pattern = re.compile(r'\d{4}년\d{1,2}월말?')
                matches = date_pattern.findall(text)
                all_dates.extend(matches)
            
            if all_dates:
                # 중복 제거 및 정렬
                unique_dates = list(set(all_dates))
                # 연도 기준으로 정렬하여 가장 최근 날짜 선택
                sorted_dates = sorted(unique_dates, key=lambda x: int(x[:4]), reverse=True)
                
                # 2025년 데이터가 있으면 우선 선택
                for date in sorted_dates:
                    if "2025년" in date:
                        self.logger.log_message(f"최신 날짜 선택: {date}", verbose=False)
                        return date
                
                # 2025년이 없으면 가장 최근 날짜 반환
                return sorted_dates[0]
            
            # 방법 3: JavaScript로 직접 추출 (더 정확함)
            js_script = """
            var dates = [];
            var allText = document.body.innerText;
            
            // 당기 관련 텍스트 우선 찾기
            var currentPeriodMatch = allText.match(/당기[^\\n]*?(\\d{4}년\\d{1,2}월말?)/);
            if (currentPeriodMatch) {
                return currentPeriodMatch[1];
            }
            
            // 모든 날짜 찾아서 최신 것 반환
            var allMatches = allText.match(/\\d{4}년\\d{1,2}월말?/g);
            if (allMatches) {
                // 연도별로 정렬
                allMatches.sort(function(a, b) {
                    return parseInt(b.substr(0, 4)) - parseInt(a.substr(0, 4));
                });
                
                // 2025년 우선
                for (var i = 0; i < allMatches.length; i++) {
                    if (allMatches[i].includes('2025년')) {
                        return allMatches[i];
                    }
                }
                
                return allMatches[0];
            }
            
            return '';
            """
            
            date_text = driver.execute_script(js_script)
            if date_text:
                self.logger.log_message(f"JavaScript로 추출한 날짜: {date_text}", verbose=False)
                return date_text
            
            return "날짜 정보 없음"
            
        except Exception as e:
            self.logger.log_message(f"날짜 정보 추출 오류: {str(e)}", verbose=False)
            return "날짜 추출 실패"
    
    def select_bank(self, driver, bank_name):
        """다양한 방법으로 은행을 선택합니다. (정확한 매칭 우선)"""
        try:
            # 메인 페이지로 접속
            driver.get(self.config.BASE_URL)
            
            # 페이지 로딩 완료 대기
            WaitUtils.wait_for_page_load(driver, self.config.PAGE_LOAD_TIMEOUT)
            WaitUtils.wait_with_random(0.5, 1)
            
            # 특수 케이스 처리를 위한 정확한 은행명 목록 (JT친애 추가)
            exact_bank_names = {
                "키움": ["키움", "키움저축은행"],
                "키움YES": ["키움YES", "키움YES저축은행"],
                "JT": ["JT", "JT저축은행"],
                "JT친애": ["JT친애", "JT친애저축은행", "친애", "친애저축은행"],  # JT친애 매핑 추가
                "상상인": ["상상인", "상상인저축은행"],
                "상상인플러스": ["상상인플러스", "상상인플러스저축은행"],
                "머스트삼일": ["머스트삼일", "머스트삼일저축은행"]
            }
            
            # 검색할 은행명 목록 결정
            search_names = exact_bank_names.get(bank_name, [bank_name, f"{bank_name}저축은행"])
            
            # 방법 1: JavaScript로 정확한 은행명 매칭 (개선된 버전)
            js_script = f"""
            var targetBankNames = {json.dumps(search_names)};
            var found = false;
            
            // 모든 테이블 셀과 링크를 검사
            var allElements = document.querySelectorAll('td, a');
            
            for(var i = 0; i < allElements.length; i++) {{
                var element = allElements[i];
                var elementText = element.textContent.trim();
                
                // 정확한 매칭 확인
                for(var j = 0; j < targetBankNames.length; j++) {{
                    if(elementText === targetBankNames[j]) {{
                        // 키움/키움YES, JT/JT친애 구분을 위한 추가 검증
                        if('{bank_name}' === '키움' && elementText.includes('YES')) {{
                            continue;  // 키움을 찾는데 키움YES가 나오면 건너뛰기
                        }}
                        if('{bank_name}' === 'JT' && elementText.includes('친애')) {{
                            continue;  // JT를 찾는데 JT친애가 나오면 건너뛰기
                        }}
                        
                        element.scrollIntoView({{block: 'center'}});
                        
                        // 링크가 있으면 링크 클릭, 없으면 셀 클릭
                        if(element.tagName === 'A') {{
                            element.click();
                            found = true;
                            break;
                        }} else {{
                            var link = element.querySelector('a');
                            if(link) {{
                                link.click();
                                found = true;
                                break;
                            }} else {{
                                element.click();
                                found = true;
                                break;
                            }}
                        }}
                    }}
                }}
                if(found) break;
            }}
            
            return found ? "정확한 매칭 성공" : false;
            """
            
            result = driver.execute_script(js_script)
            if result:
                self.logger.log_message(f"{bank_name} 은행: {result}", verbose=False)
                WaitUtils.wait_with_random(1, 1.5)  # 페이지 전환 대기 시간 증가
                
                # 페이지 전환 확인
                current_url = driver.current_url
                if current_url != self.config.BASE_URL:
                    return True
            
            # 방법 2: XPath로 정확한 텍스트 매칭 (보완)
            for search_name in search_names:
                # 추가 조건으로 더 정확한 매칭
                if bank_name == "키움":
                    xpath = f"//td[normalize-space(text())='{search_name}' and not(contains(text(), 'YES'))]"
                elif bank_name == "JT":
                    xpath = f"//td[normalize-space(text())='{search_name}' and not(contains(text(), '친애'))]"
                else:
                    xpath = f"//td[normalize-space(text())='{search_name}']"
                
                bank_elements = driver.find_elements(By.XPATH, xpath)
                
                if bank_elements:
                    for element in bank_elements:
                        try:
                            if element.is_displayed():
                                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
                                WaitUtils.wait_with_random(0.3, 0.5)
                                driver.execute_script("arguments[0].click();", element)
                                WaitUtils.wait_with_random(1, 1.5)
                                
                                # 페이지 전환 확인
                                if driver.current_url != self.config.BASE_URL:
                                    return True
                        except:
                            continue
            
            self.logger.log_message(f"{bank_name} 은행을 찾을 수 없습니다.")
            return False
            
        except Exception as e:
            self.logger.log_message(f"{bank_name} 은행 선택 실패: {str(e)}")
            return False
    
    def select_category(self, driver, category):
        """특정 카테고리 탭을 클릭합니다."""
        try:
            # 방법 1: 정확한 텍스트 매칭
            tab_xpaths = [
                f"//a[normalize-space(text())='{category}']",
                f"//a[contains(@class, 'tab') and contains(text(), '{category}')]",
                f"//li[contains(@class, 'tab') and contains(text(), '{category}')]",
                f"//span[contains(text(), '{category}')]",
                f"//button[contains(text(), '{category}')]"
            ]
            
            for xpath in tab_xpaths:
                elements = driver.find_elements(By.XPATH, xpath)
                for element in elements:
                    try:
                        if element.is_displayed():
                            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
                            WaitUtils.wait_with_random(0.3, 0.5)
                            driver.execute_script("arguments[0].click();", element)
                            WaitUtils.wait_with_random(0.5, 1)
                            return True
                    except:
                        continue
            
            # 방법 2: JavaScript로 카테고리 탭 클릭
            category_indices = {
                "영업개황": 0,
                "재무현황": 1,
                "손익현황": 2,
                "기타": 3
            }
            
            if category in category_indices:
                idx = category_indices[category]
                script = f"""
                // 모든 탭 관련 요소 찾기
                var tabContainers = document.querySelectorAll('ul.tabs, div.tab-container, nav, .tab-list, ul, div[role="tablist"]');
                
                // 정확한 텍스트 매칭
                var allElements = document.querySelectorAll('a, button, span, li, div');
                for (var k = 0; k < allElements.length; k++) {{
                    if (allElements[k].innerText.trim() === '{category}') {{
                        allElements[k].scrollIntoView({{block: 'center'}});
                        allElements[k].click();
                        return "exact_match";
                    }}
                }}
                
                // 탭 컨테이너에서 인덱스 기반 검색
                for (var i = 0; i < tabContainers.length; i++) {{
                    var tabs = tabContainers[i].querySelectorAll('a, li, button, div[role="tab"], span');
                    
                    // 먼저 텍스트로 찾기
                    for (var j = 0; j < tabs.length; j++) {{
                        if (tabs[j].innerText.includes('{category}')) {{
                            tabs[j].scrollIntoView({{block: 'center'}});
                            tabs[j].click();
                            return "text_match_in_container";
                        }}
                    }}
                    
                    // 인덱스로 찾기
                    if (tabs.length >= {idx + 1}) {{
                        tabs[{idx}].scrollIntoView({{block: 'center'}});
                        tabs[{idx}].click();
                        return "index_match";
                    }}
                }}
                
                // 모든 클릭 가능 요소에서 포함 문자열 검색
                var clickables = document.querySelectorAll('a, button, span, div, li');
                for (var j = 0; j < clickables.length; j++) {{
                    if (clickables[j].innerText.includes('{category}')) {{
                        clickables[j].scrollIntoView({{block: 'center'}});
                        clickables[j].click();
                        return "contains_match";
                    }}
                }}
                
                return false;
                """
                
                result = driver.execute_script(script)
                if result:
                    self.logger.log_message(f"{category} 탭: {result} 성공", verbose=False)
                    WaitUtils.wait_with_random(0.5, 1)
                    return True
            
            # 방법 3: 포함 문자열로 검색 (더 관대한 매칭)
            tab_broad_xpath = f"//*[contains(text(), '{category}')]"
            elements = driver.find_elements(By.XPATH, tab_broad_xpath)
            
            for element in elements:
                try:
                    if element.is_displayed() and (element.tag_name in ['a', 'li', 'span', 'button', 'div']):
                        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
                        WaitUtils.wait_with_random(0.3, 0.5)
                        driver.execute_script("arguments[0].click();", element)
                        WaitUtils.wait_with_random(0.5, 1)
                        return True
                except:
                    continue
            
            # 방법 4: CSS 선택자 시도
            tab_css = f"[role='tab'], .tab, .tab-item, .tabs li, .tabs a, nav a, ul li a"
            tabs = driver.find_elements(By.CSS_SELECTOR, tab_css)
            
            for tab in tabs:
                try:
                    if category in tab.text and tab.is_displayed():
                        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", tab)
                        WaitUtils.wait_with_random(0.3, 0.5)
                        driver.execute_script("arguments[0].click();", tab)
                        WaitUtils.wait_with_random(0.5, 1)
                        return True
                except:
                    continue
            
            self.logger.log_message(f"{category} 탭을 찾을 수 없습니다.", verbose=False)
            return False
            
        except Exception as e:
            self.logger.log_message(f"{category} 탭 클릭 실패: {str(e)}", verbose=False)
            return False
    
    def extract_tables_from_page(self, driver):
        """현재 페이지에서 모든 테이블을 추출합니다."""
        try:
            # 페이지가 완전히 로드될 때까지 대기
            WaitUtils.wait_for_page_load(driver, self.config.PAGE_LOAD_TIMEOUT)
            WaitUtils.wait_with_random(0.5, 1)
            
            # 방법 1: pandas로 테이블 추출
            try:
                html_source = driver.page_source
                dfs = pd.read_html(StringIO(html_source))
                
                if dfs:
                    valid_dfs = []
                    seen_shapes = set()
                    
                    for df in dfs:
                        # 비어있지 않은 테이블만 선택
                        if not df.empty and df.shape[0] > 0 and df.shape[1] > 0:
                            # MultiIndex 컬럼 처리
                            if isinstance(df.columns, pd.MultiIndex):
                                new_cols = []
                                for col in df.columns:
                                    if isinstance(col, tuple):
                                        clean_col = [str(c).strip() for c in col if str(c).strip() and str(c).lower() != 'nan']
                                        new_cols.append('_'.join(clean_col) if clean_col else f"Column_{len(new_cols)+1}")
                                    else:
                                        new_cols.append(str(col))
                                df.columns = new_cols
                            
                            # 중복 테이블 제거
                            try:
                                # 테이블 해시 생성 (중복 확인용)
                                shape_hash = f"{df.shape}"
                                headers_hash = f"{list(df.columns)}"
                                data_hash = ""
                                if len(df) > 0:
                                    # 데이터 첫 행의 값들을 해시 생성
                                    data_hash = f"{list(df.iloc[0].astype(str))}"
                                
                                table_hash = f"{shape_hash}_{headers_hash}_{data_hash}"
                                
                                if table_hash not in seen_shapes:
                                    valid_dfs.append(df)
                                    seen_shapes.add(table_hash)
                            except:
                                # 해시 생성 실패 시 그냥 추가
                                valid_dfs.append(df)
                    
                    return valid_dfs
            except Exception as e:
                self.logger.log_message(f"pandas 테이블 추출 실패: {str(e)}", verbose=False)
            
            # 방법 2: BeautifulSoup으로 테이블 추출 (pandas 실패 시)
            try:
                soup = BeautifulSoup(driver.page_source, 'html.parser')
                tables = soup.find_all('table')
                
                extracted_dfs = []
                table_hashes = set()  # 중복 테이블 제거용
                
                for table in tables:
                    try:
                        # 테이블 구조 파악
                        headers = []
                        rows = []
                        
                        # 헤더 추출
                        th_elements = table.select('thead th') or table.select('tr:first-child th')
                        if th_elements:
                            headers = [th.get_text(strip=True) for th in th_elements]
                        
                        # 헤더가 없으면 첫 번째 행의 td를 헤더로 사용
                        if not headers:
                            first_row_tds = table.select('tr:first-child td')
                            if first_row_tds:
                                headers = [td.get_text(strip=True) or f"Column_{i+1}" for i, td in enumerate(first_row_tds)]
                        
                        # 헤더가 없으면 기본 열 이름 생성
                        if not headers:
                            max_cols = max([len(row.select('td')) for row in table.select('tr')], default=0)
                            headers = [f'Column_{j+1}' for j in range(max_cols)]
                        
                        # 데이터 행 추출
                        header_rows = table.select('thead tr')
                        for tr in table.select('tbody tr') or table.select('tr')[1:]:
                            if tr in header_rows:
                                continue
                            
                            cells = tr.select('td')
                            if cells:
                                row_data = [td.get_text(strip=True) for td in cells]
                                if row_data and len(row_data) > 0:
                                    rows.append(row_data)
                        
                        # 행 데이터가 있으면 DataFrame 생성
                        if rows and headers:
                            # 열 개수 맞추기
                            for i, row in enumerate(rows):
                                if len(row) < len(headers):
                                    rows[i] = row + [''] * (len(headers) - len(row))
                                elif len(row) > len(headers):
                                    rows[i] = row[:len(headers)]
                            
                            df = pd.DataFrame(rows, columns=headers)
                            
                            if not df.empty:
                                # 테이블 해시 생성 (중복 확인용)
                                try:
                                    table_hash = f"{df.shape}_{hash(str(df.iloc[0].values) if len(df) > 0 else '')}"
                                    if table_hash not in table_hashes:
                                        extracted_dfs.append(df)
                                        table_hashes.add(table_hash)
                                except:
                                    extracted_dfs.append(df)
                    except Exception as e:
                        self.logger.log_message(f"개별 테이블 파싱 실패: {str(e)}", verbose=False)
                
                if extracted_dfs:
                    return extracted_dfs
            except Exception as e:
                self.logger.log_message(f"BeautifulSoup 테이블 추출 실패: {str(e)}", verbose=False)
            
            return []
            
        except Exception as e:
            self.logger.log_message(f"테이블 추출 실패: {str(e)}", verbose=False)
            return []
    
    def scrape_bank_data(self, bank_name, driver):
        """단일 은행의 데이터를 스크래핑합니다."""
        self.logger.log_message(f"[시작] {bank_name} 은행 스크래핑 시작")
        
        try:
            # 은행 선택
            if not self.select_bank(driver, bank_name):
                self.logger.log_message(f"{bank_name} 은행 선택 실패")
                return None
            
            # 현재 페이지 URL 저장
            try:
                base_bank_url = driver.current_url
                self.logger.log_message(f"{bank_name} 은행 페이지 접속 성공", verbose=False)
            except:
                self.logger.log_message(f"{bank_name} 은행 페이지 URL 획득 실패")
                return None
            
            # 날짜 정보 추출
            date_info = self.extract_date_information(driver)
            self.logger.log_message(f"{bank_name} 은행 날짜 정보: {date_info}", verbose=True)
            
            result_data = {'날짜정보': date_info}
            all_table_hashes = set()  # 중복 테이블 제거용
            
            # 각 카테고리 처리
            for category in self.config.CATEGORIES:
                try:
                    # 카테고리 탭 클릭
                    if not self.select_category(driver, category):
                        self.logger.log_message(f"{bank_name} 은행 {category} 탭 클릭 실패, 다음 카테고리로 진행")
                        continue
                    
                    # 테이블 추출
                    tables = self.extract_tables_from_page(driver)
                    if not tables:
                        self.logger.log_message(f"{bank_name} 은행 {category} 카테고리에서 테이블을 찾을 수 없습니다.")
                        continue
                    
                    # 중복 제거된 유효 테이블 저장
                    valid_tables = []
                    for df in tables:
                        # 테이블 해시 생성 (중복 확인용)
                        try:
                            shape_hash = f"{df.shape}"
                            headers_hash = f"{list(df.columns)}"
                            data_hash = ""
                            if len(df) > 0:
                                data_hash = f"{list(df.iloc[0].astype(str))}"
                            
                            table_hash = f"{shape_hash}_{headers_hash}_{data_hash}"
                            
                            if table_hash not in all_table_hashes:
                                valid_tables.append(df)
                                all_table_hashes.add(table_hash)
                        except:
                            valid_tables.append(df)
                    
                    # 유효한 테이블 저장
                    if valid_tables:
                        result_data[category] = valid_tables
                        self.logger.log_message(f"{bank_name} 은행 {category} 카테고리에서 {len(valid_tables)}개 테이블 추출")
                
                except Exception as e:
                    self.logger.log_message(f"{bank_name} 은행 {category} 카테고리 처리 실패: {str(e)}")
            
            # 데이터 수집 여부 확인
            if not any(isinstance(data, list) and data for key, data in result_data.items() if key != '날짜정보'):
                self.logger.log_message(f"{bank_name} 은행에서 데이터를 추출할 수 없습니다.")
                return None
            
            self.logger.log_message(f"[완료] {bank_name} 은행 데이터 수집 완료")
            return result_data
            
        except Exception as e:
            self.logger.log_message(f"{bank_name} 은행 처리 중 오류 발생: {str(e)}")
            return None
    
    def save_bank_data(self, bank_name, data_dict):
        """수집된 은행 데이터를 엑셀 파일로 저장합니다."""
        if not data_dict:
            return False
        
        try:
            # 날짜 정보 추출
            date_info = data_dict.get('날짜정보', '날짜정보없음')
            date_info = date_info.replace('/', '-').replace('\\', '-')  # 파일명에 사용할 수 없는 문자 제거
            
            # 파일명에 날짜 정보 포함
            excel_path = os.path.join(self.config.output_dir, f"{bank_name}_결산_{date_info}.xlsx")
            
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                # 날짜 정보 시트 생성
                date_df = pd.DataFrame({
                    '은행명': [bank_name],
                    '공시 날짜': [date_info],
                    '추출 일시': [datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                    '스크래핑 시스템': [f'결산공시 자동 스크래퍼 v{self.config.VERSION}']
                })
                date_df.to_excel(writer, sheet_name='공시정보', index=False)
                
                # 각 카테고리별 데이터 저장
                for category, tables in data_dict.items():
                    if category == '날짜정보' or not tables:
                        continue
                    
                    # 각 카테고리의 테이블을 별도 시트로 저장
                    for i, df in enumerate(tables):
                        # 시트명 생성
                        if i == 0:
                            sheet_name = category
                        else:
                            sheet_name = f"{category}_{i+1}"
                        
                        # 시트명 길이 제한 (엑셀 제한: 31자)
                        if len(sheet_name) > 31:
                            sheet_name = sheet_name[:31]
                        
                        # MultiIndex 확인 및 처리
                        if isinstance(df.columns, pd.MultiIndex):
                            new_cols = []
                            for col in df.columns:
                                if isinstance(col, tuple):
                                    col_parts = [str(c).strip() for c in col if str(c).strip() and str(c).lower() != 'nan']
                                    new_cols.append('_'.join(col_parts) if col_parts else f"Column_{len(new_cols)+1}")
                                else:
                                    new_cols.append(str(col))
                            df.columns = new_cols
                        
                        # 데이터프레임 저장
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            self.logger.log_message(f"{bank_name} 은행 데이터 저장 완료: {excel_path}")
            return True
            
        except Exception as e:
            self.logger.log_message(f"{bank_name} 은행 데이터 저장 오류: {str(e)}")
            return False
    
    def save_bank_data_to_md(self, bank_name, data_dict, is_settlement=False):
        """수집된 은행 데이터를 마크다운 파일로 저장합니다."""
        if not data_dict:
            return False
        
        try:
            # 날짜 정보 추출
            date_info = data_dict.get('날짜정보', '날짜정보없음')
            date_info = date_info.replace('/', '-').replace('\\', '-')
            
            # 파일명 설정
            file_suffix = "결산" if is_settlement else "분기"
            md_path = os.path.join(self.config.output_dir, f"{bank_name}_{file_suffix}_{date_info}.md")
            
            with open(md_path, 'w', encoding='utf-8') as f:
                # 헤더 작성
                f.write(f"# {bank_name} 저축은행 {file_suffix}공시 데이터\n\n")
                f.write(f"## 📋 기본 정보\n\n")
                f.write(f"- **🏦 은행명**: {bank_name}\n")
                f.write(f"- **📅 공시 날짜**: {date_info}\n")
                f.write(f"- **⏰ 추출 일시**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"- **🔧 스크래핑 시스템**: {file_suffix}공시 자동 스크래퍼 v{self.config.VERSION}\n")
                f.write(f"- **🌐 데이터 출처**: 저축은행중앙회 결산공시\n\n")
                
                # 목차 생성
                f.write("## 📚 목차\n\n")
                for category in data_dict.keys():
                    if category != '날짜정보' and data_dict[category]:
                        f.write(f"- [{category}](#{category.lower().replace(' ', '-')})\n")
                f.write("\n")
                
                # 데이터 개요
                f.write("## 📊 데이터 개요\n\n")
                total_tables = sum(len(tables) for tables in data_dict.values() if isinstance(tables, list))
                categories_count = len([k for k in data_dict.keys() if k != '날짜정보' and data_dict[k]])
                f.write(f"- **카테고리 수**: {categories_count}개\n")
                f.write(f"- **전체 테이블 수**: {total_tables}개\n\n")
                
                # 각 카테고리별 데이터 작성
                for category, tables in data_dict.items():
                    if category == '날짜정보' or not tables:
                        continue
                    
                    f.write(f"## {category}\n\n")
                    f.write(f"📋 **{category}** 카테고리에서 추출된 데이터입니다.\n\n")
                    
                    for i, df in enumerate(tables):
                        if i > 0:
                            f.write(f"### {category} - 테이블 {i+1}\n\n")
                        
                        # DataFrame을 마크다운 테이블로 변환
                        if not df.empty:
                            # 열 이름 정리
                            df_clean = df.copy()
                            df_clean.columns = [str(col).replace('\n', ' ').replace('|', '\\|').strip() for col in df_clean.columns]
                            
                            # 테이블 정보
                            f.write(f"📈 **테이블 크기**: {len(df_clean)}행 × {len(df_clean.columns)}열\n\n")
                            
                            # 마크다운 테이블 헤더
                            headers = '| ' + ' | '.join(df_clean.columns) + ' |\n'
                            separator = '|' + '|'.join([' --- ' for _ in df_clean.columns]) + '|\n'
                            f.write(headers)
                            f.write(separator)
                            
                            # 데이터 행 (최대 50행까지만 표시)
                            max_rows = min(50, len(df_clean))
                            for idx in range(max_rows):
                                row = df_clean.iloc[idx]
                                row_data = []
                                for value in row:
                                    # 값 정리 (파이프 문자 이스케이프, 개행 제거)
                                    str_value = str(value).replace('|', '\\|').replace('\n', ' ').replace('\r', '').strip()
                                    if str_value == 'nan' or str_value == 'None':
                                        str_value = ''
                                    # 긴 텍스트는 줄임표 처리
                                    if len(str_value) > 50:
                                        str_value = str_value[:47] + "..."
                                    row_data.append(str_value)
                                f.write('| ' + ' | '.join(row_data) + ' |\n')
                            
                            if len(df_clean) > 50:
                                f.write(f"\n⚠️ *({len(df_clean) - 50}개 행 더 있음... 전체 데이터는 엑셀 파일을 참조하세요.)*\n")
                            
                            f.write('\n')
                        else:
                            f.write("⚠️ *데이터가 없습니다.*\n\n")
                
                # 주요 정보 요약 (결산공시용)
                f.write("## 🔍 주요 정보 요약\n\n")
                f.write("### 📈 재무 현황\n\n")
                f.write("결산공시 데이터에서 추출된 주요 재무 지표들입니다:\n\n")
                
                # 재무현황에서 주요 지표 찾기
                key_indicators = {}
                for category, tables in data_dict.items():
                    if '재무' in category and tables:
                        for df in tables:
                            if not df.empty:
                                # 총자산, 자기자본 등 찾기
                                for idx in range(len(df)):
                                    for col in df.columns:
                                        try:
                                            cell_value = str(df.iloc[idx][col]).strip()
                                            if '총자산' in cell_value:
                                                for other_col in df.columns:
                                                    if other_col != col:
                                                        try:
                                                            value = pd.to_numeric(str(df.iloc[idx][other_col]).replace(',', ''), errors='coerce')
                                                            if pd.notna(value) and value > 0:
                                                                key_indicators['총자산'] = f"{value:,.0f}백만원"
                                                                break
                                                        except:
                                                            pass
                                            elif '자기자본' in cell_value and '자산' not in cell_value:
                                                for other_col in df.columns:
                                                    if other_col != col:
                                                        try:
                                                            value = pd.to_numeric(str(df.iloc[idx][other_col]).replace(',', ''), errors='coerce')
                                                            if pd.notna(value) and value > 0:
                                                                key_indicators['자기자본'] = f"{value:,.0f}백만원"
                                                                break
                                                        except:
                                                            pass
                                        except:
                                            pass
                
                if key_indicators:
                    for key, value in key_indicators.items():
                        f.write(f"- **{key}**: {value}\n")
                else:
                    f.write("- 주요 재무 지표를 자동으로 추출하지 못했습니다. 상세 데이터는 위 테이블을 참조하세요.\n")
                
                f.write("\n")
                
                # 데이터 품질 정보
                f.write("## ✅ 데이터 품질 정보\n\n")
                f.write("- **✅ 완료된 카테고리**: ")
                completed_categories = [k for k in data_dict.keys() if k != '날짜정보' and data_dict[k]]
                f.write(", ".join(completed_categories) + "\n")
                f.write(f"- **📊 총 테이블 수**: {total_tables}개\n")
                f.write(f"- **🔄 처리 상태**: 성공\n\n")
                
                # 사용 안내
                f.write("## 💡 사용 안내\n\n")
                f.write("### 📖 이 파일 활용 방법\n\n")
                f.write("1. **GitHub/GitLab**: 이 MD 파일을 저장소에 업로드하면 자동으로 포맷팅되어 표시됩니다.\n")
                f.write("2. **Notion**: 임포트 기능으로 이 파일을 불러올 수 있습니다.\n")
                f.write("3. **VS Code**: 마크다운 프리뷰로 보기 좋게 확인할 수 있습니다.\n")
                f.write("4. **Typora/Mark Text**: 전용 마크다운 에디터로 열어보세요.\n\n")
                
                f.write("### 🔗 관련 파일\n\n")
                f.write(f"- **📊 엑셀 파일**: `{bank_name}_{file_suffix}_{date_info}.xlsx`\n")
                f.write(f"- **📝 요약 보고서**: `저축은행_{file_suffix}공시_스크래핑_요약_{self.config.today}.md`\n")
                f.write(f"- **📈 통합 보고서**: `저축은행_{file_suffix}_재무데이터_통합_{self.config.today}.xlsx`\n\n")
                
                # 푸터
                f.write("---\n\n")
                f.write("### 📞 문의 및 지원\n\n")
                f.write("이 데이터에 대한 문의사항이 있으시면:\n\n")
                f.write("- 🌐 저축은행중앙회 홈페이지: https://www.fsb.or.kr\n")
                f.write("- 📧 데이터 품질 문의: 개발팀\n")
                f.write("- 🔧 스크래퍼 개선 제안: GitHub Issues\n\n")
                f.write(f"*이 문서는 {datetime.now().strftime('%Y년 %m월 %d일 %H:%M:%S')}에 자동 생성되었습니다.*\n")
            
            self.logger.log_message(f"{bank_name} 은행 MD 파일 저장 완료: {md_path}")
            return True
            
        except Exception as e:
            self.logger.log_message(f"{bank_name} 은행 MD 파일 저장 오류: {str(e)}")
            return False
    
    def worker_process_bank(self, bank_name, progress_callback=None, save_md=False):
        """단일 은행을 처리합니다."""
        driver = self.driver_manager.get_driver()
        
        try:
            # 최대 재시도 횟수만큼 시도
            for attempt in range(self.config.MAX_RETRIES):
                try:
                    # 진행 상황 업데이트
                    if progress_callback:
                        progress_callback(bank_name, "처리 중")
                    
                    # 은행 데이터 스크래핑
                    result_data = self.scrape_bank_data(bank_name, driver)
                    
                    if result_data:
                        # 엑셀 데이터 저장
                        excel_saved = self.save_bank_data(bank_name, result_data)
                        
                        # MD 데이터 저장 (옵션) - 결산공시이므로 True
                        md_saved = True
                        if save_md:
                            md_saved = self.save_bank_data_to_md(bank_name, result_data, is_settlement=True)
                        
                        if excel_saved and md_saved:
                            self.progress_manager.mark_completed(bank_name)
                            self.driver_manager.return_driver(driver)
                            
                            # 진행 상황 업데이트
                            if progress_callback:
                                progress_callback(bank_name, "완료")
                            
                            return bank_name, True
                        else:
                            if attempt < self.config.MAX_RETRIES - 1:
                                self.logger.log_message(f"{bank_name} 은행 데이터 저장 실패, 재시도 {attempt+1}/{self.config.MAX_RETRIES}...")
                                
                                # 진행 상황 업데이트
                                if progress_callback:
                                    progress_callback(bank_name, f"저장 재시도 {attempt+1}")
                            else:
                                self.logger.log_message(f"{bank_name} 은행 데이터 저장 실패, 최대 시도 횟수 초과")
                                
                                # 진행 상황 업데이트
                                if progress_callback:
                                    progress_callback(bank_name, "저장 실패")
                    else:
                        if attempt < self.config.MAX_RETRIES - 1:
                            self.logger.log_message(f"{bank_name} 은행 데이터 스크래핑 실패, 재시도 {attempt+1}/{self.config.MAX_RETRIES}...")
                            WaitUtils.wait_with_random(1, 2)  # 재시도 전 잠시 대기
                            
                            # 진행 상황 업데이트
                            if progress_callback:
                                progress_callback(bank_name, f"스크래핑 재시도 {attempt+1}")
                        else:
                            self.logger.log_message(f"{bank_name} 은행 데이터 스크래핑 실패, 최대 시도 횟수 초과")
                            
                            # 진행 상황 업데이트
                            if progress_callback:
                                progress_callback(bank_name, "스크래핑 실패")
                
                except Exception as e:
                    if attempt < self.config.MAX_RETRIES - 1:
                        self.logger.log_message(f"{bank_name} 은행 처리 중 오류: {str(e)}, 재시도 {attempt+1}/{self.config.MAX_RETRIES}...")
                        WaitUtils.wait_with_random(1, 2)
                        
                        # 진행 상황 업데이트
                        if progress_callback:
                            progress_callback(bank_name, f"오류, 재시도 {attempt+1}")
                    else:
                        self.logger.log_message(f"{bank_name} 은행 처리 실패: {str(e)}, 최대 시도 횟수 초과")
                        
                        # 진행 상황 업데이트
                        if progress_callback:
                            progress_callback(bank_name, "실패")
            
            # 모든 시도 실패
            self.progress_manager.mark_failed(bank_name)
            self.driver_manager.return_driver(driver)
            
            # 진행 상황 업데이트
            if progress_callback:
                progress_callback(bank_name, "실패")
            
            return bank_name, False
            
        except Exception as e:
            self.logger.log_message(f"{bank_name} 은행 처리 중 예상치 못한 오류: {str(e)}")
            self.progress_manager.mark_failed(bank_name)
            self.driver_manager.return_driver(driver)
            
            # 진행 상황 업데이트
            if progress_callback:
                progress_callback(bank_name, "오류")
            
            return bank_name, False
    
    def process_banks(self, banks, progress_callback=None, save_md=False):
        """은행 목록을 병렬로 처리합니다."""
        with concurrent.futures.ThreadPoolExecutor(max_workers=self.config.MAX_WORKERS) as executor:
            # 작업 제출
            future_to_bank = {
                executor.submit(self.worker_process_bank, bank, progress_callback, save_md): bank
                for bank in banks
            }
            
            # 결과 수집
            results = []
            for future in concurrent.futures.as_completed(future_to_bank):
                bank = future_to_bank[future]
                try:
                    result = future.result()
                    results.append(result)
                except Exception as e:
                    self.logger.log_message(f"{bank} 은행 처리 중 예외 발생: {e}")
                    results.append((bank, False))
            
            return results
    
    def generate_summary_report(self):
        """스크래핑 결과 요약 보고서를 생성합니다."""
        try:
            # 완료된 은행과 실패한 은행 목록
            completed_banks = self.progress_manager.progress.get('completed', [])
            failed_banks = self.progress_manager.progress.get('failed', [])
            
            # 은행별 데이터 요약
            bank_summary = []
            
            # 완료된 은행 파일 검사
            for bank in self.config.BANKS:
                # 각 은행의 엑셀 파일 찾기
                bank_files = [f for f in os.listdir(self.config.output_dir) if f.startswith(f"{bank}_결산_") and f.endswith(".xlsx")]
                
                if bank_files:
                    try:
                        # 가장 최근 파일 선택
                        latest_file = sorted(bank_files)[-1]
                        file_path = os.path.join(self.config.output_dir, latest_file)
                        
                        # 엑셀 파일 분석
                        xls = pd.ExcelFile(file_path)
                        sheet_count = len(xls.sheet_names)
                        
                        # 카테고리 추출
                        categories = []
                        for sheet in xls.sheet_names:
                            if sheet != '공시정보':
                                category = sheet.split('_')[0] if '_' in sheet else sheet
                                categories.append(category)
                        
                        # 중복 제거
                        categories = sorted(list(set(categories)))
                        
                        # 날짜 정보 추출
                        date_info = "날짜 정보 없음"
                        if '공시정보' in xls.sheet_names:
                            info_df = pd.read_excel(file_path, sheet_name='공시정보')
                            if '공시 날짜' in info_df.columns and not info_df['공시 날짜'].empty:
                                date_info = str(info_df['공시 날짜'].iloc[0])
                        
                        status = '완료' if set(categories) >= set(self.config.CATEGORIES) else '부분 완료'
                        
                        bank_summary.append({
                            '은행명': bank,
                            '스크래핑 상태': status,
                            '공시 날짜': date_info,
                            '시트 수': sheet_count - 1,  # 공시정보 시트 제외
                            '스크래핑된 카테고리': ', '.join(categories)
                        })
                    except Exception as e:
                        bank_summary.append({
                            '은행명': bank,
                            '스크래핑 상태': '파일 손상',
                            '공시 날짜': '확인 불가',
                            '시트 수': '확인 불가',
                            '스크래핑된 카테고리': f'오류: {str(e)}'
                        })
                else:
                    status = '실패' if bank in failed_banks else '미처리'
                    bank_summary.append({
                        '은행명': bank,
                        '스크래핑 상태': status,
                        '공시 날짜': '',
                        '시트 수': 0,
                        '스크래핑된 카테고리': ''
                    })
            
            # 요약 DataFrame 생성
            summary_df = pd.DataFrame(bank_summary)
            
            # 완료 상태별 정렬
            status_order = {'완료': 0, '부분 완료': 1, '파일 손상': 2, '실패': 3, '미처리': 4}
            summary_df['상태순서'] = summary_df['스크래핑 상태'].map(status_order)
            summary_df = summary_df.sort_values(['상태순서', '은행명']).drop('상태순서', axis=1)
            
            # 요약 저장
            summary_file = os.path.join(self.config.output_dir, f"저축은행_결산공시_스크래핑_요약_{self.config.today}.xlsx")
            summary_df.to_excel(summary_file, index=False)
            
            # 통계 정보
            stats = {
                '전체 은행 수': len(self.config.BANKS),
                '완료 은행 수': len([r for r in bank_summary if r['스크래핑 상태'] == '완료']),
                '부분 완료 은행 수': len([r for r in bank_summary if r['스크래핑 상태'] == '부분 완료']),
                '실패 은행 수': len([r for r in bank_summary if r['스크래핑 상태'] in ['실패', '파일 손상']]),
                '성공률': f"{len([r for r in bank_summary if r['스크래핑 상태'] in ['완료', '부분 완료']]) / len(self.config.BANKS) * 100:.2f}%"
            }
            
            self.logger.log_message("\n===== 스크래핑 결과 요약 =====")
            for key, value in stats.items():
                self.logger.log_message(f"{key}: {value}")
            
            self.logger.log_message(f"요약 파일 저장 완료: {summary_file}")
            return summary_file, stats, summary_df
            
        except Exception as e:
            self.logger.log_message(f"요약 보고서 생성 오류: {str(e)}")
            return None, {}, None
    
    def generate_summary_report_md(self):
        """스크래핑 결과 요약을 마크다운으로 생성합니다."""
        try:
            # 기존 요약 보고서 생성
            summary_file, stats, summary_df = self.generate_summary_report()
            
            if summary_df is None:
                return None
            
            # MD 파일 경로
            md_summary_file = os.path.join(self.config.output_dir, 
                                          f"저축은행_결산공시_스크래핑_요약_{self.config.today}.md")
            
            with open(md_summary_file, 'w', encoding='utf-8') as f:
                f.write(f"# 🏦 저축은행 결산공시 스크래핑 결과 요약\n\n")
                f.write(f"📅 **보고서 생성일**: {datetime.now().strftime('%Y년 %m월 %d일 %H:%M:%S')}\n")
                f.write(f"🌐 **데이터 출처**: 저축은행중앙회 결산공시 시스템\n")
                f.write(f"🔧 **스크래퍼 버전**: v{self.config.VERSION}\n\n")
                
                # 전체 통계 (더 시각적으로)
                f.write("## 📊 전체 통계\n\n")
                f.write("```\n")
                f.write("┌─────────────────────────────────────┐\n")
                f.write("│           스크래핑 결과              │\n")
                f.write("├─────────────────────────────────────┤\n")
                for key, value in stats.items():
                    f.write(f"│ {key:<15}: {str(value):>15} │\n")
                f.write("└─────────────────────────────────────┘\n")
                f.write("```\n\n")
                
                # 성공률에 따른 상태 표시
                success_rate = float(stats['성공률'].replace('%', ''))
                if success_rate >= 95:
                    status_emoji = "🟢"
                    status_text = "매우 양호"
                elif success_rate >= 85:
                    status_emoji = "🟡"
                    status_text = "양호"
                else:
                    status_emoji = "🔴"
                    status_text = "개선 필요"
                
                f.write(f"### {status_emoji} 전체 상태: {status_text}\n\n")
                
                # 상태별 은행 분류 (더 상세하게)
                f.write("## 📋 상태별 은행 현황\n\n")
                
                status_groups = summary_df.groupby('스크래핑 상태')
                for status, group in status_groups:
                    emoji_map = {
                        '완료': '✅',
                        '부분 완료': '⚠️', 
                        '실패': '❌',
                        '파일 손상': '💥',
                        '미처리': '⏳'
                    }
                    emoji = emoji_map.get(status, '📋')
                    
                    f.write(f"### {emoji} {status} ({len(group)}개 은행)\n\n")
                    
                    if len(group) > 0:
                        # 테이블 형식으로 표시
                        f.write("| 은행명 | 공시날짜 | 시트수 | 카테고리 |\n")
                        f.write("| --- | --- | --- | --- |\n")
                        
                        for _, row in group.iterrows():
                            bank_name = row['은행명']
                            date = row['공시 날짜'] if row['공시 날짜'] and row['공시 날짜'] != '' else '-'
                            sheets = str(row['시트 수']) if row['시트 수'] != '확인 불가' else '-'
                            categories = row['스크래핑된 카테고리'][:30] + "..." if len(str(row['스크래핑된 카테고리'])) > 30 else row['스크래핑된 카테고리']
                            
                            f.write(f"| {bank_name} | {date} | {sheets} | {categories} |\n")
                    f.write("\n")
                
                # 문제 은행 상세 분석
                failed_banks = summary_df[summary_df['스크래핑 상태'].isin(['실패', '파일 손상'])]
                if not failed_banks.empty:
                    f.write("## 🔧 문제 해결이 필요한 은행들\n\n")
                    f.write("다음 은행들의 데이터 수집에 문제가 있었습니다:\n\n")
                    
                    for _, row in failed_banks.iterrows():
                        f.write(f"### ❌ {row['은행명']}\n\n")
                        f.write(f"- **상태**: {row['스크래핑 상태']}\n")
                        if row['스크래핑된 카테고리'] and str(row['스크래핑된 카테고리']).startswith('오류:'):
                            f.write(f"- **오류 내용**: {row['스크래핑된 카테고리']}\n")
                        f.write("- **권장 조치**: 수동으로 재시도하거나 웹사이트 변경사항을 확인하세요.\n\n")
                
                # 권장사항 (더 구체적으로)
                f.write("## 💡 권장사항 및 다음 단계\n\n")
                
                f.write("### 🔄 즉시 조치 사항\n\n")
                failed_count = len(summary_df[summary_df['스크래핑 상태'].isin(['실패', '파일 손상'])])
                if failed_count > 0:
                    f.write(f"1. **실패한 {failed_count}개 은행 재시도**\n")
                    f.write("   - 네트워크 상태 확인\n")
                    f.write("   - 웹사이트 접속 가능 여부 확인\n")
                    f.write("   - 개별적으로 수동 재시도\n\n")
                
                partial_count = len(summary_df[summary_df['스크래핑 상태'] == '부분 완료'])
                if partial_count > 0:
                    f.write(f"2. **부분 완료된 {partial_count}개 은행 검토**\n")
                    f.write("   - 누락된 카테고리 확인\n")
                    f.write("   - 필요시 해당 카테고리만 재수집\n\n")
                
                f.write("### 📈 데이터 활용 방안\n\n")
                f.write("1. **재무 분석**: 통합 재무 보고서를 활용한 은행 간 비교 분석\n")
                f.write("2. **트렌드 분석**: 시계열 데이터 축적 후 변화 추이 분석\n")
                f.write("3. **리스크 관리**: 자기자본비율, NPL 비율 등 건전성 지표 모니터링\n")
                f.write("4. **보고서 작성**: MD 파일을 활용한 자동화된 보고서 생성\n\n")
                
                f.write("### 🔧 시스템 개선 사항\n\n")
                f.write("1. **정기 실행**: 결산 발표 시기에 맞춘 자동 스크래핑 스케줄링\n")
                f.write("2. **알림 설정**: 실패한 은행에 대한 자동 알림 기능\n")
                f.write("3. **데이터 검증**: 수집된 데이터의 품질 자동 검증 로직\n")
                f.write("4. **백업 시스템**: 중요 데이터의 자동 백업 및 버전 관리\n\n")
                
                # 파일 정보
                f.write("## 📁 생성된 파일들\n\n")
                f.write("이번 스크래핑으로 다음 파일들이 생성되었습니다:\n\n")
                
                # 개별 은행 파일들
                completed_banks = summary_df[summary_df['스크래핑 상태'].isin(['완료', '부분 완료'])]
                f.write(f"### 📊 개별 은행 데이터 ({len(completed_banks)}개)\n\n")
                f.write("| 은행명 | 엑셀 파일 | MD 파일 |\n")
                f.write("| --- | --- | --- |\n")
                
                for _, row in completed_banks.iterrows():
                    bank_name = row['은행명']
                    date = row['공시 날짜'] if row['공시 날짜'] and row['공시 날짜'] != '' else 'unknown'
                    date_clean = date.replace('/', '-').replace('\\', '-')
                    
                    excel_file = f"{bank_name}_결산_{date_clean}.xlsx"
                    md_file = f"{bank_name}_결산_{date_clean}.md"
                    
                    f.write(f"| {bank_name} | {excel_file} | {md_file} |\n")
                
                f.write("\n### 📋 요약 및 통합 파일\n\n")
                f.write("- 📊 **엑셀 요약**: `저축은행_결산공시_스크래핑_요약_" + self.config.today + ".xlsx`\n")
                f.write("- 📝 **MD 요약**: `저축은행_결산공시_스크래핑_요약_" + self.config.today + ".md` (이 파일)\n")
                f.write("- 📈 **통합 재무**: `저축은행_결산_재무데이터_통합_" + self.config.today + ".xlsx`\n")
                f.write("- 📝 **MD 통합**: `저축은행_결산_MD_통합_보고서_" + self.config.today + ".md`\n")
                f.write("- 🗜️ **압축 파일**: `저축은행_결산공시_데이터_" + self.config.today + ".zip` (선택시)\n\n")
                
                # 푸터 (더 상세하게)
                f.write("---\n\n")
                f.write("### 🔗 추가 리소스\n\n")
                f.write("- 🌐 **저축은행중앙회**: https://www.fsb.or.kr\n")
                f.write("- 📊 **결산공시 시스템**: https://www.fsb.or.kr/busmagesett_0100.act\n")
                f.write("- 📖 **스크래퍼 사용법**: 프로그램 도움말 참조\n")
                f.write("- 🐛 **버그 신고**: 개발팀에 문의\n\n")
                
                f.write("### ⚖️ 법적 고지\n\n")
                f.write("- 이 데이터는 저축은행중앙회의 공개 정보를 기반으로 수집되었습니다.\n")
                f.write("- 데이터의 정확성은 원본 공시 자료에 의존합니다.\n")
                f.write("- 투자 결정시에는 반드시 공식 공시 자료를 확인하시기 바랍니다.\n\n")
                
                f.write(f"**자동 생성**: {datetime.now().strftime('%Y년 %m월 %d일 %H:%M:%S')}\n")
                f.write(f"**스크래퍼 버전**: 저축은행 통합 데이터 스크래퍼 v{self.config.VERSION}\n")
            
            self.logger.log_message(f"MD 요약 보고서 저장 완료: {md_summary_file}")
            return md_summary_file
            
        except Exception as e:
            self.logger.log_message(f"MD 요약 보고서 생성 오류: {str(e)}")
            return None
    
    # v2.6 - MD 파일 통합 기능 추가
    def create_consolidated_md_report(self):
        """기본 출력 폴더의 모든 MD 파일을 통합한 보고서를 생성합니다."""
        try:
            self.logger.log_message("\n===== MD 파일 통합 보고서 생성 시작 =====")
            
            # 기본 출력 폴더에서 모든 MD 파일 찾기
            md_files = [f for f in os.listdir(self.config.output_dir) 
                       if f.endswith('.md') and f.startswith(tuple(self.config.BANKS))]
            
            self.logger.log_message(f"발견된 개별 은행 MD 파일: {len(md_files)}개")
            
            if not md_files:
                self.logger.log_message("통합할 MD 파일이 없습니다.")
                return None
            
            # 파일 경로 생성
            file_paths = [os.path.join(self.config.output_dir, f) for f in md_files]
            
            # 통합 보고서 생성
            return self._process_md_consolidation(file_paths, self.config.output_dir)
            
        except Exception as e:
            self.logger.log_message(f"MD 파일 통합 보고서 생성 오류: {str(e)}")
            return None
    
    def create_consolidated_md_report_from_folder(self, folder_path):
        """지정된 폴더의 모든 MD 파일을 통합한 보고서를 생성합니다."""
        try:
            self.logger.log_message(f"\n===== MD 파일 통합 보고서 생성 시작 (폴더: {folder_path}) =====")
            
            # 폴더 내 모든 MD 파일 찾기
            md_files = [f for f in os.listdir(folder_path) if f.endswith('.md')]
            self.logger.log_message(f"발견된 MD 파일: {len(md_files)}개")
            
            if not md_files:
                self.logger.log_message("통합할 MD 파일이 없습니다.")
                return None
            
            # 파일 경로 생성
            file_paths = [os.path.join(folder_path, f) for f in md_files]
            
            # 통합 보고서 생성
            return self._process_md_consolidation(file_paths, folder_path)
            
        except Exception as e:
            self.logger.log_message(f"MD 파일 통합 보고서 생성 오류: {str(e)}")
            return None
    
    def create_consolidated_md_report_from_files(self, file_list):
        """선택된 MD 파일들을 통합한 보고서를 생성합니다."""
        try:
            self.logger.log_message(f"\n===== MD 파일 통합 보고서 생성 시작 (파일 {len(file_list)}개) =====")
            
            # 출력 폴더 결정 (첫 번째 파일의 폴더 사용)
            output_folder = os.path.dirname(file_list[0]) if file_list else self.config.output_dir
            
            # 통합 보고서 생성
            return self._process_md_consolidation(file_list, output_folder)
            
        except Exception as e:
            self.logger.log_message(f"MD 파일 통합 보고서 생성 오류: {str(e)}")
            return None
    
    def _process_md_consolidation(self, file_paths, output_folder):
        """MD 파일들을 통합하여 하나의 보고서를 생성합니다. (v2.6 핵심 기능)"""
        try:
            # 통합 MD 파일 경로
            consolidated_md_file = os.path.join(output_folder, f'저축은행_결산_MD_통합_보고서_{self.config.today}.md')
            
            # 각 MD 파일에서 데이터 추출
            bank_data_list = []
            
            for file_path in file_paths:
                try:
                    bank_data = self._extract_bank_info_from_md(file_path)
                    if bank_data:
                        bank_data_list.append(bank_data)
                        self.logger.log_message(f"MD 파일 분석 완료: {os.path.basename(file_path)}")
                except Exception as e:
                    self.logger.log_message(f"MD 파일 분석 실패: {os.path.basename(file_path)} - {str(e)}")
                    continue
            
            if not bank_data_list:
                self.logger.log_message("분석 가능한 MD 파일이 없습니다.")
                return None
            
            # 통합 MD 보고서 생성
            with open(consolidated_md_file, 'w', encoding='utf-8') as f:
                # 헤더 작성
                f.write(f"# 🏦 저축은행 결산공시 MD 통합 보고서\n\n")
                f.write(f"## 📊 통합 보고서 개요\n\n")
                f.write(f"- **📅 생성일시**: {datetime.now().strftime('%Y년 %m월 %d일 %H:%M:%S')}\n")
                f.write(f"- **🏦 포함된 은행 수**: {len(bank_data_list)}개\n")
                f.write(f"- **📝 통합된 MD 파일 수**: {len(file_paths)}개\n")
                f.write(f"- **🔧 스크래퍼 버전**: v{self.config.VERSION}\n")
                f.write(f"- **🌐 데이터 출처**: 저축은행중앙회 결산공시 시스템\n\n")
                
                # 목차 생성
                f.write("## 📚 목차\n\n")
                f.write("1. [전체 요약](#전체-요약)\n")
                f.write("2. [은행별 상세 정보](#은행별-상세-정보)\n")
                f.write("3. [주요 재무 지표 비교](#주요-재무-지표-비교)\n")
                f.write("4. [데이터 품질 보고](#데이터-품질-보고)\n")
                f.write("5. [활용 가이드](#활용-가이드)\n\n")
                
                # 전체 요약
                f.write("## 📈 전체 요약\n\n")
                
                # 공시 날짜별 분류
                date_groups = {}
                for bank_data in bank_data_list:
                    date = bank_data.get('공시_날짜', '날짜 없음')
                    if date not in date_groups:
                        date_groups[date] = []
                    date_groups[date].append(bank_data['은행명'])
                
                f.write("### 📅 공시 날짜별 현황\n\n")
                for date, banks in date_groups.items():
                    f.write(f"- **{date}**: {len(banks)}개 은행 ({', '.join(sorted(banks))})\n")
                f.write("\n")
                
                # 카테고리별 데이터 완성도
                f.write("### 📊 카테고리별 데이터 완성도\n\n")
                category_stats = {}
                for category in self.config.CATEGORIES:
                    count = sum(1 for bank_data in bank_data_list if category in bank_data.get('카테고리', []))
                    category_stats[category] = count
                
                f.write("| 카테고리 | 완성된 은행 수 | 완성도 |\n")
                f.write("| --- | --- | --- |\n")
                for category, count in category_stats.items():
                    percentage = (count / len(bank_data_list)) * 100
                    progress_bar = "🟩" * int(percentage // 10) + "⬜" * (10 - int(percentage // 10))
                    f.write(f"| {category} | {count}/{len(bank_data_list)} | {percentage:.1f}% {progress_bar} |\n")
                f.write("\n")
                
                # 은행별 상세 정보
                f.write("## 🏦 은행별 상세 정보\n\n")
                
                # 은행명으로 정렬
                sorted_banks = sorted(bank_data_list, key=lambda x: x['은행명'])
                
                for i, bank_data in enumerate(sorted_banks):
                    bank_name = bank_data['은행명']
                    f.write(f"### {i+1}. {bank_name} 저축은행\n\n")
                    
                    # 기본 정보
                    f.write("#### 📋 기본 정보\n\n")
                    f.write(f"- **📅 공시 날짜**: {bank_data.get('공시_날짜', '정보 없음')}\n")
                    f.write(f"- **📊 데이터 카테고리**: {', '.join(bank_data.get('카테고리', []))}\n")
                    f.write(f"- **📈 테이블 수**: {bank_data.get('테이블_수', 0)}개\n")
                    f.write(f"- **⏰ 추출 일시**: {bank_data.get('추출_일시', '정보 없음')}\n")
                    f.write(f"- **📁 원본 파일**: `{bank_data.get('파일명', '알 수 없음')}`\n\n")
                    
                    # 주요 재무 지표 (추출 가능한 경우)
                    financial_indicators = bank_data.get('재무_지표', {})
                    if financial_indicators:
                        f.write("#### 💰 주요 재무 지표\n\n")
                        for indicator, value in financial_indicators.items():
                            f.write(f"- **{indicator}**: {value}\n")
                        f.write("\n")
                    
                    # 데이터 품질 정보
                    f.write("#### ✅ 데이터 품질\n\n")
                    quality_score = len(bank_data.get('카테고리', [])) / len(self.config.CATEGORIES) * 100
                    quality_emoji = "🟢" if quality_score >= 75 else "🟡" if quality_score >= 50 else "🔴"
                    f.write(f"- **품질 점수**: {quality_emoji} {quality_score:.1f}% ({len(bank_data.get('카테고리', []))}/{len(self.config.CATEGORIES)} 카테고리 완료)\n")
                    
                    # 누락된 카테고리
                    missing_categories = set(self.config.CATEGORIES) - set(bank_data.get('카테고리', []))
                    if missing_categories:
                        f.write(f"- **누락된 카테고리**: {', '.join(missing_categories)}\n")
                    
                    f.write("\n")
                    
                    # 구분선
                    if i < len(sorted_banks) - 1:
                        f.write("---\n\n")
                
                # 주요 재무 지표 비교
                f.write("## 📊 주요 재무 지표 비교\n\n")
                
                # 재무 지표가 있는 은행들만 추출
                banks_with_financial_data = [b for b in bank_data_list if b.get('재무_지표')]
                
                if banks_with_financial_data:
                    f.write("### 💰 재무 지표 보유 은행\n\n")
                    f.write("| 은행명 | 총자산 | 자기자본 | 기타 지표 |\n")
                    f.write("| --- | --- | --- | --- |\n")
                    
                    for bank_data in sorted(banks_with_financial_data, key=lambda x: x['은행명']):
                        indicators = bank_data.get('재무_지표', {})
                        total_assets = indicators.get('총자산', '-')
                        equity = indicators.get('자기자본', '-')
                        other_count = len(indicators) - (2 if '총자산' in indicators and '자기자본' in indicators else len(indicators))
                        
                        f.write(f"| {bank_data['은행명']} | {total_assets} | {equity} | {other_count}개 |\n")
                    
                    f.write(f"\n📈 **총 {len(banks_with_financial_data)}개 은행의 재무 지표 데이터 보유**\n\n")
                else:
                    f.write("⚠️ 재무 지표를 자동 추출할 수 있는 은행이 없습니다. 상세 데이터는 개별 MD 파일을 참조하세요.\n\n")
                
                # 데이터 품질 보고
                f.write("## ✅ 데이터 품질 보고\n\n")
                
                # 완성도별 분류
                excellent_banks = [b for b in bank_data_list if len(b.get('카테고리', [])) == len(self.config.CATEGORIES)]
                good_banks = [b for b in bank_data_list if len(self.config.CATEGORIES) > len(b.get('카테고리', [])) >= len(self.config.CATEGORIES) * 0.75]
                fair_banks = [b for b in bank_data_list if len(self.config.CATEGORIES) * 0.75 > len(b.get('카테고리', [])) >= len(self.config.CATEGORIES) * 0.5]
                poor_banks = [b for b in bank_data_list if len(b.get('카테고리', [])) < len(self.config.CATEGORIES) * 0.5]
                
                f.write("### 📈 완성도별 분류\n\n")
                f.write(f"- 🟢 **우수 (100%)**: {len(excellent_banks)}개 은행\n")
                if excellent_banks:
                    f.write(f"  - {', '.join([b['은행명'] for b in excellent_banks])}\n")
                f.write(f"- 🟡 **양호 (75% 이상)**: {len(good_banks)}개 은행\n")
                if good_banks:
                    f.write(f"  - {', '.join([b['은행명'] for b in good_banks])}\n")
                f.write(f"- 🟠 **보통 (50% 이상)**: {len(fair_banks)}개 은행\n")
                if fair_banks:
                    f.write(f"  - {', '.join([b['은행명'] for b in fair_banks])}\n")
                f.write(f"- 🔴 **개선 필요 (50% 미만)**: {len(poor_banks)}개 은행\n")
                if poor_banks:
                    f.write(f"  - {', '.join([b['은행명'] for b in poor_banks])}\n")
                f.write("\n")
                
                # 활용 가이드
                f.write("## 💡 활용 가이드\n\n")
                
                f.write("### 📖 이 통합 보고서 활용 방법\n\n")
                f.write("1. **전체 현황 파악**: 모든 은행의 데이터 수집 상태를 한눈에 확인\n")
                f.write("2. **품질 관리**: 데이터 누락이나 오류가 있는 은행 식별\n")
                f.write("3. **비교 분석**: 은행 간 공시 날짜나 데이터 완성도 비교\n")
                f.write("4. **후속 작업 계획**: 추가 수집이 필요한 은행이나 카테고리 파악\n\n")
                
                f.write("### 🔗 개별 은행 상세 정보\n\n")
                f.write("각 은행의 상세한 재무 데이터는 다음 개별 MD 파일들을 참조하세요:\n\n")
                
                for bank_data in sorted_banks:
                    f.write(f"- **{bank_data['은행명']}**: `{bank_data.get('파일명', '파일명 없음')}`\n")
                f.write("\n")
                
                f.write("### 🔧 개선 제안\n\n")
                if poor_banks:
                    f.write("**즉시 조치가 필요한 은행들**:\n")
                    for bank_data in poor_banks:
                        missing = set(self.config.CATEGORIES) - set(bank_data.get('카테고리', []))
                        f.write(f"- **{bank_data['은행명']}**: {', '.join(missing)} 카테고리 재수집 필요\n")
                    f.write("\n")
                
                f.write("**데이터 품질 향상 방안**:\n")
                f.write("1. 누락된 카테고리가 많은 은행들을 우선적으로 재스크래핑\n")
                f.write("2. 웹사이트 구조 변경 여부 확인 및 스크래퍼 업데이트\n")
                f.write("3. 정기적인 데이터 검증 및 품질 모니터링 체계 구축\n\n")
                
                # 푸터
                f.write("---\n\n")
                f.write("### 📊 통계 요약\n\n")
                f.write(f"- **📁 통합된 파일**: {len(file_paths)}개\n")
                f.write(f"- **🏦 분석된 은행**: {len(bank_data_list)}개\n")
                f.write(f"- **📅 공시 날짜 종류**: {len(date_groups)}개\n")
                f.write(f"- **🟢 완전한 데이터**: {len(excellent_banks)}개 은행\n")
                f.write(f"- **📊 평균 완성도**: {sum(len(b.get('카테고리', [])) for b in bank_data_list) / len(bank_data_list) / len(self.config.CATEGORIES) * 100:.1f}%\n\n")
                
                f.write(f"**자동 생성**: {datetime.now().strftime('%Y년 %m월 %d일 %H:%M:%S')}\n")
                f.write(f"**통합 도구**: 저축은행 결산공시 스크래퍼 v{self.config.VERSION}\n")
            
            self.logger.log_message(f"MD 통합 보고서 생성 완료: {consolidated_md_file}")
            return consolidated_md_file
            
        except Exception as e:
            self.logger.log_message(f"MD 통합 보고서 생성 오류: {str(e)}")
            import traceback
            self.logger.log_message(traceback.format_exc())
            return None
    
    def _extract_bank_info_from_md(self, md_file_path):
        """개별 MD 파일에서 은행 정보를 추출합니다."""
        try:
            with open(md_file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            bank_data = {
                '파일명': os.path.basename(md_file_path),
                '은행명': '알 수 없음',
                '공시_날짜': '정보 없음',
                '추출_일시': '정보 없음',
                '카테고리': [],
                '테이블_수': 0,
                '재무_지표': {}
            }
            
            # 은행명 추출
            bank_name_match = re.search(r'# (.+?) 저축은행', content)
            if bank_name_match:
                bank_data['은행명'] = bank_name_match.group(1)
            else:
                # 파일명에서 은행명 추출 시도
                filename = os.path.basename(md_file_path)
                for bank in self.config.BANKS:
                    if filename.startswith(bank):
                        bank_data['은행명'] = bank
                        break
            
            # 공시 날짜 추출
            date_match = re.search(r'- \*\*📅 공시 날짜\*\*: (.+)', content)
            if date_match:
                bank_data['공시_날짜'] = date_match.group(1)
            
            # 추출 일시 추출
            extract_time_match = re.search(r'- \*\*⏰ 추출 일시\*\*: (.+)', content)
            if extract_time_match:
                bank_data['추출_일시'] = extract_time_match.group(1)
            
            # 카테고리 추출 (목차에서)
            toc_section = re.search(r'## 📚 목차\n\n(.+?)\n\n', content, re.DOTALL)
            if toc_section:
                toc_content = toc_section.group(1)
                for category in self.config.CATEGORIES:
                    if category in toc_content:
                        bank_data['카테고리'].append(category)
            
            # 테이블 수 추출
            table_count_match = re.search(r'- \*\*전체 테이블 수\*\*: (\d+)개', content)
            if table_count_match:
                bank_data['테이블_수'] = int(table_count_match.group(1))
            
            # 재무 지표 추출 (주요 정보 요약 섹션에서)
            financial_section = re.search(r'### 📈 재무 현황\n\n(.+?)(?=\n##|\n---|\Z)', content, re.DOTALL)
            if financial_section:
                financial_content = financial_section.group(1)
                
                # 총자산 추출
                asset_match = re.search(r'- \*\*총자산\*\*: (.+)', financial_content)
                if asset_match:
                    bank_data['재무_지표']['총자산'] = asset_match.group(1)
                
                # 자기자본 추출
                equity_match = re.search(r'- \*\*자기자본\*\*: (.+)', financial_content)
                if equity_match:
                    bank_data['재무_지표']['자기자본'] = equity_match.group(1)
            
            return bank_data
            
        except Exception as e:
            self.logger.log_message(f"MD 파일 정보 추출 실패: {md_file_path} - {str(e)}")
            return None
    
    def extract_financial_data_from_excel(self, excel_path, bank_name):
        """개별 은행 엑셀 파일에서 당기/전년동기 재무 데이터를 추출합니다. (v2.5 개선)"""
        try:
            # 엑셀 파일 열기
            xls = pd.ExcelFile(excel_path)
            
            # 확장된 데이터 구조 (당기/전년동기/증감)
            financial_data = {
                '은행명': bank_name,
                '재무정보 날짜': '데이터 없음',
                '분기': '데이터 없음',
                
                # 당기 데이터
                '당기_총자산': None,
                '당기_자기자본': None,
                '당기_총여신': None,
                '당기_총수신': None,
                '당기_수익합계': None,
                '당기_비용합계': None,
                '당기_당기순이익': None,
                '당기_고정이하여신비율': None,
                '당기_자기자본비율': None,
                
                # 전년동기 데이터
                '전년동기_총자산': None,
                '전년동기_자기자본': None,
                '전년동기_총여신': None,
                '전년동기_총수신': None,
                '전년동기_수익합계': None,
                '전년동기_비용합계': None,
                '전년동기_당기순이익': None,
                '전년동기_고정이하여신비율': None,
                '전년동기_자기자본비율': None,
                
                # 증감 데이터 (계산 후 입력)
                '증감_총자산': None,
                '증감_자기자본': None,
                '증감_총여신': None,
                '증감_총수신': None,
                '증감_수익합계': None,
                '증감_비용합계': None,
                '증감_당기순이익': None,
                '증감_고정이하여신비율': None,
                '증감_자기자본비율': None,
                
                # 증감률 데이터 (%)
                '증감률_총자산': None,
                '증감률_자기자본': None,
                '증감률_총여신': None,
                '증감률_총수신': None,
                '증감률_수익합계': None,
                '증감률_비용합계': None,
                '증감률_당기순이익': None,
                '증감률_고정이하여신비율': None,
                '증감률_자기자본비율': None,
            }
            
            # 공시정보 시트에서 날짜 정보 추출
            if '공시정보' in xls.sheet_names:
                info_df = pd.read_excel(excel_path, sheet_name='공시정보')
                if '공시 날짜' in info_df.columns and not info_df['공시 날짜'].empty:
                    date_info = str(info_df['공시 날짜'].iloc[0])
                    financial_data['재무정보 날짜'] = date_info
                    
                    # 날짜에서 연도와 월 추출하여 분기 계산
                    import re
                    date_match = re.search(r'(\d{4})년(\d{1,2})월', date_info)
                    if date_match:
                        year = int(date_match.group(1))
                        month = int(date_match.group(2))
                        quarter = (month - 1) // 3 + 1
                        financial_data['분기'] = f"{year}년 {quarter}분기"
            
            # 각 시트에서 재무 데이터 추출 (당기/전년동기 구분)
            for sheet_name in xls.sheet_names:
                if sheet_name == '공시정보':
                    continue
                    
                try:
                    df = pd.read_excel(excel_path, sheet_name=sheet_name)
                    
                    # 데이터프레임이 비어있지 않은 경우에만 처리
                    if df.empty:
                        continue
                    
                    # 각 항목별 당기/전년동기 데이터 추출
                    self._extract_current_and_previous_data(df, financial_data, bank_name)
                    
                except Exception as e:
                    self.logger.log_message(f"{bank_name} 은행 {sheet_name} 시트 처리 중 오류: {str(e)}", verbose=False)
                    continue
            
            # 증감 및 증감률 계산
            self._calculate_changes_and_rates(financial_data)
            
            # 디버깅용 로그
            if financial_data['당기_총자산'] and financial_data['당기_자기자본']:
                self.logger.log_message(
                    f"{bank_name} - 당기 총자산: {financial_data['당기_총자산']:,.0f}, "
                    f"당기 자기자본: {financial_data['당기_자기자본']:,.0f}", 
                    verbose=False
                )
            
            return financial_data
            
        except Exception as e:
            self.logger.log_message(f"{bank_name} 은행 재무 데이터 추출 오류: {str(e)}")
            return None
    
    def _extract_current_and_previous_data(self, df, financial_data, bank_name):
        """DataFrame에서 당기와 전년동기 데이터를 추출합니다."""
        try:
            # 당기/전년동기를 구분하는 키워드 패턴
            current_keywords = ['당기', '현재', '2025', '2024년말', '24년말', '24.12']
            previous_keywords = ['전년동기', '전기', '2023', '2023년말', '23년말', '23.12']
            
            # 재무 항목별 키워드 매핑
            financial_items = {
                '총자산': ['총자산', '자산총계', '자산합계'],
                '자기자본': ['자기자본', '자본총계', '자본합계'],
                '총여신': ['총여신', '여신총계', '대출채권', '여신잔액'],
                '총수신': ['총수신', '수신총계', '예수금', '수신잔액'],
                '수익합계': ['수익합계', '영업수익', '총수익', '수익총계'],
                '비용합계': ['비용합계', '영업비용', '총비용', '비용총계'],
                '당기순이익': ['당기순이익', '순이익', '세후순이익'],
                '고정이하여신비율': ['고정이하여신비율', 'NPL비율', '부실채권비율'],
                '자기자본비율': ['자기자본비율', 'BIS비율', '위험가중자산에 대한 자기자본비율']
            }
            
            # DataFrame의 모든 셀을 검사하여 데이터 추출
            for idx in range(len(df)):
                for col in df.columns:
                    try:
                        cell_value = str(df.iloc[idx][col]).strip()
                        
                        # 각 재무 항목에 대해 검사
                        for item_key, item_keywords in financial_items.items():
                            for keyword in item_keywords:
                                if keyword in cell_value:
                                    # 해당 행에서 당기/전년동기 데이터 찾기
                                    self._extract_item_data(df, idx, item_key, financial_data, 
                                                          current_keywords, previous_keywords)
                                    break
                    
                    except Exception as e:
                        continue
            
        except Exception as e:
            self.logger.log_message(f"{bank_name} 데이터 추출 중 오류: {str(e)}", verbose=False)
    
    def _extract_item_data(self, df, row_idx, item_key, financial_data, current_keywords, previous_keywords):
        """특정 재무 항목의 당기/전년동기 데이터를 추출합니다."""
        try:
            # 해당 행의 모든 컬럼을 검사
            for col_idx, col in enumerate(df.columns):
                try:
                    # 컬럼 헤더에서 당기/전년동기 구분
                    col_header = str(col).strip()
                    
                    # 당기 데이터 찾기
                    is_current = any(keyword in col_header for keyword in current_keywords)
                    is_previous = any(keyword in col_header for keyword in previous_keywords)
                    
                    if is_current and financial_data[f'당기_{item_key}'] is None:
                        value = self._extract_numeric_value(df.iloc[row_idx][col])
                        if value is not None:
                            financial_data[f'당기_{item_key}'] = value
                    
                    elif is_previous and financial_data[f'전년동기_{item_key}'] is None:
                        value = self._extract_numeric_value(df.iloc[row_idx][col])
                        if value is not None:
                            financial_data[f'전년동기_{item_key}'] = value
                    
                    # 컬럼 헤더에 구분자가 없는 경우, 셀 값으로 판단
                    elif not is_current and not is_previous:
                        # 인접한 셀들을 검사하여 당기/전년동기 데이터 찾기
                        if col_idx + 1 < len(df.columns):
                            next_value = self._extract_numeric_value(df.iloc[row_idx][df.columns[col_idx + 1]])
                            current_value = self._extract_numeric_value(df.iloc[row_idx][col])
                            
                            if current_value is not None and financial_data[f'당기_{item_key}'] is None:
                                financial_data[f'당기_{item_key}'] = current_value
                            
                            if next_value is not None and financial_data[f'전년동기_{item_key}'] is None:
                                financial_data[f'전년동기_{item_key}'] = next_value
                
                except Exception as e:
                    continue
        
        except Exception as e:
            pass
    
    def _extract_numeric_value(self, cell_value):
        """셀 값에서 숫자를 추출합니다."""
        try:
            if pd.isna(cell_value) or cell_value == '' or str(cell_value).strip() == '':
                return None
            
            # 문자열로 변환
            str_value = str(cell_value).strip()
            
            # 퍼센트 제거
            str_value = str_value.replace('%', '')
            
            # 쉼표 제거
            str_value = str_value.replace(',', '')
            
            # 괄호 안의 음수 처리 (예: (1,000) -> -1000)
            if str_value.startswith('(') and str_value.endswith(')'):
                str_value = '-' + str_value[1:-1]
            
            # 숫자 추출 시도
            numeric_value = pd.to_numeric(str_value, errors='coerce')
            
            if pd.notna(numeric_value):
                # 비율 데이터인 경우 (0~100 범위)
                if 0 <= abs(numeric_value) <= 100 and ('비율' in str(cell_value) or '%' in str(cell_value)):
                    return float(numeric_value)
                # 금액 데이터인 경우 (큰 숫자)
                elif abs(numeric_value) > 100:
                    return float(numeric_value)
                # 기타 유효한 숫자
                else:
                    return float(numeric_value)
            
            return None
            
        except Exception as e:
            return None
    
    def _calculate_changes_and_rates(self, financial_data):
        """증감 및 증감률을 계산합니다."""
        try:
            # 계산할 항목 목록
            items = ['총자산', '자기자본', '총여신', '총수신', '수익합계', '비용합계', 
                    '당기순이익', '고정이하여신비율', '자기자본비율']
            
            for item in items:
                current_key = f'당기_{item}'
                previous_key = f'전년동기_{item}'
                change_key = f'증감_{item}'
                rate_key = f'증감률_{item}'
                
                current_value = financial_data.get(current_key)
                previous_value = financial_data.get(previous_key)
                
                if current_value is not None and previous_value is not None:
                    # 절대 증감 계산
                    change = current_value - previous_value
                    financial_data[change_key] = change
                    
                    # 증감률 계산 (0으로 나누기 방지)
                    if previous_value != 0:
                        rate = (change / previous_value) * 100
                        financial_data[rate_key] = rate
                    else:
                        financial_data[rate_key] = None
                else:
                    financial_data[change_key] = None
                    financial_data[rate_key] = None
        
        except Exception as e:
            self.logger.log_message(f"증감률 계산 중 오류: {str(e)}", verbose=False)
    
    def create_consolidated_financial_report(self):
        """79개 은행의 재무 데이터를 통합한 엑셀 보고서를 생성합니다."""
        try:
            self.logger.log_message("\n===== 재무 데이터 통합 보고서 생성 시작 =====")
            
            # 모든 은행의 재무 데이터를 저장할 리스트
            all_financial_data = []
            
            # 각 은행의 엑셀 파일에서 데이터 추출
            for bank in self.config.BANKS:
                # 해당 은행의 가장 최근 엑셀 파일 찾기
                bank_files = [f for f in os.listdir(self.config.output_dir) 
                             if f.startswith(f"{bank}_결산_") and f.endswith(".xlsx")]
                
                if bank_files:
                    # 가장 최근 파일 선택
                    latest_file = sorted(bank_files)[-1]
                    file_path = os.path.join(self.config.output_dir, latest_file)
                    
                    # 재무 데이터 추출
                    financial_data = self.extract_financial_data_from_excel(file_path, bank)
                    
                    if financial_data:
                        all_financial_data.append(financial_data)
                        self.logger.log_message(f"{bank} 은행 재무 데이터 추출 완료")
                    else:
                        # 데이터가 없는 경우에도 은행명만 포함
                        all_financial_data.append(self._create_empty_financial_data(bank))
                else:
                    # 파일이 없는 경우
                    all_financial_data.append(self._create_empty_financial_data(bank, "파일 없음"))
            
            # 나머지 처리는 _process_financial_data 메서드로 위임
            return self._process_financial_data(all_financial_data, self.config.output_dir)
            
        except Exception as e:
            self.logger.log_message(f"재무 데이터 통합 보고서 생성 오류: {str(e)}")
            import traceback
            self.logger.log_message(traceback.format_exc())
            return None, None
    
    def _create_empty_financial_data(self, bank_name, status="데이터 없음"):
        """빈 재무 데이터 구조를 생성합니다."""
        return {
            '은행명': bank_name,
            '재무정보 날짜': status,
            '분기': status,
            
            # 당기 데이터
            '당기_총자산': None,
            '당기_자기자본': None,
            '당기_총여신': None,
            '당기_총수신': None,
            '당기_수익합계': None,
            '당기_비용합계': None,
            '당기_당기순이익': None,
            '당기_고정이하여신비율': None,
            '당기_자기자본비율': None,
            
            # 전년동기 데이터
            '전년동기_총자산': None,
            '전년동기_자기자본': None,
            '전년동기_총여신': None,
            '전년동기_총수신': None,
            '전년동기_수익합계': None,
            '전년동기_비용합계': None,
            '전년동기_당기순이익': None,
            '전년동기_고정이하여신비율': None,
            '전년동기_자기자본비율': None,
            
            # 증감 데이터
            '증감_총자산': None,
            '증감_자기자본': None,
            '증감_총여신': None,
            '증감_총수신': None,
            '증감_수익합계': None,
            '증감_비용합계': None,
            '증감_당기순이익': None,
            '증감_고정이하여신비율': None,
            '증감_자기자본비율': None,
            
            # 증감률 데이터
            '증감률_총자산': None,
            '증감률_자기자본': None,
            '증감률_총여신': None,
            '증감률_총수신': None,
            '증감률_수익합계': None,
            '증감률_비용합계': None,
            '증감률_당기순이익': None,
            '증감률_고정이하여신비율': None,
            '증감률_자기자본비율': None,
        }
    
    def create_consolidated_financial_report_from_folder(self, folder_path):
        """지정된 폴더에서 재무 데이터를 통합한 엑셀 보고서를 생성합니다."""
        try:
            self.logger.log_message(f"\n===== 재무 데이터 통합 보고서 생성 시작 (폴더: {folder_path}) =====")
            
            # 폴더 내 모든 엑셀 파일 찾기
            excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
            self.logger.log_message(f"발견된 엑셀 파일: {len(excel_files)}개")
            
            # 모든 은행의 재무 데이터를 저장할 리스트
            all_financial_data = []
            
            # 각 엑셀 파일에서 데이터 추출
            for excel_file in excel_files:
                file_path = os.path.join(folder_path, excel_file)
                
                # 파일명에서 은행명 추출 시도
                bank_name = "알 수 없음"
                for bank in self.config.BANKS:
                    if excel_file.startswith(f"{bank}_"):
                        bank_name = bank
                        break
                
                if bank_name == "알 수 없음":
                    # 파일명의 첫 부분을 은행명으로 사용
                    bank_name = excel_file.split('_')[0] if '_' in excel_file else excel_file.replace('.xlsx', '')
                
                # 재무 데이터 추출
                financial_data = self.extract_financial_data_from_excel(file_path, bank_name)
                
                if financial_data:
                    all_financial_data.append(financial_data)
                    self.logger.log_message(f"{bank_name} 은행 재무 데이터 추출 완료")
            
            # 나머지 처리는 기존과 동일
            return self._process_financial_data(all_financial_data, folder_path)
            
        except Exception as e:
            self.logger.log_message(f"재무 데이터 통합 보고서 생성 오류: {str(e)}")
            return None, None
    
    def create_consolidated_financial_report_from_files(self, file_list):
        """선택된 파일들에서 재무 데이터를 통합한 엑셀 보고서를 생성합니다."""
        try:
            self.logger.log_message(f"\n===== 재무 데이터 통합 보고서 생성 시작 (파일 {len(file_list)}개) =====")
            
            # 모든 은행의 재무 데이터를 저장할 리스트
            all_financial_data = []
            
            # 각 선택된 파일에서 데이터 추출
            for file_path in file_list:
                excel_file = os.path.basename(file_path)
                
                # 파일명에서 은행명 추출 시도
                bank_name = "알 수 없음"
                for bank in self.config.BANKS:
                    if excel_file.startswith(f"{bank}_"):
                        bank_name = bank
                        break
                
                if bank_name == "알 수 없음":
                    # 파일명의 첫 부분을 은행명으로 사용
                    bank_name = excel_file.split('_')[0] if '_' in excel_file else excel_file.replace('.xlsx', '')
                
                # 재무 데이터 추출
                financial_data = self.extract_financial_data_from_excel(file_path, bank_name)
                
                if financial_data:
                    all_financial_data.append(financial_data)
                    self.logger.log_message(f"{bank_name} 은행 재무 데이터 추출 완료 ({excel_file})")
            
            # 출력 폴더 결정 (첫 번째 파일의 폴더 사용)
            output_folder = os.path.dirname(file_list[0]) if file_list else self.config.output_dir
            
            # 나머지 처리는 기존과 동일
            return self._process_financial_data(all_financial_data, output_folder)
            
        except Exception as e:
            self.logger.log_message(f"재무 데이터 통합 보고서 생성 오류: {str(e)}")
            return None, None
    
    def create_consolidated_financial_report_md(self, all_financial_data, output_folder, is_settlement=False):
        """통합 재무 데이터를 마크다운 파일로 저장합니다. (v2.5 - 당기/전년동기 비교)"""
        try:
            file_suffix = "결산" if is_settlement else "분기"
            md_file = os.path.join(output_folder, f'저축은행_{file_suffix}_재무데이터_통합_비교분석_{self.config.today}.md')
            
            # DataFrame 생성
            consolidated_df = pd.DataFrame(all_financial_data)
            
            with open(md_file, 'w', encoding='utf-8') as f:
                # 헤더
                f.write(f"# 🏦 저축은행 {file_suffix}공시 재무데이터 통합 비교분석 보고서\n\n")
                f.write(f"## 📊 요약 정보\n\n")
                
                # 통계 계산
                total_banks = len(consolidated_df)
                banks_with_data = len(consolidated_df[consolidated_df['재무정보 날짜'] != '데이터 없음'])
                
                f.write(f"- **📅 보고서 생성일**: {datetime.now().strftime('%Y년 %m월 %d일 %H:%M:%S')}\n")
                f.write(f"- **🏦 전체 은행 수**: {total_banks}개\n")
                f.write(f"- **📈 데이터 보유 은행**: {banks_with_data}개\n")
                f.write(f"- **🔄 비교 분석**: 당기 vs 전년동기 증감 분석\n")
                
                # 평균 지표 계산
                if '당기_자기자본비율' in consolidated_df.columns:
                    avg_bis = consolidated_df['당기_자기자본비율'].mean()
                    if not pd.isna(avg_bis):
                        f.write(f"- **💪 평균 자기자본비율 (당기)**: {avg_bis:.2f}%\n")
                
                if '증감률_총자산' in consolidated_df.columns:
                    avg_asset_growth = consolidated_df['증감률_총자산'].mean()
                    if not pd.isna(avg_asset_growth):
                        growth_emoji = "📈" if avg_asset_growth > 0 else "📉" if avg_asset_growth < 0 else "➡️"
                        f.write(f"- **{growth_emoji} 평균 총자산 증감률**: {avg_asset_growth:.2f}%\n")
                
                f.write(f"\n## 🚀 주요 성과 분석\n\n")
                
                # 총자산 증가율 상위 5개 은행
                if '증감률_총자산' in consolidated_df.columns:
                    asset_growth_data = consolidated_df[consolidated_df['증감률_총자산'].notna()]
                    if not asset_growth_data.empty:
                        top_asset_growth = asset_growth_data.nlargest(5, '증감률_총자산')
                        
                        f.write("### 📈 총자산 증가율 상위 5개 은행\n\n")
                        f.write("| 순위 | 은행명 | 당기 총자산 | 전년동기 총자산 | 증감률 | 트렌드 |\n")
                        f.write("| --- | --- | --- | --- | --- | --- |\n")
                        
                        for i, (_, row) in enumerate(top_asset_growth.iterrows()):
                            rank = i + 1
                            bank = row['은행명']
                            current = f"{row['당기_총자산']:,.0f}" if pd.notna(row['당기_총자산']) else 'N/A'
                            previous = f"{row['전년동기_총자산']:,.0f}" if pd.notna(row['전년동기_총자산']) else 'N/A'
                            rate = f"{row['증감률_총자산']:+.2f}%" if pd.notna(row['증감률_총자산']) else 'N/A'
                            trend = "🔥" if row['증감률_총자산'] > 10 else "📈" if row['증감률_총자산'] > 0 else "📉"
                            
                            f.write(f"| {rank} | {bank} | {current} | {previous} | {rate} | {trend} |\n")
                        f.write("\n")
                
                # 자기자본 증가율 상위 5개 은행
                if '증감률_자기자본' in consolidated_df.columns:
                    capital_growth_data = consolidated_df[consolidated_df['증감률_자기자본'].notna()]
                    if not capital_growth_data.empty:
                        top_capital_growth = capital_growth_data.nlargest(5, '증감률_자기자본')
                        
                        f.write("### 💪 자기자본 증가율 상위 5개 은행\n\n")
                        f.write("| 순위 | 은행명 | 당기 자기자본 | 전년동기 자기자본 | 증감률 | 트렌드 |\n")
                        f.write("| --- | --- | --- | --- | --- | --- |\n")
                        
                        for i, (_, row) in enumerate(top_capital_growth.iterrows()):
                            rank = i + 1
                            bank = row['은행명']
                            current = f"{row['당기_자기자본']:,.0f}" if pd.notna(row['당기_자기자본']) else 'N/A'
                            previous = f"{row['전년동기_자기자본']:,.0f}" if pd.notna(row['전년동기_자기자본']) else 'N/A'
                            rate = f"{row['증감률_자기자본']:+.2f}%" if pd.notna(row['증감률_자기자본']) else 'N/A'
                            trend = "💎" if row['증감률_자기자본'] > 15 else "💪" if row['증감률_자기자본'] > 0 else "⚠️"
                            
                            f.write(f"| {rank} | {bank} | {current} | {previous} | {rate} | {trend} |\n")
                        f.write("\n")
                
                # 당기순이익 상위 5개 은행
                if '당기_당기순이익' in consolidated_df.columns:
                    profit_data = consolidated_df[consolidated_df['당기_당기순이익'].notna()]
                    if not profit_data.empty:
                        top_profit = profit_data.nlargest(5, '당기_당기순이익')
                        
                        f.write("### 🏆 당기순이익 상위 5개 은행\n\n")
                        f.write("| 순위 | 은행명 | 당기순이익 | 전년동기 순이익 | 증감률 | 성과 |\n")
                        f.write("| --- | --- | --- | --- | --- | --- |\n")
                        
                        for i, (_, row) in enumerate(top_profit.iterrows()):
                            rank = i + 1
                            bank = row['은행명']
                            current = f"{row['당기_당기순이익']:,.0f}" if pd.notna(row['당기_당기순이익']) else 'N/A'
                            previous = f"{row['전년동기_당기순이익']:,.0f}" if pd.notna(row['전년동기_당기순이익']) else 'N/A'
                            rate = f"{row['증감률_당기순이익']:+.2f}%" if pd.notna(row['증감률_당기순이익']) else 'N/A'
                            performance = "🏆" if pd.notna(row['당기_당기순이익']) and row['당기_당기순이익'] > 50000 else "⭐" if pd.notna(row['당기_당기순이익']) and row['당기_당기순이익'] > 0 else "📉"
                            
                            f.write(f"| {rank} | {bank} | {current} | {previous} | {rate} | {performance} |\n")
                        f.write("\n")
                
                # 전체 비교 데이터 테이블
                f.write("## 📋 전체 재무현황 비교 (당기 vs 전년동기)\n\n")
                f.write("### 📊 주요 지표 요약\n\n")
                
                # 간소화된 요약 테이블
                summary_columns = ['은행명', '당기_총자산', '전년동기_총자산', '증감률_총자산', 
                                 '당기_자기자본', '전년동기_자기자본', '증감률_자기자본',
                                 '당기_당기순이익', '증감률_당기순이익']
                
                available_columns = [col for col in summary_columns if col in consolidated_df.columns]
                summary_df = consolidated_df[available_columns].copy()
                
                # 마크다운 테이블 생성
                if not summary_df.empty:
                    headers = []
                    for col in summary_df.columns:
                        if col == '은행명':
                            headers.append('은행명')
                        elif '당기_' in col:
                            headers.append(col.replace('당기_', '당기 '))
                        elif '전년동기_' in col:
                            headers.append(col.replace('전년동기_', '전년 '))
                        elif '증감률_' in col:
                            headers.append(col.replace('증감률_', '증감률(%) '))
                        else:
                            headers.append(col)
                    
                    f.write('| ' + ' | '.join(headers) + ' |\n')
                    f.write('|' + '|'.join([' --- ' for _ in headers]) + '|\n')
                    
                    # 데이터 행 (상위 20개만 표시)
                    display_df = summary_df.head(20)
                    for _, row in display_df.iterrows():
                        row_data = []
                        for col, value in zip(summary_df.columns, row):
                            if pd.isna(value):
                                formatted_value = ''
                            elif '증감률_' in col:
                                try:
                                    rate_val = float(value)
                                    emoji = "📈" if rate_val > 0 else "📉" if rate_val < 0 else "➡️"
                                    formatted_value = f"{emoji} {rate_val:+.1f}%"
                                except:
                                    formatted_value = str(value)
                            elif col in ['당기_총자산', '전년동기_총자산', '당기_자기자본', '전년동기_자기자본', 
                                       '당기_당기순이익', '전년동기_당기순이익']:
                                try:
                                    formatted_value = f"{int(value):,}"
                                except:
                                    formatted_value = str(value)
                            else:
                                formatted_value = str(value)
                            
                            row_data.append(formatted_value.replace('|', '\\|'))
                        
                        f.write('| ' + ' | '.join(row_data) + ' |\n')
                    
                    if len(summary_df) > 20:
                        f.write(f"\n⚠️ *({len(summary_df) - 20}개 은행 더 있음... 전체 데이터는 엑셀 파일을 참조하세요.)*\n")
                
                f.write('\n')
                
                # 시장 동향 분석
                f.write("## 🌟 시장 동향 분석\n\n")
                
                # 성장세 분석
                growth_banks = len(consolidated_df[(consolidated_df['증감률_총자산'] > 0) & consolidated_df['증감률_총자산'].notna()])
                decline_banks = len(consolidated_df[(consolidated_df['증감률_총자산'] < 0) & consolidated_df['증감률_총자산'].notna()])
                
                f.write(f"### 📈 성장 현황\n\n")
                f.write(f"- **성장 은행**: {growth_banks}개 (총자산 증가)\n")
                f.write(f"- **감소 은행**: {decline_banks}개 (총자산 감소)\n")
                
                if growth_banks > decline_banks:
                    f.write(f"- **시장 전망**: 🟢 긍정적 (성장 은행이 더 많음)\n")
                elif decline_banks > growth_banks:
                    f.write(f"- **시장 전망**: 🔴 주의 필요 (감소 은행이 더 많음)\n")
                else:
                    f.write(f"- **시장 전망**: 🟡 혼재 상황\n")
                f.write("\n")
                
                # 푸터
                f.write("---\n\n")
                f.write("### 💡 분석 활용 가이드\n\n")
                f.write("1. **📈 증감률 해석**: 양수는 성장, 음수는 감소를 의미합니다.\n")
                f.write("2. **🏆 순위 기준**: 각 지표별 절대값과 증감률을 종합 고려하세요.\n")
                f.write("3. **⚠️ 데이터 한계**: 일부 은행의 데이터가 누락될 수 있습니다.\n")
                f.write("4. **📊 상세 분석**: 전체 데이터는 함께 생성된 엑셀 파일을 참조하세요.\n\n")
                
                f.write("### 📁 관련 파일\n\n")
                f.write(f"- **📊 상세 엑셀**: `저축은행_{file_suffix}_재무데이터_통합_{self.config.today}.xlsx`\n")
                f.write(f"- **📋 스크래핑 요약**: `저축은행_{file_suffix}공시_스크래핑_요약_{self.config.today}.md`\n\n")
                
                f.write(f"*이 비교분석 보고서는 {datetime.now().strftime('%Y년 %m월 %d일 %H:%M:%S')}에 자동 생성되었습니다.*\n")
            
            self.logger.log_message(f"MD 통합 비교분석 보고서 저장 완료: {md_file}")
            return md_file
            
        except Exception as e:
            self.logger.log_message(f"MD 통합 보고서 생성 오류: {str(e)}")
            return None
    
    def _process_financial_data(self, all_financial_data, output_folder):
        """수집된 재무 데이터를 처리하여 통합 보고서 생성 (공통 로직 - v2.5 확장)"""
        try:
            # DataFrame 생성
            consolidated_df = pd.DataFrame(all_financial_data)
            
            # 열 순서 정렬 (당기/전년동기/증감 구조)
            column_order = [
                '은행명', '재무정보 날짜', '분기',
                
                # 당기 데이터
                '당기_총자산', '당기_자기자본', '당기_총여신', '당기_총수신', 
                '당기_수익합계', '당기_비용합계', '당기_당기순이익', 
                '당기_고정이하여신비율', '당기_자기자본비율',
                
                # 전년동기 데이터
                '전년동기_총자산', '전년동기_자기자본', '전년동기_총여신', '전년동기_총수신',
                '전년동기_수익합계', '전년동기_비용합계', '전년동기_당기순이익',
                '전년동기_고정이하여신비율', '전년동기_자기자본비율',
                
                # 증감 데이터
                '증감_총자산', '증감_자기자본', '증감_총여신', '증감_총수신',
                '증감_수익합계', '증감_비용합계', '증감_당기순이익',
                '증감_고정이하여신비율', '증감_자기자본비율',
                
                # 증감률 데이터
                '증감률_총자산', '증감률_자기자본', '증감률_총여신', '증감률_총수신',
                '증감률_수익합계', '증감률_비용합계', '증감률_당기순이익',
                '증감률_고정이하여신비율', '증감률_자기자본비율'
            ]
            
            # 존재하는 열만 선택
            existing_columns = [col for col in column_order if col in consolidated_df.columns]
            consolidated_df = consolidated_df[existing_columns]
            
            # 엑셀 파일로 저장 (결산공시용 파일명 - v2.5)
            output_file = os.path.join(output_folder, 
                                      f'저축은행_결산_재무데이터_통합_비교분석_{self.config.today}.xlsx')
            
            # 확장된 엑셀 저장 (4개 시트)
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # 시트 1: 전체 비교 데이터
                consolidated_df.to_excel(writer, sheet_name='전체_재무현황_비교', index=False)
                
                # 시트 2: 요약 통계
                self._create_summary_statistics_sheet(writer, consolidated_df)
                
                # 시트 3: 당기 재무현황만
                self._create_current_period_sheet(writer, consolidated_df)
                
                # 시트 4: 증감 분석
                self._create_change_analysis_sheet(writer, consolidated_df)
                
                # 서식 설정
                self._format_excel_sheets(writer, consolidated_df)
            
            # MD 파일로도 저장 (결산공시로 설정 - 비교분석 버전)
            md_output_file = self.create_consolidated_financial_report_md(
                all_financial_data, output_folder, is_settlement=True
            )
            
            # 통계 요약
            stats = {
                '처리된 파일': len(all_financial_data),
                '데이터 있는 은행': len([d for d in all_financial_data if d.get('재무정보 날짜', '') not in ['데이터 없음', '파일 없음', '']]),
                '평균 총자산 증감률': consolidated_df['증감률_총자산'].mean() if '증감률_총자산' in consolidated_df.columns else 0,
                '평균 자기자본 증감률': consolidated_df['증감률_자기자본'].mean() if '증감률_자기자본' in consolidated_df.columns else 0,
                '성장 은행 수': len(consolidated_df[(consolidated_df['증감률_총자산'] > 0) & consolidated_df['증감률_총자산'].notna()]) if '증감률_총자산' in consolidated_df.columns else 0
            }
            
            self.logger.log_message(f"\n통합 비교분석 보고서 저장 완료: {output_file}")
            if md_output_file:
                self.logger.log_message(f"MD 통합 비교분석 보고서 저장 완료: {md_output_file}")
            self.logger.log_message(f"처리된 파일: {stats['처리된 파일']}개")
            self.logger.log_message(f"데이터 있는 은행: {stats['데이터 있는 은행']}개")
            self.logger.log_message(f"성장 은행: {stats['성장 은행 수']}개")
            
            return output_file, consolidated_df
            
        except Exception as e:
            self.logger.log_message(f"재무 데이터 처리 오류: {str(e)}")
            import traceback
            self.logger.log_message(traceback.format_exc())
            return None, None
    
    def _create_summary_statistics_sheet(self, writer, consolidated_df):
        """요약 통계 시트를 생성합니다."""
        try:
            # 기본 통계
            stats_data = []
            
            # 전체 은행 수
            stats_data.append(['전체 은행 수', len(consolidated_df)])
            stats_data.append(['데이터 보유 은행', len(consolidated_df[consolidated_df['재무정보 날짜'] != '데이터 없음'])])
            
            # 평균 지표들
            numeric_columns = consolidated_df.select_dtypes(include=[float, int]).columns
            for col in numeric_columns:
                if '당기_' in col:
                    avg_val = consolidated_df[col].mean()
                    if pd.notna(avg_val):
                        item_name = col.replace('당기_', '')
                        if '비율' in item_name:
                            stats_data.append([f'평균 {item_name} (당기)', f"{avg_val:.2f}%"])
                        else:
                            stats_data.append([f'평균 {item_name} (당기)', f"{avg_val:,.0f}"])
                
                elif '증감률_' in col:
                    avg_rate = consolidated_df[col].mean()
                    if pd.notna(avg_rate):
                        item_name = col.replace('증감률_', '')
                        stats_data.append([f'평균 {item_name} 증감률', f"{avg_rate:+.2f}%"])
            
            # 성장 현황
            if '증감률_총자산' in consolidated_df.columns:
                growth_count = len(consolidated_df[(consolidated_df['증감률_총자산'] > 0) & consolidated_df['증감률_총자산'].notna()])
                decline_count = len(consolidated_df[(consolidated_df['증감률_총자산'] < 0) & consolidated_df['증감률_총자산'].notna()])
                stats_data.append(['총자산 성장 은행 수', growth_count])
                stats_data.append(['총자산 감소 은행 수', decline_count])
            
            # DataFrame으로 변환 후 저장
            stats_df = pd.DataFrame(stats_data, columns=['지표', '값'])
            stats_df.to_excel(writer, sheet_name='요약_통계', index=False)
            
        except Exception as e:
            self.logger.log_message(f"요약 통계 시트 생성 오류: {str(e)}", verbose=False)
    
    def _create_current_period_sheet(self, writer, consolidated_df):
        """당기 재무현황만 별도 시트로 생성합니다."""
        try:
            # 당기 데이터만 추출
            current_columns = ['은행명', '재무정보 날짜', '분기']
            current_columns.extend([col for col in consolidated_df.columns if col.startswith('당기_')])
            
            current_df = consolidated_df[current_columns].copy()
            
            # 컬럼명 정리 (당기_ 접두사 제거)
            new_columns = {}
            for col in current_df.columns:
                if col.startswith('당기_'):
                    new_columns[col] = col.replace('당기_', '')
            
            current_df = current_df.rename(columns=new_columns)
            current_df.to_excel(writer, sheet_name='당기_재무현황', index=False)
            
        except Exception as e:
            self.logger.log_message(f"당기 현황 시트 생성 오류: {str(e)}", verbose=False)
    
    def _create_change_analysis_sheet(self, writer, consolidated_df):
        """증감 분석 시트를 생성합니다."""
        try:
            # 증감 관련 컬럼만 추출
            change_columns = ['은행명']
            change_columns.extend([col for col in consolidated_df.columns if col.startswith('증감률_')])
            change_columns.extend([col for col in consolidated_df.columns if col.startswith('증감_')])
            
            change_df = consolidated_df[change_columns].copy()
            
            # 증감률 기준으로 정렬 (총자산 증감률 기준)
            if '증감률_총자산' in change_df.columns:
                change_df = change_df.sort_values('증감률_총자산', ascending=False, na_position='last')
            
            change_df.to_excel(writer, sheet_name='증감_분석', index=False)
            
        except Exception as e:
            self.logger.log_message(f"증감 분석 시트 생성 오류: {str(e)}", verbose=False)
    
    def _format_excel_sheets(self, writer, consolidated_df):
        """엑셀 시트들의 서식을 설정합니다."""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            
            workbook = writer.book
            
            # 각 시트별 서식 설정
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # 헤더 스타일
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                center_align = Alignment(horizontal="center", vertical="center")
                
                # 첫 번째 행 (헤더) 서식 적용
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                
                # 열 너비 자동 조정
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 30)  # 최대 30으로 제한
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # 숫자 포맷 설정 (시트별 차별화)
                if sheet_name == '전체_재무현황_비교':
                    self._apply_comparison_format(worksheet, consolidated_df)
                elif sheet_name == '증감_분석':
                    self._apply_change_format(worksheet)
        
        except Exception as e:
            self.logger.log_message(f"엑셀 서식 설정 중 오류: {str(e)}", verbose=False)
    
    def _apply_comparison_format(self, worksheet, consolidated_df):
        """비교 시트의 숫자 포맷을 적용합니다."""
        try:
            # 금액 컬럼과 비율 컬럼 구분
            for col_idx, col_name in enumerate(consolidated_df.columns):
                col_letter = chr(65 + col_idx)  # A, B, C, ...
                
                if any(keyword in col_name for keyword in ['총자산', '자기자본', '총여신', '총수신', '수익', '비용', '순이익']):
                    # 금액 포맷 (천단위 구분)
                    for row in range(2, len(consolidated_df) + 2):
                        try:
                            cell = worksheet[f'{col_letter}{row}']
                            if cell.value is not None:
                                cell.number_format = '#,##0'
                        except:
                            pass
                
                elif any(keyword in col_name for keyword in ['비율', '증감률']):
                    # 퍼센트 포맷
                    for row in range(2, len(consolidated_df) + 2):
                        try:
                            cell = worksheet[f'{col_letter}{row}']
                            if cell.value is not None:
                                cell.number_format = '0.00'
                        except:
                            pass
        
        except Exception as e:
            pass
    
    def _apply_change_format(self, worksheet):
        """증감 분석 시트의 포맷을 적용합니다."""
        try:
            # 증감률 컬럼에 조건부 서식 적용 (색상으로 증감 표시)
            from openpyxl.styles import PatternFill
            
            # 양수는 녹색, 음수는 빨간색 배경
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        if cell.value > 0:
                            cell.fill = green_fill
                        elif cell.value < 0:
                            cell.fill = red_fill
        
        except Exception as e:
            pass
    
    def create_zip_file(self):
        """결과 디렉토리를 ZIP 파일로 압축합니다."""
        try:
            self.logger.log_message("\n데이터 압축 중...")
            zip_filename = os.path.join(os.path.dirname(self.config.output_dir), 
                                      f'저축은행_결산공시_데이터_{self.config.today}.zip')
            
            with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for root, dirs, files in os.walk(self.config.output_dir):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arcname = os.path.relpath(file_path, os.path.dirname(self.config.output_dir))
                        zipf.write(file_path, arcname)
            
            self.logger.log_message(f"압축 파일 생성 완료: {zip_filename}")
            return zip_filename
        except Exception as e:
            self.logger.log_message(f"압축 파일 생성 오류: {str(e)}")
            return None


# GUI 클래스 (탭 버전 - v2.6 업데이트)
class SettlementScraperTab:
    def __init__(self, parent):
        self.parent = parent
        self.frame = ttk.Frame(parent)  # 탭용 프레임 생성
        
        # 구성 설정
        self.config = Config()
        self.logger = Logger(self.config, self)
        self.progress_status = {}  # 은행별 진행 상태 저장
        
        # 실행 상태 변수
        self.running = False
        self.scraper = None
        self.driver_manager = None
        self.progress_manager = None
        
        # 자동 압축 변수 초기화
        self.auto_zip_var = tk.BooleanVar(value=self.config.auto_zip)
        
        # MD 생성 옵션 추가
        self.save_md_var = tk.BooleanVar(value=False)
        
        # 메인 프레임
        self.create_widgets()
        
        # 초기 은행 목록 로드
        self.load_bank_list()
        
        # 설정 로드
        self.load_settings()
    
    def create_widgets(self):
        """GUI 위젯을 생성합니다."""
        # 메인 프레임 설정 - self.frame을 부모로 사용
        self.main_frame = ttk.Frame(self.frame, padding="10")
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 상단 프레임 (설정)
        self.settings_frame = ttk.LabelFrame(self.main_frame, text="설정", padding="5")
        self.settings_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # 출력 디렉토리 선택
        ttk.Label(self.settings_frame, text="출력 디렉토리:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.output_dir_var = tk.StringVar(value=self.config.output_dir)
        ttk.Entry(self.settings_frame, textvariable=self.output_dir_var, width=50).grid(row=0, column=1, sticky=tk.W+tk.E, padx=5, pady=5)
        ttk.Button(self.settings_frame, text="찾아보기", command=self.browse_output_dir).grid(row=0, column=2, padx=5, pady=5)
        
        # ChromeDriver 경로 선택
        ttk.Label(self.settings_frame, text="ChromeDriver 경로:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        self.chrome_driver_path_var = tk.StringVar(value=self.config.chrome_driver_path if self.config.chrome_driver_path else "")
        ttk.Entry(self.settings_frame, textvariable=self.chrome_driver_path_var, width=50).grid(row=1, column=1, sticky=tk.W+tk.E, padx=5, pady=5)
        ttk.Button(self.settings_frame, text="찾아보기", command=self.browse_chrome_driver_path).grid(row=1, column=2, padx=5, pady=5)
        
        # 작업자 수 설정
        ttk.Label(self.settings_frame, text="병렬 작업자 수:").grid(row=2, column=0, sticky=tk.W, padx=5, pady=5)
        self.workers_var = tk.IntVar(value=self.config.MAX_WORKERS)
        worker_spinbox = ttk.Spinbox(self.settings_frame, from_=1, to=10, textvariable=self.workers_var, width=5)
        worker_spinbox.grid(row=2, column=1, sticky=tk.W, padx=5, pady=5)
        
        # 자동 압축 옵션
        ttk.Checkbutton(self.settings_frame, text="스크래핑 완료 후 자동 압축", variable=self.auto_zip_var).grid(row=2, column=2, sticky=tk.W, padx=5, pady=5)
        
        # MD 생성 옵션 추가
        ttk.Checkbutton(self.settings_frame, text="📝 MD 파일도 함께 생성", 
                       variable=self.save_md_var).grid(row=3, column=0, sticky=tk.W, padx=5, pady=5)
        
        # 중앙 프레임 (은행 선택)
        self.bank_frame = ttk.LabelFrame(self.main_frame, text="은행 선택", padding="5")
        self.bank_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 은행 선택 버튼 및 체크박스
        self.select_all_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(self.bank_frame, text="전체 선택", variable=self.select_all_var, command=self.toggle_all_banks).pack(anchor=tk.W, padx=5, pady=5)
        
        # 은행 목록 표시 영역 (스크롤 가능)
        self.bank_list_frame = ttk.Frame(self.bank_frame)
        self.bank_list_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 스크롤바
        scrollbar = ttk.Scrollbar(self.bank_list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 트리뷰 (은행 목록) - 수정된 부분
        self.bank_tree = ttk.Treeview(self.bank_list_frame, columns=("bank", "status"), show="headings", yscrollcommand=scrollbar.set)
        self.bank_tree.heading("bank", text="은행명")
        self.bank_tree.heading("status", text="상태")
        self.bank_tree.column("bank", width=200)
        self.bank_tree.column("status", width=100, anchor=tk.CENTER)
        self.bank_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar.config(command=self.bank_tree.yview)
        
        # 하단 프레임 (로그 및 버튼)
        self.bottom_frame = ttk.Frame(self.main_frame)
        self.bottom_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 로그 표시 영역
        log_frame = ttk.LabelFrame(self.bottom_frame, text="로그")
        log_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, wrap=tk.WORD, width=80, height=10)
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.log_text.config(state=tk.DISABLED)
        
        # 버튼 프레임 (v2.6 - MD 통합 버튼 추가)
        button_frame = ttk.Frame(self.main_frame)
        button_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # 첫 번째 행 - 주요 기능
        main_buttons_frame = ttk.Frame(button_frame)
        main_buttons_frame.pack(fill=tk.X, pady=2)
        
        self.start_button = ttk.Button(main_buttons_frame, text="스크래핑 시작", command=self.start_scraping)
        self.start_button.pack(side=tk.LEFT, padx=5)
        
        self.stop_button = ttk.Button(main_buttons_frame, text="중지", command=self.stop_scraping, state=tk.DISABLED)
        self.stop_button.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(main_buttons_frame, text="설정 저장", command=self.save_settings).pack(side=tk.LEFT, padx=5)
        ttk.Button(main_buttons_frame, text="결과 폴더 열기", command=self.open_output_folder).pack(side=tk.LEFT, padx=5)
        
        # 두 번째 행 - 보고서 기능
        report_buttons_frame = ttk.Frame(button_frame)
        report_buttons_frame.pack(fill=tk.X, pady=2)
        
        ttk.Button(report_buttons_frame, text="요약 보고서 생성", command=self.generate_report).pack(side=tk.LEFT, padx=5)
        ttk.Button(report_buttons_frame, text="📝 MD 요약 보고서", 
                   command=self.generate_md_summary_report).pack(side=tk.LEFT, padx=5)
        
        # v2.5 - 기존 엑셀 통합 재무 보고서 버튼
        ttk.Button(report_buttons_frame, text="📊 통합 재무 보고서 (당기/전년 비교)", 
                   command=self.create_financial_consolidation_with_selection).pack(side=tk.LEFT, padx=5)
        
        # 세 번째 행 - 기타 기능 (v2.6 - MD 통합 기능 추가)
        other_buttons_frame = ttk.Frame(button_frame)
        other_buttons_frame.pack(fill=tk.X, pady=2)
        
        # v2.6 - 새로운 MD 통합 보고서 버튼
        ttk.Button(other_buttons_frame, text="📝 MD 통합 보고서 생성", 
                   command=self.create_md_consolidation_with_selection).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(other_buttons_frame, text="데이터 압축 및 다운로드", command=self.compress_and_download).pack(side=tk.LEFT, padx=5)
        ttk.Button(other_buttons_frame, text="진행 상태 초기화", command=self.reset_progress).pack(side=tk.LEFT, padx=5)
    
    # v2.6 - MD 통합 기능 추가
    def create_md_consolidation_with_selection(self):
        """MD 파일 통합 보고서를 생성합니다. (v2.6 새 기능)"""
        try:
            # 스크래퍼 초기화
            if self.scraper is None:
                self.progress_manager = ProgressManager(self.config, self.logger)
                self.scraper = BankScraper(self.config, self.logger, None, self.progress_manager)
            
            # MD 소스 선택 대화상자 표시
            source_dialog = MDSourceDialog(self.parent, self.config)
            self.parent.wait_window(source_dialog.dialog)
            
            if source_dialog.result:
                source_type, source_path = source_dialog.result
                
                # 상태 업데이트
                self.update_log(f"📝 MD 파일 통합 보고서 생성 시작: {source_type} - {source_path}")
                
                # 별도 스레드에서 실행
                threading.Thread(
                    target=self._run_md_consolidation, 
                    args=(source_type, source_path),
                    daemon=True
                ).start()
            
        except Exception as e:
            messagebox.showerror("오류", f"MD 파일 통합 중 오류 발생: {str(e)}")
    
    def _run_md_consolidation(self, source_type, source_path):
        """MD 파일 통합 실행 (v2.6)"""
        try:
            # 소스 타입에 따른 처리
            if source_type == "default_folder":
                # 기본 출력 폴더에서 모든 MD 파일 통합
                output_file = self.scraper.create_consolidated_md_report()
                
            elif source_type == "custom_folder":
                # 사용자 지정 폴더에서 MD 파일 통합
                output_file = self.scraper.create_consolidated_md_report_from_folder(source_path)
                
            elif source_type == "selected_files":
                # 선택한 MD 파일들만 통합
                output_file = self.scraper.create_consolidated_md_report_from_files(source_path)
            
            if output_file and os.path.exists(output_file):
                # 성공 메시지
                self.frame.after(0, lambda: self.update_log(f"📝 MD 통합 보고서 생성 완료: {output_file}"))
                self.frame.after(0, lambda: messagebox.showinfo("완료", f"📝 MD 파일 통합 보고서가 생성되었습니다!\n\n✅ 모든 은행의 MD 데이터를 하나로 통합\n📊 데이터 품질 분석 포함\n📈 카테고리별 완성도 리포트\n\n{os.path.basename(output_file)}"))
                
                # MD 파일 열기 여부 확인
                self.frame.after(0, lambda: self._ask_open_md_file(output_file))
            else:
                self.frame.after(0, lambda: messagebox.showerror("오류", "MD 파일 통합 보고서 생성에 실패했습니다."))
                
        except Exception as e:
            self.frame.after(0, lambda: messagebox.showerror("오류", f"MD 파일 통합 중 오류 발생: {str(e)}"))
    
    def _ask_open_md_file(self, md_file_path):
        """MD 파일 열기 여부를 묻고 열기"""
        if messagebox.askyesno("파일 열기", "생성된 MD 통합 보고서를 열어보시겠습니까?"):
            self.open_md_file(md_file_path)
    
    def browse_chrome_driver_path(self):
        """ChromeDriver 파일을 선택합니다."""
        filetypes = []
        if sys.platform == 'win32':
            filetypes = [("ChromeDriver", "*.exe"), ("모든 파일", "*.*")]
        else:
            filetypes = [("ChromeDriver", "*"), ("모든 파일", "*.*")]
            
        file_path = filedialog.askopenfilename(
            title="ChromeDriver 선택",
            filetypes=filetypes
        )
        
        if file_path:
            self.chrome_driver_path_var.set(file_path)
            self.config.update_chrome_driver_path(file_path)
            self.logger.log_message(f"ChromeDriver 경로 변경: {file_path}")
    
    def save_settings(self):
        """설정을 저장합니다."""
        # 출력 디렉토리 업데이트
        try:
            self.config.update_output_dir(self.output_dir_var.get())
        except Exception as e:
            messagebox.showerror("오류", f"출력 디렉토리 설정 실패: {str(e)}")
            return
        
        # ChromeDriver 경로 업데이트
        chrome_driver_path = self.chrome_driver_path_var.get()
        if chrome_driver_path and os.path.exists(chrome_driver_path):
            self.config.update_chrome_driver_path(chrome_driver_path)
        elif not chrome_driver_path:
            self.config.update_chrome_driver_path(None)
        
        # 작업자 수 업데이트
        self.config.MAX_WORKERS = self.workers_var.get()
        
        # 자동 압축 옵션 업데이트
        self.config.update_auto_zip(self.auto_zip_var.get())
        
        # 설정 저장
        self.config.save_settings()
        
        messagebox.showinfo("설정", "설정이 저장되었습니다.")
        self.logger.log_message("설정 저장 완료")
    
    def load_bank_list(self):
        """은행 목록을 로드하고 UI에 표시합니다."""
        # 트리뷰 항목 삭제
        for item in self.bank_tree.get_children():
            self.bank_tree.delete(item)
        
        # 은행 추가 - v2.3에서 수정된 트리뷰 구조 사용
        for bank in self.config.BANKS:
            # insert 메서드에서 values 파라미터 사용
            self.bank_tree.insert("", tk.END, iid=bank, values=(bank, "대기 중"))
            self.progress_status[bank] = "대기 중"
        
        self.update_log("은행 목록 로드 완료")
    
    def load_settings(self):
        """설정을 로드합니다."""
        # 이미 완료된 은행의 상태 업데이트
        self.progress_manager = ProgressManager(self.config, self.logger)
        completed_banks = self.progress_manager.progress.get('completed', [])
        failed_banks = self.progress_manager.progress.get('failed', [])
        
        for bank in completed_banks:
            self.update_bank_status(bank, "완료")
        
        for bank in failed_banks:
            self.update_bank_status(bank, "실패")
        
        self.update_log(f"설정 로드 완료. 출력 디렉토리: {self.config.output_dir}")
        self.update_log(f"완료된 은행: {len(completed_banks)}개, 실패한 은행: {len(failed_banks)}개")
    
    def browse_output_dir(self):
        """출력 디렉토리를 선택합니다."""
        directory = filedialog.askdirectory(initialdir=self.config.output_dir)
        if directory:
            self.output_dir_var.set(directory)
            try:
                self.config.update_output_dir(directory)
                self.logger.log_message(f"출력 디렉토리 변경: {directory}")
            except Exception as e:
                messagebox.showerror("오류", f"디렉토리 변경 실패: {str(e)}")
    
    def toggle_all_banks(self):
        """모든 은행을 선택하거나 선택 해제합니다."""
        select_all = self.select_all_var.get()
        for item in self.bank_tree.get_children():
            if select_all:
                self.bank_tree.selection_add(item)
            else:
                self.bank_tree.selection_remove(item)
    
    def update_log(self, message):
        """로그 메시지를 추가합니다."""
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, f"{message}\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
        # self.parent.update_idletasks() 대신 상위 위젯 업데이트
        try:
            self.frame.update_idletasks()
        except:
            pass
    
    def update_bank_status(self, bank_name, status):
        """은행의 상태를 업데이트합니다."""
        # 상태 저장
        self.progress_status[bank_name] = status
        
        # UI 업데이트 - v2.3에서 수정된 트리뷰 구조 사용
        try:
            if self.bank_tree.exists(bank_name):
                self.bank_tree.item(bank_name, values=(bank_name, status))
                self.frame.update_idletasks()
        except Exception as e:
            print(f"상태 업데이트 오류: {e}")
    
    def update_progress_callback(self, bank_name, status):
        """스크래핑 진행 상태 콜백"""
        self.update_bank_status(bank_name, status)
    
    def start_scraping(self):
        """스크래핑을 시작합니다."""
        if self.running:
            messagebox.showwarning("경고", "이미 스크래핑이 실행 중입니다.")
            return
        
        # 설정 업데이트
        try:
            self.config.update_output_dir(self.output_dir_var.get())
        except Exception as e:
            messagebox.showerror("오류", f"출력 디렉토리 설정 실패: {str(e)}")
            return
            
        self.config.MAX_WORKERS = self.workers_var.get()
        
        chrome_driver_path = self.chrome_driver_path_var.get()
        if chrome_driver_path:
            if os.path.exists(chrome_driver_path):
                self.config.update_chrome_driver_path(chrome_driver_path)
            else:
                messagebox.showwarning("경고", f"지정한 ChromeDriver 경로가 존재하지 않습니다: {chrome_driver_path}\n자동으로 찾습니다.")
                self.config.update_chrome_driver_path(None)
        else:
            self.config.update_chrome_driver_path(None)
        
        # 선택된 은행 확인
        selected_banks = list(self.bank_tree.selection())
        if not selected_banks:
            messagebox.showwarning("경고", "스크래핑할 은행을 선택하세요.")
            return
        
        # UI 업데이트
        self.start_button.config(state=tk.DISABLED)
        self.stop_button.config(state=tk.NORMAL)
        self.running = True
        
        # 스크래핑 스레드 시작
        self.scraping_thread = threading.Thread(target=self.run_scraping, args=(selected_banks,))
        self.scraping_thread.daemon = True
        self.scraping_thread.start()
    
    def run_scraping(self, selected_banks):
        """별도 스레드에서 스크래핑을 실행합니다."""
        try:
            # 드라이버 및 진행 관리자 초기화
            self.driver_manager = DriverManager(self.config, self.logger)
            self.progress_manager = ProgressManager(self.config, self.logger)
            self.scraper = BankScraper(self.config, self.logger, self.driver_manager, self.progress_manager)
            
            # 드라이버 초기화
            self.driver_manager.initialize_drivers()
            
            # 스크래핑 실행
            self.logger.log_message(f"\n===== 저축은행 중앙회 결산공시 데이터 스크래핑 시작 [{self.config.today}] =====\n")
            self.logger.log_message(f"선택된 은행 수: {len(selected_banks)}")
            
            # 스크래핑 시작 전 상태 초기화
            for bank in selected_banks:
                self.update_bank_status(bank, "대기 중")
            
            # 스크래핑 실행 (MD 옵션 포함)
            start_time = time.time()
            save_md = self.save_md_var.get()  # MD 저장 옵션 확인
            results = self.scraper.process_banks(selected_banks, self.update_progress_callback, save_md)
            
            # 결과 처리
            successful_banks = [r[0] for r in results if r[1]]
            failed_banks = [r[0] for r in results if not r[1]]
            
            # 실행 시간 계산
            end_time = time.time()
            total_duration = end_time - start_time
            minutes, seconds = divmod(total_duration, 60)
            
            # 요약 메시지
            self.logger.log_message(f"\n스크래핑 완료!")
            self.logger.log_message(f"성공: {len(successful_banks)}개, 실패: {len(failed_banks)}개")
            self.logger.log_message(f"총 실행 시간: {int(minutes)}분 {int(seconds)}초")
            
            # 요약 보고서 생성 (MD 포함)
            summary_file, stats, _ = self.scraper.generate_summary_report()
            if save_md:
                md_summary_file = self.scraper.generate_summary_report_md()
                if md_summary_file:
                    self.logger.log_message(f"MD 요약 보고서 생성 완료: {md_summary_file}")
            
            # UI 업데이트 - after 메서드로 메인 스레드에서 실행
            self.frame.after(0, self.on_scraping_complete)
            
        except Exception as e:
            self.logger.log_message(f"스크래핑 중 오류 발생: {str(e)}")
            import traceback
            self.logger.log_message(traceback.format_exc())
            
            # UI 업데이트
            self.frame.after(0, self.on_scraping_error)
        finally:
            # 드라이버 종료
            if self.driver_manager:
                self.driver_manager.close_all()
    
    def on_scraping_complete(self):
        """스크래핑 완료 후 UI 업데이트"""
        self.running = False
        self.start_button.config(state=tk.NORMAL)
        self.stop_button.config(state=tk.DISABLED)
        messagebox.showinfo("완료", "저축은행 결산공시 데이터 스크래핑이 완료되었습니다.")
        
        # 자동 압축 옵션이 활성화된 경우
        if self.auto_zip_var.get():
            self.frame.after(500, self.compress_and_download)  # 약간 지연 후 압축 시작
    
    def on_scraping_error(self):
        """스크래핑 오류 발생 시 UI 업데이트"""
        self.running = False
        self.start_button.config(state=tk.NORMAL)
        self.stop_button.config(state=tk.DISABLED)
        messagebox.showerror("오류", "스크래핑 중 오류가 발생했습니다. 로그를 확인하세요.")
    
    def stop_scraping(self):
        """스크래핑을 중지합니다."""
        if not self.running:
            return
        
        if messagebox.askyesno("중지 확인", "스크래핑을 중지하시겠습니까? 현재 진행 중인 작업이 완료된 후 중지됩니다."):
            self.running = False
            
            # 드라이버 종료
            if self.driver_manager:
                self.driver_manager.close_all()
            
            self.start_button.config(state=tk.NORMAL)
            self.stop_button.config(state=tk.DISABLED)
    
    def open_output_folder(self):
        """출력 폴더를 엽니다."""
        output_dir = self.output_dir_var.get()
        if os.path.exists(output_dir):
            try:
                if sys.platform == 'win32':
                    os.startfile(output_dir)
                elif sys.platform == 'darwin':  # macOS
                    import subprocess
                    subprocess.Popen(['open', output_dir])
                else:  # Linux
                    import subprocess
                    subprocess.Popen(['xdg-open', output_dir])
            except Exception as e:
                messagebox.showerror("오류", f"폴더를 열 수 없습니다: {str(e)}")
        else:
            messagebox.showerror("오류", f"폴더가 존재하지 않습니다: {output_dir}")
    
    def generate_report(self):
        """요약 보고서를 생성합니다."""
        try:
            # 스크래퍼 초기화
            if self.scraper is None:
                self.progress_manager = ProgressManager(self.config, self.logger)
                self.scraper = BankScraper(self.config, self.logger, None, self.progress_manager)
            
            # 보고서 생성
            summary_file, stats, summary_df = self.scraper.generate_summary_report()
            
            if summary_file and os.path.exists(summary_file):
                messagebox.showinfo("완료", f"요약 보고서가 생성되었습니다: {summary_file}")
                
                # 요약 창 표시
                self.show_summary_window(stats, summary_df)
            else:
                messagebox.showerror("오류", "요약 보고서 생성에 실패했습니다.")
        
        except Exception as e:
            messagebox.showerror("오류", f"요약 보고서 생성 중 오류 발생: {str(e)}")
    
    def generate_md_summary_report(self):
        """MD 요약 보고서를 생성합니다."""
        try:
            if self.scraper is None:
                self.progress_manager = ProgressManager(self.config, self.logger)
                self.scraper = BankScraper(self.config, self.logger, None, self.progress_manager)
            
            md_summary_file = self.scraper.generate_summary_report_md()
            
            if md_summary_file and os.path.exists(md_summary_file):
                messagebox.showinfo("완료", f"📝 결산공시 MD 요약 보고서가 생성되었습니다!\n\n{os.path.basename(md_summary_file)}")
                
                if messagebox.askyesno("파일 열기", "생성된 MD 파일을 열어보시겠습니까?"):
                    self.open_md_file(md_summary_file)
            else:
                messagebox.showerror("오류", "MD 요약 보고서 생성에 실패했습니다.")
        
        except Exception as e:
            messagebox.showerror("오류", f"MD 요약 보고서 생성 중 오류 발생: {str(e)}")

    def open_md_file(self, file_path):
        """마크다운 파일을 엽니다."""
        if os.path.exists(file_path):
            try:
                if sys.platform == 'win32':
                    os.startfile(file_path)
                elif sys.platform == 'darwin':  # macOS
                    import subprocess
                    subprocess.Popen(['open', file_path])
                else:  # Linux
                    import subprocess
                    subprocess.Popen(['xdg-open', file_path])
            except Exception as e:
                messagebox.showerror("오류", f"파일을 열 수 없습니다: {str(e)}")
        else:
            messagebox.showerror("오류", f"파일이 존재하지 않습니다: {file_path}")
    
    def show_summary_window(self, stats, summary_df):
        """요약 통계를 보여주는 창을 표시합니다."""
        if summary_df is None:
            return
        
        summary_window = tk.Toplevel(self.parent)
        summary_window.title("스크래핑 결과 요약")
        summary_window.geometry("600x400")
        
        # 통계 프레임
        stats_frame = ttk.LabelFrame(summary_window, text="통계")
        stats_frame.pack(fill=tk.X, padx=10, pady=10)
        
        row = 0
        for key, value in stats.items():
            ttk.Label(stats_frame, text=key).grid(row=row, column=0, sticky=tk.W, padx=5, pady=2)
            ttk.Label(stats_frame, text=str(value)).grid(row=row, column=1, sticky=tk.W, padx=5, pady=2)
            row += 1
        
        # 은행별 상태 프레임
        status_frame = ttk.LabelFrame(summary_window, text="은행별 상태")
        status_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 트리뷰
        tree = ttk.Treeview(status_frame, columns=list(summary_df.columns), show="headings")
        for col in summary_df.columns:
            tree.heading(col, text=col)
            tree.column(col, width=100)
        
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # 스크롤바
        scrollbar = ttk.Scrollbar(status_frame, orient=tk.VERTICAL, command=tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        tree.configure(yscrollcommand=scrollbar.set)
        
        # 데이터 추가
        for _, row in summary_df.iterrows():
            values = list(row.values)
            tree.insert("", tk.END, values=values)
        
        # 닫기 버튼
        ttk.Button(summary_window, text="닫기", command=summary_window.destroy).pack(pady=10)
    
    def create_financial_consolidation_with_selection(self):
        """재무 데이터 통합 보고서를 생성합니다. (폴더/파일 선택 가능) - v2.5"""
        try:
            # 스크래퍼 초기화
            if self.scraper is None:
                self.progress_manager = ProgressManager(self.config, self.logger)
                self.scraper = BankScraper(self.config, self.logger, None, self.progress_manager)
            
            # 데이터 소스 선택 대화상자 표시
            source_dialog = FinancialDataSourceDialog(self.parent, self.config)
            self.parent.wait_window(source_dialog.dialog)
            
            if source_dialog.result:
                source_type, source_path = source_dialog.result
                
                # 상태 업데이트
                self.update_log(f"📊 당기/전년동기 비교분석 재무 데이터 통합 시작: {source_type} - {source_path}")
                
                # 별도 스레드에서 실행
                threading.Thread(
                    target=self._run_financial_consolidation_flexible, 
                    args=(source_type, source_path),
                    daemon=True
                ).start()
            
        except Exception as e:
            messagebox.showerror("오류", f"재무 데이터 통합 중 오류 발생: {str(e)}")
    
    def _run_financial_consolidation_flexible(self, source_type, source_path):
        """유연한 재무 데이터 통합 실행 - v2.5 확장"""
        try:
            # 소스 타입에 따른 처리
            if source_type == "default_folder":
                # 기존 방식: 기본 출력 폴더에서 모든 파일 읽기
                output_file, consolidated_df = self.scraper.create_consolidated_financial_report()
                
            elif source_type == "custom_folder":
                # 사용자 지정 폴더에서 파일 읽기
                output_file, consolidated_df = self.scraper.create_consolidated_financial_report_from_folder(source_path)
                
            elif source_type == "selected_files":
                # 선택한 파일들만 처리
                output_file, consolidated_df = self.scraper.create_consolidated_financial_report_from_files(source_path)
            
            if output_file and os.path.exists(output_file):
                # 성공 메시지
                self.frame.after(0, lambda: self.update_log(f"📊 당기/전년동기 비교분석 완료: {output_file}"))
                self.frame.after(0, lambda: messagebox.showinfo("완료", f"📊 당기/전년동기 비교분석 재무 데이터 통합 보고서가 생성되었습니다!\n\n✅ 4개 시트로 구성된 상세 분석\n📈 당기 vs 전년동기 증감 분석\n📋 MD 비교분석 보고서 포함\n\n{os.path.basename(output_file)}"))
                
                # 통합 결과 창 표시 (v2.5 확장된 버전)
                self.frame.after(0, lambda: self.show_financial_consolidation_window_v25(consolidated_df))
            else:
                self.frame.after(0, lambda: messagebox.showerror("오류", "재무 데이터 통합 보고서 생성에 실패했습니다."))
                
        except Exception as e:
            self.frame.after(0, lambda: messagebox.showerror("오류", f"재무 데이터 통합 중 오류 발생: {str(e)}"))
    
    def show_financial_consolidation_window_v25(self, consolidated_df):
        """재무 데이터 통합 결과를 보여주는 창을 표시합니다. (v2.5 - 당기/전년동기 비교)"""
        if consolidated_df is None:
            return
        
        # 새 창 생성
        consolidation_window = tk.Toplevel(self.parent)
        consolidation_window.title("📊 당기/전년동기 재무 데이터 비교분석 결과")
        consolidation_window.geometry("1400x700")
        
        # 메인 프레임
        main_frame = ttk.Frame(consolidation_window, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 탭 노트북 생성 (3개 탭)
        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill=tk.BOTH, expand=True)
        
        # 탭 1: 요약 보기
        summary_frame = ttk.Frame(notebook)
        notebook.add(summary_frame, text="📋 요약 보기")
        
        # 탭 2: 상세 데이터
        detail_frame = ttk.Frame(notebook)
        notebook.add(detail_frame, text="📊 상세 데이터")
        
        # 탭 3: 증감 분석
        analysis_frame = ttk.Frame(notebook)
        notebook.add(analysis_frame, text="📈 증감 분석")
        
        # === 탭 1: 요약 보기 ===
        self._create_summary_tab(summary_frame, consolidated_df)
        
        # === 탭 2: 상세 데이터 ===
        self._create_detail_tab(detail_frame, consolidated_df)
        
        # === 탭 3: 증감 분석 ===
        self._create_analysis_tab(analysis_frame, consolidated_df)
        
        # 하단 버튼 프레임
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(button_frame, text="📊 엑셀 파일 열기", 
                  command=lambda: self.open_excel_file(os.path.join(self.config.output_dir, 
                                                                   f'저축은행_결산_재무데이터_통합_비교분석_{self.config.today}.xlsx'))).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(button_frame, text="📝 MD 보고서 열기", 
                  command=lambda: self.open_md_file(os.path.join(self.config.output_dir, 
                                                                f'저축은행_결산_재무데이터_통합_비교분석_{self.config.today}.md'))).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(button_frame, text="닫기", command=consolidation_window.destroy).pack(side=tk.RIGHT, padx=5)
    
    def _create_summary_tab(self, parent_frame, consolidated_df):
        """요약 탭을 생성합니다."""
        # 요약 통계 프레임
        stats_frame = ttk.LabelFrame(parent_frame, text="📊 전체 요약 통계")
        stats_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # 통계 계산 및 표시
        total_banks = len(consolidated_df)
        banks_with_data = len(consolidated_df[consolidated_df['재무정보 날짜'] != '데이터 없음'])
        
        stats_text = ttk.Frame(stats_frame)
        stats_text.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(stats_text, text=f"🏦 전체 은행 수: {total_banks}개", font=("", 10)).pack(anchor=tk.W)
        ttk.Label(stats_text, text=f"📈 데이터 보유 은행: {banks_with_data}개", font=("", 10)).pack(anchor=tk.W)
        
        # 평균 증감률 표시
        if '증감률_총자산' in consolidated_df.columns:
            avg_asset_growth = consolidated_df['증감률_총자산'].mean()
            if not pd.isna(avg_asset_growth):
                growth_emoji = "📈" if avg_asset_growth > 0 else "📉" if avg_asset_growth < 0 else "➡️"
                ttk.Label(stats_text, text=f"{growth_emoji} 평균 총자산 증감률: {avg_asset_growth:.2f}%", font=("", 10)).pack(anchor=tk.W)
        
        if '증감률_자기자본' in consolidated_df.columns:
            avg_capital_growth = consolidated_df['증감률_자기자본'].mean()
            if not pd.isna(avg_capital_growth):
                growth_emoji = "💪" if avg_capital_growth > 0 else "⚠️" if avg_capital_growth < 0 else "➡️"
                ttk.Label(stats_text, text=f"{growth_emoji} 평균 자기자본 증감률: {avg_capital_growth:.2f}%", font=("", 10)).pack(anchor=tk.W)
        
        # 상위 성과 은행 표시
        top_frame = ttk.LabelFrame(parent_frame, text="🏆 상위 성과 은행")
        top_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 총자산 증가율 상위 5개
        if '증감률_총자산' in consolidated_df.columns:
            asset_growth_data = consolidated_df[consolidated_df['증감률_총자산'].notna()]
            if not asset_growth_data.empty:
                top_asset_growth = asset_growth_data.nlargest(5, '증감률_총자산')
                
                ttk.Label(top_frame, text="📈 총자산 증가율 상위 5개:", font=("", 10, "bold")).pack(anchor=tk.W, pady=5)
                
                for i, (_, row) in enumerate(top_asset_growth.iterrows()):
                    rate = row['증감률_총자산']
                    trend = "🔥" if rate > 10 else "📈"
                    ttk.Label(top_frame, text=f"  {i+1}. {row['은행명']}: {trend} {rate:+.2f}%").pack(anchor=tk.W)
    
    def _create_detail_tab(self, parent_frame, consolidated_df):
        """상세 데이터 탭을 생성합니다."""
        # 트리뷰 생성
        tree_frame = ttk.Frame(parent_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 주요 컬럼만 선택해서 표시 (너무 많으면 복잡함)
        display_columns = ['은행명', '당기_총자산', '전년동기_총자산', '증감률_총자산', 
                          '당기_자기자본', '전년동기_자기자본', '증감률_자기자본']
        
        available_columns = [col for col in display_columns if col in consolidated_df.columns]
        display_df = consolidated_df[available_columns]
        
        # 트리뷰 생성
        columns = list(display_df.columns)
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings")
        
        # 컬럼 헤더 설정
        for col in columns:
            if col == '은행명':
                tree.heading(col, text='은행명')
                tree.column(col, width=100, anchor=tk.CENTER)
            elif '당기_' in col:
                tree.heading(col, text=col.replace('당기_', '당기 '))
                tree.column(col, width=120, anchor=tk.E)
            elif '전년동기_' in col:
                tree.heading(col, text=col.replace('전년동기_', '전년 '))
                tree.column(col, width=120, anchor=tk.E)
            elif '증감률_' in col:
                tree.heading(col, text=col.replace('증감률_', '증감률 '))
                tree.column(col, width=100, anchor=tk.E)
        
        # 스크롤바
        v_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # 배치
        tree.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # 데이터 추가
        for _, row in display_df.iterrows():
            values = []
            for col, value in zip(display_df.columns, row):
                if pd.isna(value):
                    values.append('')
                elif '증감률_' in col:
                    try:
                        rate_val = float(value)
                        values.append(f"{rate_val:+.2f}%")
                    except:
                        values.append(str(value))
                elif col in ['당기_총자산', '전년동기_총자산', '당기_자기자본', '전년동기_자기자본']:
                    try:
                        values.append(f"{int(value):,}")
                    except:
                        values.append(str(value))
                else:
                    values.append(str(value))
            
            tree.insert("", tk.END, values=values)
    
    def _create_analysis_tab(self, parent_frame, consolidated_df):
        """증감 분석 탭을 생성합니다."""
        # 분석 결과 프레임
        analysis_frame = ttk.LabelFrame(parent_frame, text="📈 증감 분석 결과")
        analysis_frame.pack(fill=tk.X, padx=10, pady=10)
        
        analysis_content = ttk.Frame(analysis_frame)
        analysis_content.pack(fill=tk.X, padx=10, pady=10)
        
        # 성장/감소 은행 수 계산
        if '증감률_총자산' in consolidated_df.columns:
            growth_banks = len(consolidated_df[(consolidated_df['증감률_총자산'] > 0) & consolidated_df['증감률_총자산'].notna()])
            decline_banks = len(consolidated_df[(consolidated_df['증감률_총자산'] < 0) & consolidated_df['증감률_총자산'].notna()])
            
            ttk.Label(analysis_content, text=f"📈 성장 은행 (총자산 증가): {growth_banks}개", font=("", 10)).pack(anchor=tk.W)
            ttk.Label(analysis_content, text=f"📉 감소 은행 (총자산 감소): {decline_banks}개", font=("", 10)).pack(anchor=tk.W)
            
            # 시장 전망
            if growth_banks > decline_banks:
                ttk.Label(analysis_content, text="🟢 시장 전망: 긍정적 (성장 은행이 더 많음)", 
                         font=("", 10, "bold"), foreground="green").pack(anchor=tk.W, pady=5)
            elif decline_banks > growth_banks:
                ttk.Label(analysis_content, text="🔴 시장 전망: 주의 필요 (감소 은행이 더 많음)", 
                         font=("", 10, "bold"), foreground="red").pack(anchor=tk.W, pady=5)
            else:
                ttk.Label(analysis_content, text="🟡 시장 전망: 혼재 상황", 
                         font=("", 10, "bold"), foreground="orange").pack(anchor=tk.W, pady=5)
        
        # 증감률 차트 (간단한 히스토그램 정보)
        chart_frame = ttk.LabelFrame(parent_frame, text="📊 증감률 분포")
        chart_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        if '증감률_총자산' in consolidated_df.columns:
            rates = consolidated_df['증감률_총자산'].dropna()
            if not rates.empty:
                # 간단한 텍스트 기반 분포 표시
                ranges = [
                    ("📉 -20% 이하", len(rates[rates <= -20])),
                    ("📉 -20% ~ -10%", len(rates[(rates > -20) & (rates <= -10)])),
                    ("📉 -10% ~ 0%", len(rates[(rates > -10) & (rates <= 0)])),
                    ("📈 0% ~ 10%", len(rates[(rates > 0) & (rates <= 10)])),
                    ("📈 10% ~ 20%", len(rates[(rates > 10) & (rates <= 20)])),
                    ("🔥 20% 이상", len(rates[rates > 20]))
                ]
                
                for range_text, count in ranges:
                    if count > 0:
                        ttk.Label(chart_frame, text=f"{range_text}: {count}개 은행").pack(anchor=tk.W, pady=2)
    
    def open_excel_file(self, file_path):
        """엑셀 파일을 엽니다."""
        if os.path.exists(file_path):
            try:
                if sys.platform == 'win32':
                    os.startfile(file_path)
                elif sys.platform == 'darwin':  # macOS
                    import subprocess
                    subprocess.Popen(['open', file_path])
                else:  # Linux
                    import subprocess
                    subprocess.Popen(['xdg-open', file_path])
            except Exception as e:
                messagebox.showerror("오류", f"파일을 열 수 없습니다: {str(e)}")
        else:
            messagebox.showerror("오류", f"파일이 존재하지 않습니다: {file_path}")
    
    def compress_and_download(self):
        """결과를 압축 파일로 만들고 원하는 위치에 저장합니다."""
        try:
            # 스크래퍼 초기화
            if self.scraper is None:
                self.progress_manager = ProgressManager(self.config, self.logger)
                self.scraper = BankScraper(self.config, self.logger, None, self.progress_manager)
            
            # 기본 저장 경로 및 파일명 설정
            default_filename = f'저축은행_결산공시_데이터_{self.config.today}.zip'
            default_dir = os.path.dirname(self.config.output_dir)
            
            # 파일 저장 대화상자
            save_path = filedialog.asksaveasfilename(
                title="압축 파일 저장",
                initialdir=default_dir,
                initialfile=default_filename,
                defaultextension=".zip",
                filetypes=[("ZIP 파일", "*.zip"), ("모든 파일", "*.*")]
            )
            
            if not save_path:
                return  # 사용자가 취소함
            
            # 압축 작업은 별도 스레드에서 실행 (UI 응답성 유지)
            threading.Thread(target=self._create_zip_file, args=(save_path,), daemon=True).start()
            
        except Exception as e:
            messagebox.showerror("오류", f"압축 파일 생성 중 오류 발생: {str(e)}")
    
    def _create_zip_file(self, save_path):
        """별도 스레드에서 실행되는 압축 파일 생성 함수"""
        try:
            # 압축 작업 수행
            with zipfile.ZipFile(save_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                # 진행 상황 모니터링 변수
                total_files = 0
                for root, _, files in os.walk(self.config.output_dir):
                    total_files += len(files)
                
                files_processed = 0
                
                # 파일 압축
                for root, dirs, files in os.walk(self.config.output_dir):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arcname = os.path.relpath(file_path, os.path.dirname(self.config.output_dir))
                        zipf.write(file_path, arcname)
                        
                        # 진행 상황 업데이트
                        files_processed += 1
                        progress = int(files_processed / total_files * 100)
                        self.frame.after(0, lambda p=progress: self.update_log(f"압축 중... {p}%"))
            
            # 완료 메시지
            self.frame.after(0, lambda: self.logger.log_message(f"압축 파일 생성 완료: {save_path}"))
            self.frame.after(0, lambda: messagebox.showinfo("완료", f"압축 파일이 생성되었습니다:\n{save_path}"))
            
            # 파일 탐색기에서 압축 파일 열기 (선택적)
            self.frame.after(0, lambda: self._open_file_location(save_path))
            
        except Exception as e:
            self.frame.after(0, lambda: self.logger.log_message(f"압축 파일 생성 오류: {str(e)}"))
            self.frame.after(0, lambda: messagebox.showerror("오류", f"압축 파일 생성 중 오류 발생: {str(e)}"))
    
    def _open_file_location(self, file_path):
        """파일이 위치한 폴더를 엽니다."""
        try:
            directory = os.path.dirname(file_path)
            if sys.platform == 'win32':
                os.startfile(directory)
            elif sys.platform == 'darwin':  # macOS
                import subprocess
                subprocess.Popen(['open', directory])
            else:  # Linux
                import subprocess
                subprocess.Popen(['xdg-open', directory])
        except Exception as e:
            self.logger.log_message(f"폴더 열기 실패: {str(e)}", verbose=False)
    
    def reset_progress(self):
        """진행 상태를 초기화합니다."""
        if messagebox.askyesno("초기화 확인", "진행 상태를 초기화하시겠습니까? 이 작업은 되돌릴 수 없습니다."):
            try:
                # 진행 관리자 초기화
                self.progress_manager = ProgressManager(self.config, self.logger)
                self.progress_manager.reset_progress()
                
                # 상태 업데이트
                for bank in self.config.BANKS:
                    self.update_bank_status(bank, "대기 중")
                
                messagebox.showinfo("완료", "진행 상태가 초기화되었습니다.")
                self.logger.log_message("진행 상태 초기화 완료")
            
            except Exception as e:
                messagebox.showerror("오류", f"진행 상태 초기화 중 오류 발생: {str(e)}")


# main 함수와 실행 부분은 주석 처리 (탭 버전에서는 불필요)
# def main():
#     """프로그램의 메인 진입점"""
#     try:
#         # Tkinter 루트 윈도우 생성
#         root = tk.Tk()
#         
#         # 앱 인스턴스 생성
#         app = SettlementScraperTab(root)
#         
#         # 이벤트 루프 시작
#         root.mainloop()
#         
#     except Exception as e:
#         print(f"프로그램 실행 중 오류 발생: {str(e)}")
#         import traceback
#         traceback.print_exc()


# 프로그램 진입점 주석 처리 (탭 버전에서는 불필요)
# if __name__ == "__main__":
#     main()
